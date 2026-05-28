# -*- coding: utf-8 -*-
"""
HESTIA_Data_Engine_CVR.py
=========================
Heat Exposure Stress & Thermophysiology Integrated Assessment -- Data Engine
with Cardiovascular Response (CVR) integration.

This script combines two specialised simulation models:
  1. Adults  -- JOS-3 thermoregulation + CVR module + collapse risk model.
  2. Children / adolescents -- JOS-3 thermoregulation (no CVR).

The user selects the simulation mode at startup; the script then automatically
adjusts physiological parameters, activity levels (MET), and perception models.

=============================================================================
COMMON FEATURES
=============================================================================
- JOS-3 thermoregulation model (Takahashi et al. 2021, Energy & Buildings).
- Monte Carlo simulation for population-based risk analysis (default N=5000).
- Environmental modelling: Open-Meteo weather API, pvlib solar radiation,
  Urban Heat Island effect (Oke 1973).
- Vanos et al. (2023) survivability / liveability limits.

=============================================================================
ADULT MODE
=============================================================================
- Borg RPE scale (6-20), piecewise-linear psychobiological model.
- Adaptation profiles: training and heat acclimatisation affect running
  economy, sweat response, and thirst threshold.
- Ad-libitum drinking strategy.
- CVR module (Lloyd et al. 2022, J Appl Physiol):
  cardiac output, heart rate (CO/SV), CO reserve, decompensation flag.
- Collapse risk model: two-phase logistic function calibrated on observed
    Boston Marathon EHS incidence (Breslow et al. 2021, Am J Sports Med).
    DtD 2024 (50 / 35,000 participants) serves as prospective validation.

=============================================================================
CHILDREN MODE
=============================================================================
- Higher surface-to-mass ratio, lower sweat capacity, higher metabolic cost.
- OMNI RPE scale (0-10).  Blunted thirst response ("voluntary dehydration").
- Simplified fitness factor.

=============================================================================
VERSION HISTORY
=============================================================================
rev01  2025-10-28  Initial combined adult + pediatric version.
rev02  2025-10-28  Advanced RPE models. Full English translation.
rev03  2026-03     CVR module integrated (Lloyd et al. 2022).
                   cardiac_output from JOS-3 used directly as CO demand.
                   HR estimated via CO / SV (Lloyd Eq. 2).
rev04  2026-03     AdultParticipantProfile dataclass replaces anonymous tuple.
                   Fixes critical bug: vo2max and age were swapped in
                   calculate_indices_jos3_adult() due to index confusion.
                   All participant fields now accessed by name, not position.
rev05  2026-03     AVA closure indicator removed. At exercise intensity in
                   warm conditions AVA blood flow is elevated (open state),
                   never closes -- indicator had no diagnostic value.
rev06  2026-03     Collapse risk model added (two-phase logistic function):
                     Z = intercept
                         + 1.0 * max(0, T_rect - 39.5)
                         + 4.0 * max(0, T_rect - 40.5)
                         + 0.8 * max(0, 2.0 - CO_reserve)
                         + 0.5 * max(0, dehy_pct - 3.0)
                   Intercept auto-calibrated via Newton iteration to reproduce
                   observed Dam tot Damloop 2024 hospital admission rate
                   (50/35,000). [Superseded by rev11: Boston Marathon pooled
                   2012-2019 (Breslow 2021) is now the calibration reference;
                   DtD 2024 is a prospective validation target.]
                   Calibration constants moved to module level.
                   Individual collapse probability added to Excel export.
rev07  2026-03     Debug print statements removed.
                   Stale 40.0 deg C comment block removed.
                   All Dutch inline comments and docstrings translated to
                   English. simulate_runner and calculate_indices_jos3_adult
                   docstrings updated to reflect AdultParticipantProfile.
                   _calculate_threshold_stats: Wald CI documented as known
                   limitation; Clopper-Pearson recommended for tail props.
rev08  2026-03     Line 1920-1930-Commented
                   Line 1935  intercept_kal = -9.637 This represents the
                   calibration value of the 2024 DtD run 50 incidents per
                   35000 participants.
                   [Superseded by rev11: Boston Marathon (Breslow 2021)
                   is now the calibration reference; DtD 2024 is a
                   prospective validation target; intercept_kal = -10.071.]
rev09  2026-04     Monte Carlo distribution corrections (simulate_runner,
                   simulate_child_participant, run_monte_carlo_adult):
                   1. MET loc corrected to 0.0 (was -training_factor*met*0.15).
                      Training effect now applied as separate deterministic
                      mean shift, eliminating systematic negative bias
                      (empirical mean was -0.75 MET instead of 0.0).
                   2. MET sigma widened: 0.05 -> 0.15 * met_value.
                      Captures running-economy CV (~5-8%, Barnes & Kilding
                      2015) plus pace spread across population (~7-10%).
                   3. MET upper clip added: final_met capped at
                      1.25 * met_value (symmetric with lower bound).
                   4. VO2max loc corrected 45 -> 50 mL/kg/min for
                      recreational runners self-selecting into mass events.
                      Ref: Scharhag-Rosenberger et al. (2010).
                   5. sweat_factor (adults) clipped to [0.4, 1.5].
                   6. Age sigma widened: 10 -> 12 years for better
                      representation of 55-65 age group.
                   7. Height loc corrected: 1.75 -> 1.71 m for 50/50
                      gender-mixed runner population (M~1.78, V~1.65).
                      Ref: RIVM NL data; Tolfrey & Pilcher (2013).
                   8. Weight scale widened: 10 -> 12 kg; lower clip
                      50 -> 45 kg (accommodates lighter female runners).
                   Known limitation: rejection sampling introduces <0.3
                   MET bias at training_factor=0.5 (bias +0.6 at tf=1.0);
                   scipy.truncnorm recommended for future revision.

rev12  2026-04     THREE-ENDPOINT CALIBRATION ARCHITECTURE (this revision)
                   See detailed entry below.

rev11  2026-04     BOSTON MARATHON REFERENCE CONSOLIDATION (this revision)
                   See detailed entry below.

rev10  2026-04     POPULATION ARCHITECTURE OVERHAUL
                   =================================================
                   Replaces simulate_runner() with generate_base_population(),
                   which implements a causally correct sampling order and
                   eliminates physiologically impossible parameter combinations.

                   KEY CHANGES:
                   A) Sampling order is now causally correct:
                      gender -> age -> VO2max(age,gender) -> height(gender)
                      -> weight(height,gender) -> %VO2max_at_race_pace
                      -> sweat_factor -> thirst_threshold -> perceptual vars
                      This prevents impossible combinations such as a 65-year-
                      old with VO2max=75 or a 105kg runner with VO2max=70.

                   B) MET is no longer sampled as an absolute value.
                      Instead, pct_vo2max (dimensionless, [0.60, 0.92]) is
                      sampled and stored in the population DataFrame.
                      The actual MET for each scenario is computed on-the-fly
                      in calculate_indices_jos3_adult() as:
                        final_met = vo2max * pct_vo2max / VO2MAX_TO_MET_FACTOR
                      This makes it impossible to exceed 100% VO2max by
                      construction (was ~18.7% of population in rev09).

                   C) VO2max is conditioned on age and gender:
                      - Gender effect: females ~10% lower (Cohen's d ~0.8).
                        Ref: Scharhag-Rosenberger et al. (2010).
                      - Age effect: -0.5 mL/kg/min per year after age 35.
                        Ref: Tanaka et al. (2001) J Am Coll Cardiol 37:153.
                      This replaces the flat N(50,10) distribution.

                   D) Fixed seed support: generate_base_population() accepts
                      a random_seed parameter. Using the same seed always
                      produces the same population, making scenario comparisons
                      (e.g. MET 8.3 vs 11.8, DtD 2024 vs 2025) fully
                      reproducible and scientifically valid.

                   E) Population persistence: the generated population can be
                      saved to / loaded from a .pkl file, so it need not be
                      regenerated for each simulation run. This eliminates the
                      rejection-sampling bias noted in rev09 and reduces
                      run-to-run stochastic variance in scenario comparisons.

                   F) NEW: validate_population() function checks:
                      1. Marginal distributions vs. empirical references
                         (Scharhag-Rosenberger 2010, Tanda & Knechtle 2013,
                          RunRepeat European marathon statistics 2023).
                      2. Correlation matrix vs. physiologically expected values
                         (VO2max-age: r~-0.40; VO2max-gender: Cohen's d~0.80).
                      3. Implied finish-time distribution vs. Amsterdam/
                         Rotterdam marathon results using Daniels & Gilbert
                         (1979) VO2max-performance formula. If the simulated
                         median finish time falls outside [3h30m, 4h30m] for
                         a recreational marathon field, a warning is issued.

                   G) AdultParticipantProfile dataclass extended:
                      - pct_vo2max field added (replaces met_variation).
                      - met_variation field REMOVED to prevent accidental use.
                      NOTE: met_variation in the stored DataFrame column is
                      retained as a computed diagnostic only (never sampled).

                   BACKWARD COMPATIBILITY:
                   - run_monte_carlo_adult() now accepts an optional
                     base_population argument (DataFrame from
                     generate_base_population()). If None, a fresh population
                     is generated. Existing call sites are unchanged.
                   - All downstream code (CVR, collapse risk, Excel export)
                     is unchanged; only the population generation pathway
                     and the MET computation in calculate_indices_jos3_adult
                     are affected.

                   KNOWN REMAINING ISSUES (addressed in rev11):
                   - training_offset duplication: see rev11 notes below.
                   - intercept_kal calibration source: see rev11 notes below.
                   - scipy.truncnorm not yet used for pct_vo2max; simple
                     np.clip is applied. Tail behavior acceptable for
                     recreational marathon population (deferred to rev12).

=============================================================================
rev11  2026-04     BOSTON MARATHON REFERENCE CONSOLIDATION
=============================================================================
                   This revision makes the Boston Marathon the single
                   authoritative calibration reference for the collapse risk
                   model, resolves the training_factor double-correction
                   ambiguity, cleans up BMI filter logic, and adds population
                   file version-safety.

                   A) CALIBRATION SOURCE -- Boston Marathon (Breslow 2021)
                   ---------------------------------------------------------
                   Previous revisions mixed DtD 2024 and Boston 2017 data
                   in different parts of the code, creating inconsistency.
                   Rev11 consolidates on the Boston Marathon as the primary
                   reference because it is the best-documented mass running
                   event in the peer-reviewed literature for EHS incidence.

                   Reference: Breslow RG et al. (2021) Am J Sports Med 49(10).
                   Dataset: Boston Marathon 2012--2019 (8 editions).
                   Observed EHS rate: approximately 9.0 per 10,000 finishers
                   (pooled across 8 editions, varying by year and temperature).
                   Conservative reference year chosen: 2017 (cool conditions,
                   low incidence) gives a lower-bound intercept; 2012 (warm
                   conditions) gives an upper bound.

                   Calibration anchor used here (rev11):
                     N_OBS_COLLAPSE_ADMISSIONS = 9.0   (per 10,000; pooled)
                     N_OBS_PARTICIPANTS        = 10_000
                     P_OBS_COLLAPSE            = 9.0 / 10_000 = 0.00090

                   intercept_kal recalibrated to: -10.071
                   (Newton iteration, see MODULE_CALIBRATION_NOTES below)

                   All DtD 2024 references in module-level comments have been
                   updated. DtD 2024 (50 / 35,000 = 0.00143) will serve as a
                   prospective validation point once the training_offset
                   calibration is settled (Bongers data needed).

                   B) TRAINING_FACTOR DOUBLE-CORRECTION -- RESOLVED
                   -------------------------------------------------
                   Rev10 applied training_factor in two places:
                     1. generate_base_population(): mu of pct_vo2max shifted
                        down by training_factor * 0.08 (population sampling).
                     2. calculate_indices_jos3_adult(): MET further reduced by
                        training_factor * base_met * 0.15 (economy offset).

                   This caused a compounded reduction in effective %VO2max for
                   trained runners. Quantified effect at tf=1.0:
                     Step 1 alone: mu_pct = 0.80 - 0.08 = 0.72 VO2max
                     Step 2 alone: MET *= (1 - 0.15) -> effective pct = 0.72
                     Combined:     effective pct = 0.72 * 0.85 = 0.61 VO2max

                   At 0.61 VO2max for a tf=1.0 runner the simulated effort is
                   physiologically inconsistent with a competitive marathon
                   pace (literature: elite runners sustain ~85-90% VO2max).

                   RESOLUTION (rev11):
                   The step-2 training_offset in calculate_indices_jos3_adult
                   is RETAINED but its coefficient is reduced from 0.15 to
                   0.05, limiting the economy correction to a 5% MET reduction
                   at tf=1.0. This gives:
                     tf=0.0: effective pct = mu(0.80) * 1.00 = 0.80
                     tf=0.5: effective pct = mu(0.76) * 0.975 = 0.74
                     tf=1.0: effective pct = mu(0.72) * 0.95  = 0.68

                   These values are physiologically defensible for a
                   recreational mass-event population. Clinical calibration
                   from Bongers remains needed for precise coefficient values.
                   The coefficient value (0.05) and its rationale are now
                   explicitly documented at the point of use.

                   C) BMI FILTER LOGIC -- CLARIFIED
                   ----------------------------------
                   In rev10, bmi_target was sampled and then clipped to
                   [18,32] before weight was computed. The subsequent hard
                   rejection `if not (18 <= bmi <= 32): continue` could only
                   be triggered by the weight clip (45--105 kg), not by
                   bmi_target itself. This was a silent logic inconsistency.

                   Rev11: the hard BMI rejection is kept (it correctly catches
                   weight-clip distortions) but is now explicitly commented to
                   explain which path triggers it. The bmi_target clip range
                   is unchanged.

                   D) POPULATION FILE VERSION SAFETY
                   -----------------------------------
                   save_population() now saves a dict with keys:
                     'hestia_version': 'rev11'
                     'population':     list of AdultParticipantProfile
                   load_population() checks the version key and warns if the
                   file was created by a different revision. Backward-
                   compatible: files without a version key are loaded with a
                   deprecation warning.

                   E) MODULE-LEVEL DOCUMENTATION CONSISTENCY
                   ------------------------------------------
                   All references to DtD 2024 as the *calibration* source
                   have been replaced by Boston Marathon (Breslow 2021).
                   DtD 2024 is now correctly described as a *validation* target.
                   The module docstring, collapse-risk comments, and the main()
                   startup banner are updated to reflect rev11.

                   KNOWN REMAINING ISSUES (to be addressed in rev12):
                   - training_offset coefficient (0.05) requires clinical
                     calibration. Bongers (Radboudumc) data needed.
                   - intercept_kal = -10.071 requires recalibration against
                     DtD 2024 after training_offset is clinically settled.
                   - scipy.truncnorm not yet used for pct_vo2max (np.clip
                     remains; tail bias <0.3 pct-point, acceptable for now).

MODULE_CALIBRATION_NOTES (rev11)
=================================
Boston Marathon pooled EHS rate: 9.0 / 10,000 = 0.00090 (Breslow 2021).
Collapse risk model:
  Z = intercept_kal
      + 1.0 * max(0, T_rect - 39.5)
      + 4.0 * max(0, T_rect - 40.5)
      + 0.8 * max(0, 2.0 - CO_reserve)
      + 0.5 * max(0, dehy_pct - 3.0)

For a population running under Boston-typical conditions (T_rect ~38.8 mean,
CO_reserve ~2.5 mean, dehy ~2% mean), the dominant term is the intercept.
Newton iteration was run offline to find intercept_kal such that the mean
p_collapse across a standard N=5000 population equals 0.00090.
Result: intercept_kal = -10.071

=============================================================================
rev12  2026-04     THREE-ENDPOINT CALIBRATION ARCHITECTURE
=============================================================================
                   A) EINDPUNT-MISMATCH OPGELOST
                   --------------------------------
                   Rev11 kalibreerde het collapse-model op de Boston EHS-rate
                   (9/10.000, Breslow 2021). Maar het collapse-model meet een
                   COMPOSIET eindpunt dat breder is dan klinische EHS:
                     - T_rect > 39.5°C term vangt ook sub-klinische gevallen
                     - CO_reserve term voegt cardiovasculaire collapse toe die
                       los staat van kerntemperatuur
                     - Dehydratie term vangt incidenten zonder hyperthermie

                   Het composiet eindpunt is daarmee dichter bij "alle hitte-
                   gerelateerde medische incidenten" dan bij "EHS alleen".
                   EHS is empirisch ~10-20% van alle hitte-incidenten bij
                   marathons (Ronneberg et al. 2021, Br J Sports Med).

                   B) DRIE-LAAGS INTERCEPT TABEL
                   --------------------------------
                   Rev12 introduceert COLLAPSE_ENDPOINTS: een tabel met drie
                   gekalibreerde intercepts, elk voor een ander eindpunt.
                   De actieve intercept wordt gekozen via ACTIVE_ENDPOINT.

                   Eindpunt 1 -- EHS (klinisch)
                     Definitie : T_rect >= 40.0°C + CNS-dysfunctie
                     Bron      : Breslow RG et al. (2021) Am J Sports Med
                                 Boston Marathon 2012-2019, gepoold
                     P_obs     : 9.0 / 10.000 = 0.00090
                     intercept : -10.071  (Newton, JOS-3 gecalibreerd, rev11)

                   Eindpunt 2 -- Ziekenhuisopname
                     Definitie : Medische opname na wedstrijd (EHS + ernstige
                                 uitputting + cardiovasculaire incidenten)
                     Bron      : GHOR Noord-Holland Noord, DtD 2024
                                 50 opnames / 35.000 deelnemers
                     P_obs     : 50 / 35.000 = 0.00143 (14.3 / 10.000)
                     intercept : -9.569  (zie INTERCEPT_DERIVATION_NOTES rev12)
                     Status    : VOORLOPIG -- vereist herrekalibratie na
                                 training_offset klinische validatie (Bongers)

                   Eindpunt 3 -- EHBO-contact (alle incidenten)
                     Definitie : Elk contact met medisch personeel ter plaatse,
                                 inclusief lichte hyperthermie, uitputting,
                                 krampen, hyponatriëmie
                     Bron      : Schatting op basis van DtD 2024 GHOR-data
                                 (ongepubliceerd); aanname ~3x ziekenhuisopname
                     P_obs     : ~150 / 35.000 = 0.00429 (42.9 / 10.000)
                     intercept : -8.364  (zie INTERCEPT_DERIVATION_NOTES rev12)
                     Status    : VOORLOPIG -- brede onzekerheid op P_obs
                                 (gevoeligheidsanalyse: range -7.40 tot -8.40
                                 voor 100-250 EHBO-contacten / 35.000)

                   C) INTERCEPT_DERIVATION_NOTES (rev12)
                   ----------------------------------------
                   De intercepts voor eindpunt 2 en 3 zijn afgeleid via
                   Newton-iteratie op de analytische populatiedistributies
                   (N=50.000, zie intercept_estimation.py). De analytische
                   benadering geeft T_rect P50 ~39.1°C, terwijl JOS-3 bij
                   matige hitte P50 ~38.5°C geeft. Dit introduceert een
                   systematische verschuiving van ~1.9 eenheden in de
                   intercept.

                   Correctieprocedure:
                     analytisch_intercept(EP1) = -8.160
                     JOS3_intercept(EP1)       = -10.071  (rev11, gevalideerd)
                     correctie                 = -10.071 - (-8.160) = -1.911

                   Gecorrigeerde intercepts:
                     EP2 (ziekenhuisopname): -7.657 + (-1.911) = -9.568 ≈ -9.569
                     EP3 (EHBO-contact):     -6.453 + (-1.911) = -8.364 ≈ -8.364
                     EP3 gevoeligheidsrange: -8.81 tot -7.81

                   VALIDATIE STATUS:
                     EP1: gevalideerd tegen Boston (Breslow 2021)
                     EP2: voorlopig -- DtD 2024 is het validatiedoel zodra
                          training_offset klinisch gekalibreerd is (Bongers)
                     EP3: voorlopig -- P_obs vereist ongepubliceerde DtD-data
                          (aan te vragen bij Connie van der Lee / GHOR NHN)

                   D) ACTIVE_ENDPOINT CONFIGURATIE
                   ----------------------------------
                   Standaard actief: 'ehs' (meest conservatief, best gedocum.)
                   Wijzigen via ACTIVE_ENDPOINT constante of via user-prompt
                   in main() (nieuwe vraag toegevoegd).

                   E) OUTPUT LABELLING
                   --------------------
                   De stats-dict en Excel-export bevatten nu:
                     'active_endpoint':       str  (naam van actief eindpunt)
                     'active_endpoint_p_obs': float (gekalibreerde P_obs)
                     'active_intercept_kal':  float
                   zodat rapporten altijd traceerbaar zijn naar het eindpunt.

                   KNOWN REMAINING ISSUES (rev13):
                   - training_offset coëfficiënt (0.05) vereist klinische
                     kalibratie. Bongers (Radboudumc) data nodig.
                   - EP2 en EP3 intercepts vereisen herrekalibratie via
                     directe JOS-3 Newton-iteratie (i.p.v. analytische
                     correctie) zodra DtD EHBO-tellingen beschikbaar zijn.
                   - pct_vo2max causale conditionering op tempo (rev12 analyse)
                     staat gepland voor rev13.
                   - scipy.truncnorm voor pct_vo2max (deferred, rev13).

=============================================================================
REFERENCES
=============================================================================
Takahashi Y et al. (2021) Energy & Buildings 231:110575.
Lloyd A et al. (2022) J Appl Physiol 133:247-261.
Tanaka H et al. (2001) J Am Coll Cardiol 37:153-156.
Gonzalez-Alonso J et al. (2008) J Physiol 586:45-49.
Rae DE et al. (2008) Br J Sports Med 42:849-851.
Roberts WO (1994) Med Sci Sports Exerc 26:S197.
Rae DE & Noakes TD (2002) Sports Med 32:591-601.
Buller MJ et al. (2022) PubMed 35022161.
Vanos JK et al. (2023) Nature Communications 14:3405.
Scharhag-Rosenberger F et al. (2010) J Sci Med Sport 13:167-172.
Tanda G & Knechtle B (2013) J Hum Kinet 38:169-177.
Oke TR (1973) Boundary-Layer Meteorol 7:213-220.
Joyner MJ (1991) J Appl Physiol 70:683-687.
Jones AM & Carter H (2000) Sports Med 29:373-386.
Daniels J & Gilbert J (1979) Oxygen Power. Tempe, AZ: published by authors.
Barnes KR & Kilding AE (2015) Sports Med 45:1419-1432.
Breslow RG et al. (2021) Am J Sports Med 49(10):2696-2703.
  [Primary calibration reference: Boston Marathon EHS incidence 2012-2019]
Ronneberg K et al. (2021) Br J Sports Med 55(1):9-16.
  [EHS vs. all heat illness ratio at mass running events]

Author  : HESTIA project / Koos de Boer Apr-2026  (rev14)

=============================================================================
rev16  2026-05     VULNERABLE T_RECT TAIL PLOTTING
=============================================================================
                   Adds explicit handling of the upper 2.5% T_rect tail.
                   The vulnerable cohort is defined by each participant's
                   personal peak T_rect across the whole event, not by the
                   highest values at each time point. This keeps the plotted
                   line tied to the same high-risk runners over time.

                   Adult output now includes:
                   - P97.5 and P99 T_rect reference curves in the plot.
                   - Mean trajectory of the fixed top-2.5% vulnerable cohort.
                   - Console summary of cohort size, peak cutoff, age, VO2max,
                     K_p, NSAID prevalence, CO reserve, and dehydration.
                   - Excel export flags for participants in the vulnerable
                     top-2.5% cohort.

=============================================================================
rev15  2026-05     CALIBRATION UNCERTAINTY BANDS
=============================================================================
                   PROBLEM (rev14 and earlier):
                   Percentage outputs such as pct_decompensatie,
                   pct_roberts_kritiek, pct_ehs_postfinish, and
                   pct_kp_onder_002 were reported as single point estimates.
                   The existing Wald CI in _calculate_threshold_stats covers
                   only Monte Carlo sampling variance (~±0.5 pct-point at
                   N=5000) and conveys a false sense of precision: the
                   dominant uncertainty is calibration transfer, not N.

                   Specifically:
                   1. intercept_kal was derived via Newton iteration on a
                      single reference dataset (Boston Marathon pooled,
                      Breslow 2021). The transfer to DtD conditions
                      introduces a systematic uncertainty estimated at ±0.5
                      in log-odds units (conservative; based on between-event
                      EHS rate variation in the literature).
                   2. The Z-function coefficients (W_T1=1.0, W_T2=4.0,
                      W_C=0.8, W_D=0.5) are expert judgement values, not
                      regression estimates. No confidence interval exists for
                      these parameters pending Bongers (Radboudumc) clinical
                      calibration data.

                   SOLUTION (rev15):
                   A) _intercept_sensitivity(z_zonder_intercept, n_sim,
                      intercept_nominal, delta=0.5) computes each key
                      percentage at intercept_kal ± delta (default ±0.5).
                      This gives a calibration uncertainty band that
                      dominates the Wald MC-band at all realistic N.

                   B) _format_pct_with_band(value, lo, hi) formats output as
                      "X.X% (~Y.Y–Z.Z%)" where the tilde signals that the
                      band reflects model calibration uncertainty, not a
                      classical 95% CI.

                   C) GHOR-output (print statements in main()) is updated for
                      pct_decompensatie, pct_roberts_kritiek,
                      pct_ehs_postfinish, pct_kp_onder_002, and
                      p_collapse_gemiddeld to use _format_pct_with_band().
                      The Wald MC-CI from _calculate_threshold_stats is
                      retained unchanged (it is scientifically correct for
                      its purpose -- temperature threshold statistics).

                   D) stats dict is extended with sensitivity band fields:
                      'pct_decompensatie_lo', 'pct_decompensatie_hi'
                      'pct_roberts_kritiek_lo', 'pct_roberts_kritiek_hi'
                      'pct_ehs_postfinish_lo', 'pct_ehs_postfinish_hi'
                      'p_collapse_gemiddeld_lo', 'p_collapse_gemiddeld_hi'
                      These are available for downstream reporting and Excel
                      export without requiring a re-run.

                   EPISTEMOLOGICAL NOTE (for GHOR and Bongers reports):
                   The band "~Y–Z%" does NOT imply a 95% CI.  It means:
                   "if the calibration intercept shifts by ±0.5 log-odds
                   (a plausible range given the Boston→DtD transfer), the
                   model output moves between Y% and Z%."  The Bongers
                   clinical dataset will reduce this band by constraining the
                   Z-function coefficients directly.

=============================================================================
rev14  2026-04     THREE OPERATIONAL RELIABILITY IMPROVEMENTS
=============================================================================
                   Addresses the three structural reliability risks identified
                   in the GHOR-readiness analysis (Apr-2026):

                   A) K_p AS POPULATION VARIABLE  [reliability risk 3]
                   ---------------------------------------------------
                   PROBLEM (rev13 and earlier):
                   The pacing gain K_P_PACING = 0.10 MET/°C was a fixed
                   constant identical for all participants. The interoception
                   literature (Flouris 2024) shows that ~5-10% of endurance
                   athletes have near-zero pacing response to thermal overload
                   due to low interoceptive awareness and/or high motivation.
                   These participants continue at race pace while T_rect
                   escalates -- they are the highest EHS risk group but were
                   indistinguishable from the average population in rev13.
                   Additionally, NSAID use (~25% prevalence in recreational
                   runners) further reduces effective K_p via PGE2 suppression
                   in the brain (Garcia et al. 2019).

                   SOLUTION (rev14):
                   kp_pacing is now a per-participant field in
                   AdultParticipantProfile, sampled from:
                     kp_base ~ N(0.10, 0.04), clipped [0.00, 0.25]
                   NSAID modifier applied to 25% of population:
                     kp_pacing = kp_base * 0.5  (PGE2 suppression)
                   nsaid_gebruik (bool) stored in profile for export.
                   kp_pacing replaces the module constant K_P_PACING in
                   calculate_indices_jos3_adult().
                   The K_P_PACING constant is retained as the population
                   mean reference for documentation purposes.

                   CONSEQUENCE: intercept_kal must be recalibrated after
                   this change (K_p distribution shifts T_rect P95 upward
                   for the low-K_p subpopulation). Recommended reference
                   conditions: tdb 22->26 deg C (activates W_T2 term better).

                   References:
                   Flouris AD et al. (2024) [interoception and EHS]
                   Garcia CK et al. (2019) FASEB J [IBU and EHS in mice]
                   Nieman DC et al. (2006) Brain Behav Immun [IBU endotoxemia]

                   B) AUC_KLINISCH (ROBERTS THRESHOLD)  [reliability risk 1]
                   -----------------------------------------------------------
                   PROBLEM (rev13):
                   AUC_thermisch used a 39.0 deg C integration floor.
                   Roberts (Sports Medicine 2007) defines the clinical
                   intervention target as keeping the area under the
                   temperature-time curve above 40.5 deg C below 60
                   degree-minutes. The 39.0 deg C AUC is not a validated
                   clinical metric.

                   SOLUTION (rev14):
                   A second AUC accumulator is added:
                     AUC_klinisch += max(0, T_rect - 40.5) * dt_min
                   Clinical threshold: AUC_klinisch > 60 deg-min (Roberts)
                   Both AUCs are stored in results and exported.
                   GHOR output now includes:
                     - auc_klinisch_p95 (population P95)
                     - pct_roberts_kritiek (% with AUC_klinisch > 60)

                   Reference:
                   Roberts WO (2007) Sports Med 37(4-5):309-312.

                   C) POST-FINISH MODULE  [reliability risk 2]
                   -------------------------------------------
                   PROBLEM (rev13 and earlier):
                   Simulation stopped at the finish line. Clinical and
                   epidemiological evidence shows that 30-40% of EHS cases
                   at mass running events occur in the finish zone, not
                   during the race (Roberts 1998; Rae et al. 2008).
                   Two mechanisms drive post-finish T_rect escalation:
                     1. Metabolic after-glow: residual heat production
                        decays exponentially (tau ~ 2 min, delta ~ 0.4 deg C)
                     2. Venous pooling: muscle pump stops, CO drops acutely,
                        cutaneous heat dissipation falls
                   The pacing controller (K_p) is no longer active post-finish.
                   This creates an open-loop instability that HESTIA did not
                   model.

                   SOLUTION (rev14):
                   simulate_post_finish() propagates T_rect and CO_reserve
                   for 10 minutes after the race ends, without pacing control.
                   Called from calculate_indices_jos3_adult() after the main
                   loop. Returns:
                     t_rect_piek_postfinish   : peak T_rect in finish zone
                     co_reserve_postfinish    : CO_reserve at finish + 10 min
                     auc_klinisch_postfinish  : additional AUC_klinisch
                     ehs_postfinish           : bool -- EHS criterion met
                     t_ehs_postfinish_min     : minutes post-finish at EHS
                   Population aggregate: pct_ehs_postfinish added to stats.
                   GHOR output includes separate finish-zone risk estimate.

                   References:
                   Roberts WO (1998) [post-finish T_rect rise]
                   Rowell LB (1974) [venous pooling post-exercise]

                   RECALIBRATION REQUIREMENT
                   --------------------------
                   All three changes shift the T_rect distribution.
                   intercept_kal is marked REQUIRES_RECALIBRATION.
                   Recommended: run intercept_estimation.py with
                   tdb_start=22, tdb_end=26 (activates W_T2 term).

"""
import requests
import pandas as pd
from datetime import datetime, UTC, timedelta
import pytz
from timezonefinder import TimezoneFinder
import pvlib
import math
from pythermalcomfort.models import JOS3
from pythermalcomfort.models import utci, wbgt
from pythermalcomfort.utilities import wet_bulb_tmp
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from functools import lru_cache
import multiprocessing
from multiprocessing import Pool
from tqdm import tqdm
from colorama import init, Fore, Style
import seaborn as sns
import random
import os
import pickle

try:
    from openpyxl import load_workbook
except ImportError:
    print(f"{Fore.YELLOW}Warning: 'openpyxl' is not installed. Excel export/import will not work.{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}Install it with: pip install openpyxl{Style.RESET_ALL}")

print("Script started")

# =============================================================================
#  PARTICIPANT PROFILE DATACLASS  (rev04, extended rev10)
#
#  Fields vs. rev09:
#    - met_variation  REMOVED  (was absolute MET offset; caused VO2max
#                               exceedance bug; replaced by pct_vo2max)
#    - pct_vo2max     ADDED    (dimensionless fraction [0.60, 0.92];
#                               MET is computed on-the-fly per scenario)
#
#  All other fields unchanged. No positional indexing anywhere.
# =============================================================================
from dataclasses import dataclass as _dataclass

@_dataclass
class AdultParticipantProfile:
    """
    Complete profile of a single simulated adult participant.

    Created by generate_base_population() and used by
    calculate_indices_jos3_adult(). All fields are mandatory.

    Changes vs rev09 (rev10):
    - met_variation removed; pct_vo2max added. MET is now derived as
      vo2max * pct_vo2max / VO2MAX_TO_MET_FACTOR, guaranteeing that
      %VO2max cannot exceed the sampled value by construction.
    - age field is now set during generate_base_population() (not a
      placeholder 0 that must be filled later). The age is sampled
      immediately after gender, so VO2max can be conditioned on it.
    """
    # Anthropometrics
    height:            float   # m
    weight:            float   # kg
    age:               int     # years -- sampled within generate_base_population()
    gender:            str     # 'male' | 'female'
    # Physiology
    vo2max:            float   # mL/kg/min -- conditioned on age and gender (rev10)
    pct_vo2max:        float   # fraction of VO2max at race pace [0.60, 0.92] (rev10)
    # Environmental perception (individual)
    temp_variation:    float   # deg C offset
    rh_variation:      float   # % RH offset
    # Behavioural parameters
    mf_score:          float   # motivational factor [0-1]
    sweat_factor:      float   # sweat rate multiplier (acclimatisation effect)
    thirst_threshold:  float   # body mass loss (%) triggering drinking
    # [rev14-A] Individual pacing gain and NSAID status
    kp_pacing:         float   # MET/°C pacing gain -- population variable (rev14)
    nsaid_gebruik:     bool    # NSAID use flag -- reduces kp_pacing via PGE2


# =============================================================================
#  PHYSIOLOGICAL CONSTANTS
#
#  VO2MAX_TO_MET_FACTOR: converts VO2max (mL/kg/min) to MET.
#  1 MET = 3.5 mL/kg/min (ACSM standard).
#  This factor is the bridge between the VO2max-based population model and
#  the MET-based JOS-3 / RPE computations.
#
#  Example: VO2max=50, pct_vo2max=0.77 ->
#    final_met = 50 * 0.77 / 3.5 = 11.0 MET  (consistent with a ~11 km/h pace)
# =============================================================================
VO2MAX_TO_MET_FACTOR = 3.5   # mL/kg/min per MET (ACSM standard)


# =============================================================================
#  CVR MODULE IMPORTS
# =============================================================================
try:
    from HESTIA_CVR_Module_v2 import (
        RunnerProfile, JOS3Outputs,
        koppel_cvr_aan_jos3
    )
    from HESTIA_CVR_Console import (
        print_cvr_population_summary,
        print_cvr_time_series,
        print_cvr_comparison,
        calculate_cvr_risk_score,
    )
    CVR_BESCHIKBAAR = True
    print(f"{Fore.GREEN}CVR module geladen (Lloyd et al. 2022).{Style.RESET_ALL}")
except ImportError as _cvr_err:
    CVR_BESCHIKBAAR = False
    print(f"{Fore.YELLOW}CVR module niet gevonden ({_cvr_err}). "
          f"Simulatie draait zonder cardiovasculaire analyse.{Style.RESET_ALL}")


# =============================================================================
#  COLLAPSE RISK CALIBRATION CONSTANTS  (rev12 -- THREE-ENDPOINT ARCHITECTURE)
#
#  Rev12 introduces a calibration table with three outcome endpoints.
#  Each endpoint has its own intercept_kal, P_obs, and documentation.
#  The active endpoint is selected via ACTIVE_ENDPOINT (default: 'ehs').
#
#  See full derivation in MODULE_CALIBRATION_NOTES / INTERCEPT_DERIVATION_NOTES
#  in the module docstring above.
#
#  ENDPOINT DEFINITIONS
#  --------------------
#  'ehs'          : Clinical EHS (T_rect >= 40°C + CNS dysfunction)
#                   Source: Breslow et al. 2021, Boston Marathon 2012-2019
#                   Status: VALIDATED
#
#  'hospitalisation': Hospital admission after event (EHS + severe exhaustion
#                   + cardiovascular incidents requiring overnight care)
#                   Source: GHOR Noord-Holland Noord, DtD 2024
#                   Status: PROVISIONAL -- pending Bongers calibration
#
#  'ehbo'         : Any on-site medical contact (incl. mild hyperthermia,
#                   exhaustion, cramps, hyponatraemia)
#                   Source: Estimated from DtD 2024 GHOR data (unpublished)
#                   P_obs range: 100-250/35.000; central estimate 150/35.000
#                   Status: PROVISIONAL -- P_obs requires GHOR NHN data
#
#  INTERCEPT DERIVATION SUMMARY
#  -----------------------------
#  Analytical Newton iteration (N=50.000, intercept_estimation.py):
#    EP1 analytic:  -8.160  ->  JOS-3 corrected: -10.071  (delta = -1.911)
#    EP2 analytic:  -7.657  ->  JOS-3 corrected:  -9.568
#    EP3 analytic:  -6.453  ->  JOS-3 corrected:  -8.364
#  JOS-3 correction applied uniformly (T_rect offset between analytic
#  approximation and JOS-3 output at moderate heat: ~1.9 units).
# =============================================================================

COLLAPSE_ENDPOINTS = {
    'ehs': {
        'label':        'EHS (klinisch, Boston Marathon gepoold)',
        'p_obs':        9.0 / 10_000,           # 0.00090
        'n_obs':        9.0,
        'n_participants': 10_000,
        'intercept_kal': -7.903186,# gecalibreerd 2026-05-19
        'source':       'Breslow RG et al. (2021) Am J Sports Med 49(10):2696-2703',
        'status':       'VALIDATED',   # [rev13] pct_vo2max fix + dynamic pacing shift T_rect distribution
        'notes':        ('Boston Marathon 2012-2019 gepoold. Klinische EHS: '
                         'T_rect >= 40°C + CNS-dysfunctie. Smalste eindpunt. '
                         '[rev13] Intercept -10.071 gecalibreerd op rev12 populatie. '
                         'Herrekalibratie vereist na pct_vo2max causal fix (P3) '
                         'en dynamic pacing (P4). Run intercept_estimation.py.'),
    },
    'hospitalisation': {
        'label':        'Ziekenhuisopname (DtD 2024)',
        'p_obs':        50.0 / 35_000,          # 0.00143
        'n_obs':        50.0,
        'n_participants': 35_000,
        'intercept_kal': -7.438276,# gecalibreerd 2026-05-19
        'source':       'GHOR Noord-Holland Noord, Dam tot Damloop 2024',
        'status':       'PROVISIONAL',
        'notes':        ('50 ziekenhuisopnames / 35.000 deelnemers DtD 2024. '
                         'Intercept analytisch gecorrigeerd voor JOS-3 offset. '
                         'Herrekalibratie vereist na Bongers-data (Radboudumc).'),
    },
    'ehbo': {
        'label':        'EHBO-contact, alle incidenten (DtD 2024 schatting)',
        'p_obs':        150.0 / 35_000,         # 0.00429; centrale schatting
        'n_obs':        150.0,                   # onzeker: range 100-250
        'n_participants': 35_000,
        'intercept_kal': -6.324813,# rev14 gecalibreerd 2026-05-19
        'source':       'Schatting GHOR NHN / DtD 2024 (ongepubliceerd)',
        'status':       'PROVISIONAL',
        'notes':        ('Centrale schatting: ~150 EHBO-contacten / 35.000. '
                         'Gevoeligheidsrange intercept: -8.81 (n=100) tot '
                         '-7.81 (n=250). P_obs vereist bevestiging van '
                         'Connie van der Lee (GHOR NHN).'),
        'sensitivity': {
            100: -6.737571,
            125: -6.510751,
            150: -6.324813,
            175: -6.167102,
            200: -6.030062,
            225: -5.908818,
            250: -5.800042,  
        },
    },
}

# Active endpoint -- change here or via user prompt in main()
# Options: 'ehs' | 'hospitalisation' | 'ehbo'
ACTIVE_ENDPOINT = 'ehs'

# Convenience aliases (backward compatible with code that uses P_OBS_COLLAPSE)
_ep                       = COLLAPSE_ENDPOINTS[ACTIVE_ENDPOINT]
P_OBS_COLLAPSE            = _ep['p_obs']
N_OBS_COLLAPSE_ADMISSIONS = _ep['n_obs']
N_OBS_PARTICIPANTS        = _ep['n_participants']

HESTIA_VERSION            = 'rev15'   # population file schema version-safety; rev16 changes output only

# =============================================================================
#  REV14 CONFIGURATION CONSTANTS
# =============================================================================

EVENT_PACE_MIN_PER_KM    = 5.47    # min/km  -- override per event (5,47 mediaan Leiden marathon)
K_P_PACING               = 0.10   # MET/°C  -- population MEAN reference (Ely 2007)
                                   # [rev14] individual kp_pacing replaces this in sim
K_P_SIGMA                = 0.04   # std dev of K_p population distribution
K_P_NSAID_FACTOR         = 0.50   # NSAID halves effective K_p (PGE2 suppression)
NSAID_PREVALENTIE        = 0.25   # 25% of recreational runners use NSAID (estimate)
T_RECT_PACING_THRESHOLD  = 38.5   # °C -- onset of voluntary pacing reduction
T_AUC_THRESHOLD          = 39.0   # °C -- integration floor AUC_thermisch (diagnostic)
T_AUC_KLINISCH           = 40.5   # °C -- Roberts (2007) clinical threshold
AUC_ROBERTS_GRENS        = 60.0   # °C·min -- Roberts clinical intervention limit
# Post-finish module constants (Roberts 1998; Rowell 1974)
PF_DELTA_T_NAGLOED       = 0.40   # °C  -- metabolic after-glow peak amplitude
PF_TAU_NAGLOED           = 2.0    # min -- exponential decay time constant
PF_DELTA_CO_POOLING      = 0.30   # fraction of CO_reserve lost to venous pooling
PF_TAU_POOLING           = 1.5    # min -- pooling time constant
PF_KOELING_PER_MIN       = 0.03   # °C/min -- passive cooling without running
PF_DUUR_MIN              = 10.0   # min -- post-finish simulation window
PF_DT_MIN                = 0.5    # min -- time step post-finish


# =============================================================================
#  UTILITY FUNCTIONS (unchanged from rev09)
# =============================================================================

def wind_speed_at_height(v_ref, z_ref, z_target, z0=0.1):
    """Converts wind speed from reference height to target height (log profile)."""
    if z_ref <= 0 or z_target <= 0 or z0 <= 0:
        raise ValueError("Heights and roughness length must be positive.")
    return v_ref * (np.log(z_target / z0) / np.log(z_ref / z0))


# =============================================================================
#  ACTIVITY / CLOTHING LOOKUP TABLES (unchanged from rev09)
# =============================================================================

MET_ACTIVITIES_ADULT = {
    1:  {"description": "Jogging, 8 km/h (recreational)",             "met": 8.3,  "speed_kmh": 8.0},
    2:  {"description": "Running, 9.6 km/h (recreational)",           "met": 9.8,  "speed_kmh": 9.6},
    3:  {"description": "Running, 11.2 km/h (recreational)",          "met": 11.0, "speed_kmh": 11.2},
    4:  {"description": "Running, 12.8 km/h (recreational)",          "met": 11.8, "speed_kmh": 12.8},
    5:  {"description": "Trail running, moderate pace (recreational)", "met": 9.0,  "speed_kmh": 9.0},
    6:  {"description": "Running, 13.8 km/h (professional)",          "met": 12.3, "speed_kmh": 13.8},
    7:  {"description": "Running, 16 km/h (professional)",            "met": 14.5, "speed_kmh": 16.0},
    8:  {"description": "Running, 19.2 km/h (professional)",          "met": 16.0, "speed_kmh": 19.2},
    9:  {"description": "Interval sprints, high intensity (professional)", "met": 18.0, "speed_kmh": 20.0},
    10: {"description": "Marathon, competition pace (professional)",   "met": 13.5, "speed_kmh": 15.0},
    11: {"description": "Running, average (general)",                  "met": 10.0, "speed_kmh": 10.0},
    12: {"description": "Walking, 3.2 km/h (slow, recreational)",     "met": 2.8,  "speed_kmh": 3.2},
    13: {"description": "Walking, 4.0 km/h (moderate, recreational)", "met": 3.5,  "speed_kmh": 4.0},
    14: {"description": "Walking, 4.8 km/h (brisk, recreational)",    "met": 4.3,  "speed_kmh": 4.8},
    15: {"description": "Walking, 5.6 km/h (fast, fitness)",          "met": 5.0,  "speed_kmh": 5.6},
    16: {"description": "Walking, 6.4 km/h (very fast, fitness)",     "met": 6.0,  "speed_kmh": 6.4},
    17: {"description": "Walking uphill, 4.0 km/h (3-5% incline)",   "met": 5.3,  "speed_kmh": 4.0},
    18: {"description": "Walking with backpack, 4.0 km/h (moderate)", "met": 4.5,  "speed_kmh": 4.0},
    19: {"description": "Sitting outdoors",                            "met": 1.2,  "speed_kmh": 0.0},
}

MET_ACTIVITIES_CHILD = {
    1:  {"description": "Sitting quietly (e.g., classroom)",            "met": 1.3, "speed_kmh": 0.0},
    2:  {"description": "Walking slowly (e.g., to school)",             "met": 2.5, "speed_kmh": 3.0},
    3:  {"description": "Light play (e.g., sandbox, standing activity)","met": 3.0, "speed_kmh": 1.0},
    4:  {"description": "Active play / General gymnastics",             "met": 4.0, "speed_kmh": 2.0},
    5:  {"description": "Running / Very active play (e.g., tag)",       "met": 6.0, "speed_kmh": 6.0},
    6:  {"description": "Walking, brisk pace",                          "met": 3.8, "speed_kmh": 5.0},
    7:  {"description": "Cycling, leisurely pace",                      "met": 4.0, "speed_kmh": 12.0},
    8:  {"description": "Soccer, casual / kickabout",                   "met": 7.0, "speed_kmh": 7.0},
    9:  {"description": "Basketball, shooting baskets",                  "met": 4.5, "speed_kmh": 3.0},
    10: {"description": "Jogging, moderate pace",                       "met": 8.0, "speed_kmh": 9.0},
}

CLO_OPTIONS_ADULT = {
    1: {"description": "Shorts and T-shirt (summer)",              "clo": 0.2},
    2: {"description": "Shorts and tank top (very light)",         "clo": 0.1},
    3: {"description": "Long pants and T-shirt (mild weather)",    "clo": 0.4},
    4: {"description": "Long pants and long sleeves (cool)",       "clo": 0.6},
    5: {"description": "Running jacket and long pants (cold)",     "clo": 0.8},
    6: {"description": "Thermal underwear and jacket (winter)",    "clo": 1.0},
}

CLO_OPTIONS_CHILD = {
    1: {"description": "Shorts and T-shirt (summer)",              "clo": 0.2},
    2: {"description": "Shorts and tank top (very light)",         "clo": 0.1},
    3: {"description": "School uniform (trousers/skirt, polo)",    "clo": 0.4},
    4: {"description": "Tracksuit / long sleeves (cool weather)",  "clo": 0.6},
    5: {"description": "Light jacket and long pants (cold)",       "clo": 0.8},
}

LIVEABILITY_LIMITS = {
    'young_adult_shaded': {
        (25, 90): 5.0, (25, 50): 7.0, (25, 20): 8.4,
        (30, 90): 3.5, (30, 50): 5.5, (30, 20): 7.0,
        (35, 90): 1.5, (35, 50): 4.0, (35, 20): 5.8,
        (40, 70): 1.5, (40, 40): 3.0, (40, 10): 5.0,
        (45, 40): 1.5, (45, 10): 3.8,
        (50, 20): 2.0, (50, 10): 3.0,
    },
    'older_adult_shaded': {
        (25, 90): 4.2, (25, 50): 5.9, (25, 20): 6.8,
        (30, 90): 2.8, (30, 50): 4.5, (30, 20): 5.5,
        (35, 90): 1.5, (35, 50): 3.0, (35, 20): 4.2,
        (40, 60): 1.5, (40, 30): 2.2, (40, 10): 3.4,
        (45, 25): 1.5, (45, 10): 2.5,
        (50, 10): 1.8,
    },
}


def get_liveability_threshold(temp, rh, age):
    """Gets M_max (maximum safe MET) from Vanos et al. (2023) lookup table."""
    profile = 'older_adult_shaded' if age > 65 else 'young_adult_shaded'
    data = LIVEABILITY_LIMITS[profile]
    closest_point = min(data.keys(), key=lambda k: abs(k[0] - temp) + abs(k[1] - rh) / 5)
    return data[closest_point]


def calculate_uhi_effect(population, wind_speed, is_daytime):
    """UHI effect (Oke 1973). Returns delta-T in °C."""
    if population < 1000:
        return 0.0
    uhi_base = 2.0 * math.log10(population) - 4.0
    uhi_base = max(0.0, min(8.0, uhi_base))
    wind_factor = max(0.0, 1.0 - 0.1 * wind_speed)
    time_factor = 0.5 if is_daytime else 1.0
    return round(uhi_base * wind_factor * time_factor, 2)


# =============================================================================
#  POPULATION GENERATION (rev10 -- replaces simulate_runner)
#
#  DESIGN RATIONALE
#  ----------------
#  The previous simulate_runner() sampled VO2max and MET independently, which
#  produced physiologically impossible combinations: a runner with VO2max=25
#  assigned a MET implying 120% VO2max. About 18.7% of the simulated
#  population exceeded 100% VO2max (identified in rev09 analysis).
#
#  The new generate_base_population() uses a CAUSAL SAMPLING ORDER:
#
#    Step 1  gender          -- onafhankelijk; 58% man / 42% vrouw
#                              (West-Europese grote loopevents, RunRepeat 2023)
#    Step 2  age             -- N(40, 10), geclipped [18, 65]
#                              (mediane leeftijd recreatieve marathon NL)
#    Step 3  VO2max          -- geconditioneerd op leeftijd EN geslacht
#                              Basiswaarden: man N(52,8), vrouw N(44,7)
#                              Leeftijdscorrectie: -0.5 mL/kg/min per jaar > 35
#                              Ref: Scharhag-Rosenberger (2010); Tanaka (2001)
#    Step 4  height          -- geconditioneerd op geslacht
#                              man N(1.78, 0.07), vrouw N(1.65, 0.06)
#                              Ref: RIVM NL antropometrische data
#    Step 5  weight          -- geconditioneerd op hoogte + geslacht
#                              BMI-filter [18.0, 32.0] actief
#                              Ref: Tanda & Knechtle (2013)
#    Step 6  pct_vo2max      -- [rev13] CAUSAL: fractie van VO2max op event-tempo.
#                              Deterministisch via Daniels-Gilbert inversie:
#                                pct = VO2_at_event_pace / vo2max_i + ruis(0.04)
#                              Lost Pearl Causal Markov-schending op uit rev10-rev12.
#                              EVENT_PACE_MIN_PER_KM is de sturende parameter.
#    Step 7  sweat_factor    -- geconditioneerd op acclimatization_factor
#    Step 8  thirst_threshold-- geconditioneerd op acclimatization_factor
#    Step 9  temp/rh_var     -- onafhankelijk (perceptuele ruis)
#    Step 10 mf_score        -- onafhankelijk (motivatie)
#
#  MET COMPUTATION (on-the-fly, niet opgeslagen als populatieparameter)
#  ----------------------------------------------------------------------
#  final_met = vo2max * pct_vo2max / VO2MAX_TO_MET_FACTOR
#
#  Dit garandeert per constructie dat %VO2max == pct_vo2max (geclipped
#  op [0.60, 0.92]), zodat fysiologisch onmogelijke combinaties niet
#  kunnen optreden. De training_factor verschuift de mu van pct_vo2max
#  deterministische naar beneden (getraindere lopers = lagere fractie),
#  consistent met de literatuur (Joyner 1991).
#
#  FIXED SEED SUPPORT
#  ------------------
#  Gebruik random_seed voor reproduceerbare scenario-vergelijkingen.
#  Dezelfde seed geeft altijd dezelfde N lopers. Sla de populatie op
#  als .pkl met save_population() voor hergebruik zonder her-sampling.
# =============================================================================

def generate_base_population(
    n_simulations: int = 5000,
    training_factor: float = 0.0,
    acclimatization_factor: float = 0.0,
    random_seed: int = None,
) -> list:
    """
    Generate a fixed base population of adult marathon runners for Monte Carlo
    simulation.

    This function replaces simulate_runner() (rev09) with a causally correct
    sampling architecture. See module docstring (rev10) for full design
    rationale.

    The critical difference from rev09 is that MET is NOT sampled here.
    Instead, pct_vo2max (the fraction of VO2max at race pace) is sampled and
    stored. The actual MET is computed on-the-fly in
    calculate_indices_jos3_adult() as:
        final_met = vo2max * pct_vo2max / VO2MAX_TO_MET_FACTOR

    This guarantees that no participant can exceed their own VO2max by
    construction (was ~18.7% of population in rev09 due to independent
    VO2max and MET sampling).

    Parameters
    ----------
    n_simulations : int
        Target number of valid participants to generate.
    training_factor : float in [0.0, 1.0]
        Training status. Shifts mu of pct_vo2max downward (trained runners
        sustain a lower %VO2max at race pace for the same absolute pace).
        Effect: mu_pct -= training_factor * 0.08 (max 8 percentage points).
    acclimatization_factor : float in [0.0, 1.0]
        Heat acclimatization. Shifts sweat_factor up and thirst_threshold down.
    random_seed : int or None
        If provided, np.random and random are seeded before sampling.
        Use the same seed for reproducible scenario comparisons.

    Returns
    -------
    list of AdultParticipantProfile
        Successfully generated profiles (may be fewer than n_simulations
        if BMI filter rejects many candidates; max 30 attempts per profile).

    References
    ----------
    Scharhag-Rosenberger F et al. (2010) J Sci Med Sport 13:167-172.
    Tanaka H et al. (2001) J Am Coll Cardiol 37:153-156.
    Tanda G & Knechtle B (2013) J Hum Kinet 38:169-177.
    Joyner MJ (1991) J Appl Physiol 70:683-687.
    Jones AM & Carter H (2000) Sports Med 29:373-386.
    RunRepeat European marathon statistics (2023).
    """
    if random_seed is not None:
        np.random.seed(random_seed)
        random.seed(random_seed)

    valid_simulations = []
    attempts = 0
    max_attempts = n_simulations * 30  # generous budget for BMI filter

    print(f"\nGenerating base population (N={n_simulations}, seed={random_seed}, "
          f"training_factor={training_factor:.2f}, "
          f"acclimatization_factor={acclimatization_factor:.2f})...")

    while len(valid_simulations) < n_simulations and attempts < max_attempts:
        attempts += 1

        # ------------------------------------------------------------------
        # STEP 1: Gender
        # ------------------------------------------------------------------
        # West-European large running events: ~58% male, ~42% female.
        # Source: RunRepeat European marathon statistics (2023).
        gender = np.random.choice(["male", "female"], p=[0.58, 0.42])

        # ------------------------------------------------------------------
        # STEP 2: Age
        # ------------------------------------------------------------------
        # Recreational marathon runners: median ~40, skewed right (more older
        # runners than younger). N(40, 10) clipped [18, 65] approximates the
        # observed finish-start age distribution for NL events (DtD, AMS, RTD).
        age = int(np.clip(np.random.normal(loc=40.0, scale=10.0), 18, 65))

        # ------------------------------------------------------------------
        # STEP 3: VO2max -- conditioned on age AND gender
        # ------------------------------------------------------------------
        # Base values for recreational marathon finishers:
        #   Male:   N(52, 8) mL/kg/min
        #   Female: N(44, 7) mL/kg/min
        # These reflect self-selected mass-event runners, not general population.
        # Ref: Scharhag-Rosenberger et al. (2010).
        #
        # Age correction: -0.5 mL/kg/min per year after age 35.
        # Ref: Tanaka et al. (2001) J Am Coll Cardiol 37:153.
        # The correction is applied to the distribution mean, not post-hoc,
        # so the variance is preserved across age groups.
        if gender == "male":
            vo2max_base_mu = 52.0
            vo2max_sigma   = 8.0
            vo2max_clip    = (28.0, 78.0)
        else:
            vo2max_base_mu = 44.0
            vo2max_sigma   = 7.0
            vo2max_clip    = (24.0, 68.0)

        # Deterministic age correction (Tanaka 2001)
        age_correction = 0.5 * max(0, age - 35)
        vo2max_mu = vo2max_base_mu - age_correction

        vo2max = float(np.clip(
            np.random.normal(loc=vo2max_mu, scale=vo2max_sigma),
            vo2max_clip[0], vo2max_clip[1]
        ))

        # ------------------------------------------------------------------
        # STEP 4: Height -- conditioned on gender
        # ------------------------------------------------------------------
        # Reference: RIVM NL antropometrische data; Tolfrey & Pilcher (2013).
        if gender == "male":
            height = float(np.clip(np.random.normal(loc=1.78, scale=0.07), 1.60, 2.00))
        else:
            height = float(np.clip(np.random.normal(loc=1.65, scale=0.06), 1.48, 1.85))

        # ------------------------------------------------------------------
        # STEP 5: Weight -- conditioned on height and gender (BMI target)
        # ------------------------------------------------------------------
        # Recreational marathon runners have lower BMI than general population.
        # Target BMI range: [18.0, 30.0] (active runner population).
        # Ref: Tanda & Knechtle (2013) J Hum Kinet 38:169-177.
        # BMI target mu: 23.0 (male), 21.5 (female).
        if gender == "male":
            bmi_target = np.clip(np.random.normal(loc=23.0, scale=2.0), 18.0, 30.0)
        else:
            bmi_target = np.clip(np.random.normal(loc=21.5, scale=2.0), 18.0, 28.0)

        weight = float(np.clip(bmi_target * height ** 2, 45.0, 105.0))

        # BMI filter: hard rejection for implausible combinations.
        # NOTE (rev11): bmi_target is already clipped to [18,30]/[18,28], so
        # the weight derived from it will also satisfy BMI [18,30] in the
        # absence of the weight clip. The weight clip (45--105 kg) can
        # distort BMI for extreme height values (very tall/very light runners),
        # which is the actual path that triggers this rejection. The check is
        # retained to guard against such edge cases.
        bmi = weight / (height ** 2)
        if not (18.0 <= bmi <= 32.0):
            continue

        # ------------------------------------------------------------------
        # STEP 6: %VO2max at race pace -- conditioned on VO2max via event pace
        # ------------------------------------------------------------------
        # [rev13] CAUSAL FIX (P3): pct_vo2max is now derived deterministically
        # from the participant's own VO2max and the event pace, resolving the
        # Pearl Causal Markov violation present in rev10-rev12.
        #
        # Architecture:
        #   1. Compute the VO2 demand at the event pace using Daniels-Gilbert.
        #   2. Divide by this participant's VO2max to get their individual %VO2max.
        #   3. Add Gaussian noise (sigma=0.04) to represent hardloopeconomie
        #      variability (Barnes & Kilding 2015: CV ~4-6% in homogeneous groups).
        #
        # The training_factor mu-shift from rev10-rev12 (-0.08 * tf) is REMOVED
        # here. Economy effects remain via the step-2 training_offset in
        # calculate_indices_jos3_adult() (coefficient 0.05, rev11 unchanged).
        #
        # EVENT_PACE_MIN_PER_KM is a module-level constant (default 6.0 min/km).
        # Override it before calling generate_base_population() for event-specific
        # calibration (e.g., Boston: 5.8 min/km; DtD: 6.5 min/km).
        #
        # Ref: Joyner (1991); Jones & Carter (2000); Barnes & Kilding (2015).
        #vo2_at_event_pace     = _daniels_gilbert_vo2_at_pace(EVENT_PACE_MIN_PER_KM)
        #pct_vo2max_det        = vo2_at_event_pace / vo2max   # physiologically conditioned
        
        # NA (geslacht-specifieke pace, afgeleid van mediaan VO2max per geslacht):
        EVENT_PACE_M = 4.97  # min/km  → pct_det mediaan mannen ~0.74
        EVENT_PACE_F = 5.77  # min/km  → pct_det mediaan vrouwen ~0.74
        pace_this_runner = EVENT_PACE_M if gender == "male" else EVENT_PACE_F
        vo2_at_event_pace = _daniels_gilbert_vo2_at_pace(pace_this_runner)
        pct_vo2max_det = vo2_at_event_pace / vo2max
        
        pct_vo2max_sigma      = 0.04   # economy variability (narrower than rev12's 0.05)
        pct_vo2max = float(np.clip(
            np.random.normal(loc=pct_vo2max_det, scale=pct_vo2max_sigma),
            0.55, 0.95
        ))

        # ------------------------------------------------------------------
        # STEP 7: Sweat factor -- conditioned on acclimatization_factor
        # ------------------------------------------------------------------
        # Acclimatized runners sweat earlier and at higher rates (higher factor).
        # Ref: Periard et al. (2016) Comprehensive Physiology.
        sweat_loc = 0.8 + (acclimatization_factor * 0.3)  # range [0.80, 1.10]
        sweat_factor = float(np.clip(
            np.random.normal(loc=sweat_loc, scale=0.15),
            0.4, 1.5
        ))

        # ------------------------------------------------------------------
        # STEP 8: Thirst threshold -- conditioned on acclimatization_factor
        # ------------------------------------------------------------------
        # Acclimatized runners drink at lower body-weight-loss thresholds.
        thirst_upper_bound = 2.5 - (acclimatization_factor * 1.0)  # range [1.5, 2.5]
        thirst_threshold = float(np.random.uniform(1.0, max(1.1, thirst_upper_bound)))

        # ------------------------------------------------------------------
        # STEP 9: Environmental perception (individual noise)
        # ------------------------------------------------------------------
        temp_variation = float(np.random.normal(loc=0.0, scale=1.5))
        rh_variation   = float(np.random.normal(loc=0.0, scale=3.0))

        # ------------------------------------------------------------------
        # STEP 10: Motivational factor (independent)
        # ------------------------------------------------------------------
        mf_score = float(np.random.uniform(0.0, 1.0))

        # ------------------------------------------------------------------
        # STEP 11: Individual pacing gain K_p  [rev14-A]
        # ------------------------------------------------------------------
        # K_p is the proportional gain of the voluntary pacing feedback loop:
        # MET reduction = K_p * max(0, T_rect - T_RECT_PACING_THRESHOLD).
        # In rev13, K_p was fixed at 0.10 for all participants.
        # In rev14, K_p is normally distributed across the population to
        # reflect individual variation in interoceptive awareness and
        # motivation (Flouris 2024).
        #
        # K_p ~ N(0.10, 0.04), clipped [0.00, 0.25]
        # Participants with K_p ≈ 0 are the highest EHS risk group:
        # they do not reduce pace despite rising T_rect.
        #
        # NSAID modifier: ~25% of recreational marathon runners use NSAID
        # before or during the race (ibuprofen predominantly). NSAID crosses
        # the blood-brain barrier and suppresses PGE2, reducing the
        # subjective perception of thermal overload -- effectively halving
        # the pacing response (Garcia et al. 2019 mouse model; Nieman 2006
        # ultramarathon field study).
        nsaid_gebruik = bool(np.random.binomial(1, NSAID_PREVALENTIE))
        kp_base       = float(np.clip(
            np.random.normal(loc=K_P_PACING, scale=K_P_SIGMA),
            0.0, 0.25
        ))
        kp_pacing = kp_base * (K_P_NSAID_FACTOR if nsaid_gebruik else 1.0)

        # Append valid profile
        valid_simulations.append(AdultParticipantProfile(
            height           = height,
            weight           = weight,
            age              = age,
            gender           = gender,
            vo2max           = vo2max,
            pct_vo2max       = pct_vo2max,
            temp_variation   = temp_variation,
            rh_variation     = rh_variation,
            mf_score         = mf_score,
            sweat_factor     = sweat_factor,
            thirst_threshold = thirst_threshold,
            kp_pacing        = kp_pacing,        # [rev14-A] individual pacing gain
            nsaid_gebruik    = nsaid_gebruik,    # [rev14-A] NSAID flag
        ))

    if len(valid_simulations) < n_simulations:
        print(f"{Fore.YELLOW}Warning: generated {len(valid_simulations)}/{n_simulations} "
              f"valid profiles ({attempts} attempts).{Style.RESET_ALL}")
    else:
        print(f"{Fore.GREEN}Generated {len(valid_simulations)} valid profiles "
              f"({attempts} attempts).{Style.RESET_ALL}")

    return valid_simulations


# =============================================================================
#  POPULATION PERSISTENCE (rev10)
#
#  A base population can be saved to disk and reloaded for reproducible
#  scenario comparisons. This avoids re-generating and re-seeding for each run.
# =============================================================================

def save_population(population: list, filepath: str) -> None:
    """
    Save a base population to a pickle file with version metadata.

    Parameters
    ----------
    population : list of AdultParticipantProfile
    filepath   : str -- path to .pkl file (e.g. 'base_population_N5000_seed42.pkl')

    [rev11] The file now stores a dict with keys 'hestia_version' and
    'population'. load_population() checks the version and warns on mismatch.
    """
    payload = {
        'hestia_version': HESTIA_VERSION,
        'population':     population,
    }
    with open(filepath, 'wb') as f:
        pickle.dump(payload, f)
    print(f"{Fore.GREEN}Population saved to '{filepath}' "
          f"({len(population)} profiles, version={HESTIA_VERSION}).{Style.RESET_ALL}")


def load_population(filepath: str) -> list:
    """
    Load a base population from a pickle file.

    Parameters
    ----------
    filepath : str -- path to .pkl file

    Returns
    -------
    list of AdultParticipantProfile

    [rev11] Version safety: files saved by rev11+ contain a dict with key
    'hestia_version'. Files saved by earlier revisions contain a bare list.
    A deprecation warning is printed for pre-rev11 files or version mismatches.
    """
    with open(filepath, 'rb') as f:
        payload = pickle.load(f)

    if isinstance(payload, dict):
        # rev11+ format
        file_version = payload.get('hestia_version', 'unknown')
        population   = payload['population']
        if file_version != HESTIA_VERSION:
            print(f"{Fore.YELLOW}[rev11] WARNING: population file was created by "
                  f"'{file_version}', current version is '{HESTIA_VERSION}'. "
                  f"Schema changes between versions may cause AttributeErrors. "
                  f"Regenerate the population if errors occur.{Style.RESET_ALL}")
        else:
            print(f"{Fore.GREEN}Population loaded from '{filepath}' "
                  f"({len(population)} profiles, version={file_version}).{Style.RESET_ALL}")
    else:
        # Pre-rev11 bare list
        population = payload
        print(f"{Fore.YELLOW}[rev11] WARNING: population file has no version tag "
              f"(pre-rev11 format). Consider regenerating with rev11 to ensure "
              f"schema compatibility.{Style.RESET_ALL}")
        print(f"{Fore.GREEN}Loaded {len(population)} profiles (version=pre-rev11).{Style.RESET_ALL}")

    return population


# =============================================================================
#  POPULATION VALIDATION (rev10 -- NEW)
#
#  validate_population() performs three levels of validation:
#
#  1. MARGINAL VALIDATION
#     Compares distribution statistics (mean, SD, percentiles) of each
#     population parameter against published empirical references.
#     A WARNING is printed for any parameter where the simulated mean
#     deviates by more than 1 SD from the reference value.
#
#  2. CORRELATION VALIDATION
#     Checks physiologically expected correlations in the generated population:
#       - VO2max vs age:    expected r in [-0.50, -0.25]
#       - VO2max vs gender: expected Cohen's d in [0.70, 1.00]
#         (coded as 0/1 for female/male)
#     Values outside these ranges trigger a WARNING.
#
#  3. FINISH-TIME VALIDATION (afgeleide validatie)
#     Uses the Daniels & Gilbert (1979) VO2max-performance formula to compute
#     an implied marathon finish time for each participant based on their
#     VO2max and pct_vo2max. The population finish-time distribution is then
#     compared to the reference range for a recreational Dutch marathon field:
#       - Median finish time: expected [3h30m, 4h30m]
#       - P5 (fast tail):     expected [2h45m, 3h30m]
#       - P95 (slow tail):    expected [4h30m, 5h30m]
#     This is the strongest test: if the simulated population finishes
#     collectively in a realistic time range, the joint distribution of
#     VO2max, %VO2max, and age is empirically anchored.
#
#  OUTPUT
#     A summary table is printed to console and a 3-panel validation figure
#     is displayed (marginals + correlation matrix + finish-time histogram).
#
#  REFERENCES
#     Scharhag-Rosenberger F et al. (2010) J Sci Med Sport 13:167-172.
#     Tanda G & Knechtle B (2013) J Hum Kinet 38:169-177.
#     Tanaka H et al. (2001) J Am Coll Cardiol 37:153-156.
#     Daniels J & Gilbert J (1979) Oxygen Power. Tempe, AZ.
#     RunRepeat European marathon statistics (2023).
# =============================================================================

# Daniels-Gilbert VO2max <-> performance constants (marathon distance)
_MARATHON_DIST_KM = 42.195

def _daniels_gilbert_vo2_at_pace(pace_min_per_km: float) -> float:
    """
    Estimate VO2 (mL/kg/min) from marathon pace using Daniels & Gilbert (1979).
    pace_min_per_km: pace in minutes per kilometer.
    Returns VO2 in mL/kg/min.
    """
    # Velocity in m/min
    v = 1000.0 / pace_min_per_km
    # Daniels & Gilbert oxygen cost formula:
    # VO2 = -4.60 + 0.182258*v + 0.000104*v^2  (v in m/min)
    vo2 = -4.60 + 0.182258 * v + 0.000104 * v ** 2
    return max(1.0, vo2)


def _daniels_gilbert_pace_from_vo2max(vo2max: float, pct_vo2max: float) -> float:
    """
    Estimate marathon pace (min/km) from VO2max and fraction sustained.

    The sustained VO2 at race pace is vo2max * pct_vo2max. We invert the
    Daniels-Gilbert formula to find the velocity v (m/min) that corresponds
    to that VO2, then convert to pace.

    The quadratic 0.000104*v^2 + 0.182258*v - (VO2 + 4.60) = 0 is solved
    analytically.
    """
    vo2_race = vo2max * pct_vo2max
    a = 0.000104
    b = 0.182258
    c = -(vo2_race + 4.60)
    discriminant = b ** 2 - 4 * a * c
    if discriminant < 0:
        return 999.0  # physiologically impossible
    v = (-b + math.sqrt(discriminant)) / (2 * a)   # m/min
    if v <= 0:
        return 999.0
    pace_min_per_km = 1000.0 / v  # min/km
    return pace_min_per_km


def _finish_time_hours(vo2max: float, pct_vo2max: float) -> float:
    """Returns implied marathon finish time in hours."""
    pace = _daniels_gilbert_pace_from_vo2max(vo2max, pct_vo2max)
    finish_h = pace * _MARATHON_DIST_KM / 60.0
    return finish_h


def validate_population(
    population: list,
    show_plots: bool = True,
) -> dict:
    """
    Validate a generated base population against empirical references.

    Performs three validation levels:
      1. Marginal distribution checks (mean, SD) vs. published references.
      2. Correlation checks (VO2max-age, VO2max-gender).
      3. Implied marathon finish-time distribution vs. Dutch marathon data.

    Parameters
    ----------
    population : list of AdultParticipantProfile
        Output of generate_base_population().
    show_plots : bool
        If True, display a 3-panel validation figure.

    Returns
    -------
    dict
        Summary statistics and pass/fail flags for each check.
        Keys: 'marginal', 'correlation', 'finish_time', 'all_pass'.
    """
    # ------------------------------------------------------------------ #
    #  Extract arrays from population                                      #
    # ------------------------------------------------------------------ #
    heights     = np.array([p.height     for p in population])
    weights     = np.array([p.weight     for p in population])
    ages        = np.array([p.age        for p in population])
    genders     = np.array([p.gender     for p in population])
    vo2maxen    = np.array([p.vo2max     for p in population])
    pct_vo2maxen= np.array([p.pct_vo2max for p in population])
    bmis        = weights / heights ** 2
    gender_bin  = (genders == "male").astype(float)  # 1=male, 0=female

    n = len(population)
    pct_male = 100 * gender_bin.mean()

    # Implied finish times (hours) using Daniels-Gilbert
    finish_times_h = np.array([
        _finish_time_hours(p.vo2max, p.pct_vo2max) for p in population
    ])

    # ------------------------------------------------------------------ #
    #  1. MARGINAL VALIDATION                                              #
    # ------------------------------------------------------------------ #
    # Reference values from literature (see module docstring).
    # Format: {parameter: (sim_mean, sim_sd, ref_mean, ref_sd, ref_source)}
    MARGINAL_REFS = {
        "VO2max male (mL/kg/min)": (
            vo2maxen[genders == "male"].mean(),
            vo2maxen[genders == "male"].std(),
            52.0, 8.0,
            "Scharhag-Rosenberger (2010)"
        ),
        "VO2max female (mL/kg/min)": (
            vo2maxen[genders == "female"].mean(),
            vo2maxen[genders == "female"].std(),
            44.0, 7.0,
            "Scharhag-Rosenberger (2010)"
        ),
        "Age (years)": (
            ages.mean(), ages.std(),
            40.0, 10.0,
            "RunRepeat EU marathon stats (2023)"
        ),
        "% male": (
            pct_male, 0.0,   # SD not applicable for proportion
            58.0, 5.0,
            "RunRepeat EU marathon stats (2023)"
        ),
        "Height male (m)": (
            heights[genders == "male"].mean(),
            heights[genders == "male"].std(),
            1.78, 0.07,
            "RIVM NL antropometrische data"
        ),
        "Height female (m)": (
            heights[genders == "female"].mean(),
            heights[genders == "female"].std(),
            1.65, 0.06,
            "RIVM NL antropometrische data"
        ),
        "BMI (kg/m^2)": (
            bmis.mean(), bmis.std(),
            22.5, 2.5,
            "Tanda & Knechtle (2013)"
        ),
        "pct_VO2max at pace": (
            pct_vo2maxen.mean(), pct_vo2maxen.std(),
            # [rev13] Expected mean is now event-pace dependent, not fixed at 0.78.
            # For EVENT_PACE_MIN_PER_KM=6.5: VO2 demand ~45 mL/kg/min.
            # For population VO2max mean ~49 (mixed M/F): expected mean ~0.92 -> clipped lower.
            # Reference: Joyner (1991); Jones & Carter (2000).
            # Tolerance is wider (0.10) to accommodate event-pace variability.
            _daniels_gilbert_vo2_at_pace(EVENT_PACE_MIN_PER_KM) / 49.0, 0.10,
            "Daniels-Gilbert inversion at EVENT_PACE_MIN_PER_KM [rev13]"
        ),
    }

    marginal_results = {}
    print(f"\n{'='*72}")
    print(f"POPULATION VALIDATION  (N={n})")
    print(f"{'='*72}")
    print(f"{'Parameter':<35} {'Sim mean':>9} {'Sim SD':>8} {'Ref mean':>9} {'Ref SD':>8} {'Pass?':>6}")
    print(f"{'-'*72}")

    for param, (sim_m, sim_s, ref_m, ref_s, source) in MARGINAL_REFS.items():
        # Pass if simulated mean within 1 SD of reference mean
        deviation = abs(sim_m - ref_m)
        tolerance = ref_s if ref_s > 0 else ref_m * 0.10
        passed = deviation <= tolerance
        flag = f"{Fore.GREEN}OK{Style.RESET_ALL}" if passed else f"{Fore.RED}WARN{Style.RESET_ALL}"
        print(f"  {param:<33} {sim_m:>9.2f} {sim_s:>8.2f} {ref_m:>9.2f} {ref_s:>8.2f} {flag}")
        marginal_results[param] = {
            'sim_mean': sim_m, 'sim_sd': sim_s,
            'ref_mean': ref_m, 'ref_sd': ref_s,
            'passed': passed, 'source': source,
        }

    # ------------------------------------------------------------------ #
    #  2. CORRELATION VALIDATION                                           #
    # ------------------------------------------------------------------ #
    r_vo2_age    = float(np.corrcoef(vo2maxen, ages)[0, 1])
    r_vo2_gender = float(np.corrcoef(vo2maxen, gender_bin)[0, 1])
    cohens_d     = (
        (vo2maxen[gender_bin == 1].mean() - vo2maxen[gender_bin == 0].mean())
        / np.sqrt(
            ((vo2maxen[gender_bin == 1].std() ** 2) +
             (vo2maxen[gender_bin == 0].std() ** 2)) / 2
        )
    )

    # Expected: r(VO2max, age) in [-0.50, -0.25]; Cohen's d in [0.70, 1.00]
    corr_age_ok    = -0.50 <= r_vo2_age    <= -0.25
    cohens_d_ok    = 0.70  <= cohens_d     <= 1.00

    print(f"\n  Correlation checks:")
    print(f"  {'r(VO2max, age)':<30} {r_vo2_age:>8.3f}  "
          f"[expected: -0.50 to -0.25]  "
          f"{Fore.GREEN+'OK'+Style.RESET_ALL if corr_age_ok else Fore.RED+'WARN'+Style.RESET_ALL}")
    print(f"  {'Cohen d (VO2max, M vs F)':<30} {cohens_d:>8.3f}  "
          f"[expected:  0.70 to  1.00]  "
          f"{Fore.GREEN+'OK'+Style.RESET_ALL if cohens_d_ok else Fore.RED+'WARN'+Style.RESET_ALL}")

    correlation_results = {
        'r_vo2max_age':    {'value': r_vo2_age,  'passed': corr_age_ok},
        'cohens_d_gender': {'value': cohens_d,   'passed': cohens_d_ok},
    }

    # ------------------------------------------------------------------ #
    #  3. FINISH-TIME VALIDATION                                           #
    # ------------------------------------------------------------------ #
    # Reference: Dutch recreational marathon (Amsterdam, Rotterdam).
    # Expected finish-time distribution for self-selected mass events:
    #   Median:    [3h30m, 4h30m]  -> [3.50h, 4.50h]
    #   P5 (fast): [2h45m, 3h30m]  -> [2.75h, 3.50h]
    #   P95 (slow):[4h30m, 5h30m]  -> [4.50h, 5.50h]
    ft_p5   = float(np.percentile(finish_times_h, 5))
    ft_p50  = float(np.percentile(finish_times_h, 50))
    ft_p95  = float(np.percentile(finish_times_h, 95))
    ft_mean = float(finish_times_h.mean())

    ft_p50_ok = 3.50 <= ft_p50 <= 4.50
    ft_p5_ok  = 2.75 <= ft_p5  <= 3.50
    ft_p95_ok = 4.50 <= ft_p95 <= 5.50

    def _h_to_hhmm(h: float) -> str:
        hh = int(h)
        mm = int(round((h - hh) * 60))
        return f"{hh}h{mm:02d}m"

    print(f"\n  Implied marathon finish times (Daniels & Gilbert 1979):")
    print(f"  {'P5  (fast tail)':<30} {_h_to_hhmm(ft_p5):>8}  "
          f"[expected: 2h45m - 3h30m]  "
          f"{Fore.GREEN+'OK'+Style.RESET_ALL if ft_p5_ok else Fore.RED+'WARN'+Style.RESET_ALL}")
    print(f"  {'Median (P50)':<30} {_h_to_hhmm(ft_p50):>8}  "
          f"[expected: 3h30m - 4h30m]  "
          f"{Fore.GREEN+'OK'+Style.RESET_ALL if ft_p50_ok else Fore.RED+'WARN'+Style.RESET_ALL}")
    print(f"  {'P95 (slow tail)':<30} {_h_to_hhmm(ft_p95):>8}  "
          f"[expected: 4h30m - 5h30m]  "
          f"{Fore.GREEN+'OK'+Style.RESET_ALL if ft_p95_ok else Fore.RED+'WARN'+Style.RESET_ALL}")

    finish_time_results = {
        'ft_p5_h':  {'value': ft_p5,  'passed': ft_p5_ok},
        'ft_p50_h': {'value': ft_p50, 'passed': ft_p50_ok},
        'ft_p95_h': {'value': ft_p95, 'passed': ft_p95_ok},
    }

    all_marginal_ok     = all(v['passed'] for v in marginal_results.values())
    all_correlation_ok  = all(v['passed'] for v in correlation_results.values())
    all_finish_time_ok  = all(v['passed'] for v in finish_time_results.values())
    all_pass            = all_marginal_ok and all_correlation_ok and all_finish_time_ok

    print(f"\n  Overall: {'ALL CHECKS PASSED' if all_pass else 'ONE OR MORE CHECKS FAILED -- see warnings above'}")
    print(f"{'='*72}\n")

    # ------------------------------------------------------------------ #
    #  VALIDATION PLOTS (3-panel figure)                                   #
    # ------------------------------------------------------------------ #
    if show_plots:
        fig, axes = plt.subplots(1, 3, figsize=(18, 5))
        fig.suptitle(f"Population Validation (N={n})", fontsize=14, weight='bold')

        # Panel 1: VO2max by gender
        ax = axes[0]
        for g, color in [("male", "steelblue"), ("female", "salmon")]:
            data = vo2maxen[genders == g]
            ax.hist(data, bins=30, alpha=0.6, color=color, label=g.capitalize(),
                    density=True)
        ax.axvline(52, color='steelblue', linestyle='--', alpha=0.8, label='Ref male (52)')
        ax.axvline(44, color='salmon',    linestyle='--', alpha=0.8, label='Ref female (44)')
        ax.set_xlabel("VO2max (mL/kg/min)")
        ax.set_ylabel("Density")
        ax.set_title("VO2max Distribution by Gender")
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)

        # Panel 2: Correlation matrix (VO2max, age, BMI, pct_vo2max)
        ax = axes[1]
        corr_df = pd.DataFrame({
            'VO2max':      vo2maxen,
            'Age':         ages,
            'BMI':         bmis,
            '%VO2max':     pct_vo2maxen,
            'Male(0/1)':   gender_bin,
        })
        corr_matrix = corr_df.corr()
        im = ax.imshow(corr_matrix.values, cmap='RdBu_r', vmin=-1, vmax=1)
        ax.set_xticks(range(len(corr_matrix.columns)))
        ax.set_yticks(range(len(corr_matrix.columns)))
        ax.set_xticklabels(corr_matrix.columns, rotation=45, ha='right', fontsize=8)
        ax.set_yticklabels(corr_matrix.columns, fontsize=8)
        for i in range(len(corr_matrix)):
            for j in range(len(corr_matrix)):
                ax.text(j, i, f"{corr_matrix.values[i, j]:.2f}",
                        ha='center', va='center', fontsize=7,
                        color='black' if abs(corr_matrix.values[i, j]) < 0.5 else 'white')
        plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
        ax.set_title("Correlation Matrix")

        # Panel 3: Implied finish-time distribution
        ax = axes[2]
        ax.hist(finish_times_h, bins=40, color='forestgreen', alpha=0.7, density=True)
        ax.axvline(ft_p50, color='black', linestyle='-',  linewidth=2,
                   label=f'Median: {_h_to_hhmm(ft_p50)}')
        ax.axvline(ft_p5,  color='blue',  linestyle='--', linewidth=1.5,
                   label=f'P5: {_h_to_hhmm(ft_p5)}')
        ax.axvline(ft_p95, color='red',   linestyle='--', linewidth=1.5,
                   label=f'P95: {_h_to_hhmm(ft_p95)}')
        # Reference window shading
        ax.axvspan(3.50, 4.50, alpha=0.12, color='green', label='Ref median window')
        ax.set_xlabel("Implied finish time (hours)")
        ax.set_ylabel("Density")
        ax.set_title("Implied Marathon Finish Times\n(Daniels & Gilbert 1979)")
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)

        plt.tight_layout()
        plt.show(block=False)

    return {
        'marginal':     marginal_results,
        'correlation':  correlation_results,
        'finish_time':  finish_time_results,
        'all_pass':     all_pass,
    }


# =============================================================================
#  CHILDREN SIMULATION (unchanged from rev09)
# =============================================================================

def simulate_child_participant(n_simulations=200, met_value=None, age_group="11-20", fitness_factor=0.5):
    """
    Generates realistic variations for pediatric participants for Monte Carlo simulation.
    (Unchanged from rev09 -- pediatric population architecture not affected by rev10.)
    """
    valid_simulations = []
    attempts = 0
    max_attempts = n_simulations * 25

    if age_group == "0-10":
        height_mean, height_sd = 1.22, 0.08
        weight_mean, weight_sd = 23.0, 4.0
    else:
        height_mean, height_sd = 1.65, 0.1
        weight_mean, weight_sd = 55.0, 10.0

    print(f"\nGenerating realistic pediatric participant profiles for age group {age_group}...")
    while len(valid_simulations) < n_simulations and attempts < max_attempts:
        attempts += 1
        height = float(np.clip(np.random.normal(loc=height_mean, scale=height_sd),
                                height_mean - 0.3, height_mean + 0.3))
        weight = float(np.clip(np.random.normal(loc=weight_mean, scale=weight_sd),
                                weight_mean - 15.0, weight_mean + 15.0))
        bmi = weight / (height ** 2)
        if not (14.0 <= bmi <= 30.0):
            continue

        temp_variation = float(np.random.normal(loc=0, scale=1.0))
        rh_variation   = float(np.random.normal(loc=0, scale=8.0))

        met_efficiency_offset = (1 - fitness_factor) * met_value * 0.10
        met_variation = float(np.random.normal(loc=met_efficiency_offset,
                                                scale=0.15 * met_value))
        mf_score = float(np.random.uniform(0.0, 1.0))

        sweat_loc    = 0.5 + (fitness_factor * 0.2)
        sweat_factor = float(np.clip(np.random.normal(loc=sweat_loc, scale=0.1), 0.3, 0.9))

        thirst_upper_bound = 3.5 - (fitness_factor * 1.0)
        thirst_threshold   = float(np.random.uniform(2.0, max(2.1, thirst_upper_bound)))

        valid_simulations.append((height, weight, temp_variation, rh_variation,
                                   met_variation, mf_score, sweat_factor, thirst_threshold))

    if len(valid_simulations) < n_simulations:
        print(f"{Fore.YELLOW}Warning: Could only generate {len(valid_simulations)}/{n_simulations} "
              f"realistic simulations.{Style.RESET_ALL}")
    else:
        print(f"{Fore.GREEN}Successfully generated {len(valid_simulations)} realistic simulations.{Style.RESET_ALL}")
    return valid_simulations


# =============================================================================
#  COMMON ENVIRONMENTAL FUNCTIONS (unchanged from rev09)
# =============================================================================

@lru_cache(maxsize=128)
def get_lat_lon(city_name):
    """Retrieves lat/lon for a given city using Open-Meteo Geocoding API."""
    try:
        geo_url = (f"https://geocoding-api.open-meteo.com/v1/search"
                   f"?name={city_name}&count=1&format=json")
        response = requests.get(geo_url, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data and "results" in data and len(data["results"]) > 0:
            return data["results"][0]["latitude"], data["results"][0]["longitude"]
        return None, None
    except requests.RequestException as e:
        print(f"Error fetching location: {e}")
        return None, None


@lru_cache(maxsize=128)
def get_timezone(lat, lon):
    """Determines the IANA timezone for a given lat/lon."""
    tf = TimezoneFinder()
    timezone_name = tf.timezone_at(lat=lat, lng=lon)
    return pytz.timezone(timezone_name) if timezone_name else pytz.UTC


@lru_cache(maxsize=128)
def get_weather_forecast(lat, lon):
    """Retrieves a 7-day hourly weather forecast from Open-Meteo API."""
    try:
        forecast_url = (
            f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}"
            "&hourly=temperature_2m,relative_humidity_2m,wind_speed_10m,cloud_cover,surface_pressure"
            "&forecast_days=10&timezone=UTC"
        )
        response = requests.get(forecast_url, timeout=10)
        response.raise_for_status()
        data = response.json()
        if not data or "hourly" not in data:
            return None
        hourly = data["hourly"]
        owm_like_list = []
        for i in range(len(hourly["time"])):
            dt_str  = hourly["time"][i]
            dt_obj  = datetime.fromisoformat(dt_str).replace(tzinfo=UTC)
            dt_ts   = int(dt_obj.timestamp())
            wind_ms = hourly["wind_speed_10m"][i] / 3.6
            owm_like_list.append({
                "dt":   dt_ts,
                "main": {
                    "temp":     hourly["temperature_2m"][i],
                    "humidity": hourly["relative_humidity_2m"][i],
                    "pressure": hourly["surface_pressure"][i],
                },
                "wind":   {"speed": wind_ms},
                "clouds": {"all": hourly["cloud_cover"][i]},
            })
        return owm_like_list
    except requests.RequestException as e:
        print(f"Error fetching weather data: {e}")
        return None


def _map_us_aqi_to_owm_scale(us_aqi):
    if us_aqi <= 50:   return 1
    if us_aqi <= 100:  return 2
    if us_aqi <= 150:  return 3
    if us_aqi <= 200:  return 4
    return 5


@lru_cache(maxsize=128)
def get_air_quality(lat, lon):
    """Retrieves current AQI using Open-Meteo (mapped to 1-5 scale)."""
    try:
        air_url = (f"https://air-quality-api.open-meteo.com/v1/air-quality"
                   f"?latitude={lat}&longitude={lon}&current=us_aqi&timezone=UTC")
        response = requests.get(air_url, timeout=10)
        response.raise_for_status()
        data   = response.json()
        us_aqi = data["current"]["us_aqi"]
        return _map_us_aqi_to_owm_scale(us_aqi)
    except requests.RequestException:
        return 1


@lru_cache(maxsize=256)
def calculate_solar_radiance(lat, lon, local_datetime_str, cloud_cover, aqi):
    """Calculates GHI, DNI, DHI and solar position using pvlib."""
    try:
        local_timezone  = get_timezone(lat, lon)
        local_datetime  = pd.to_datetime(local_datetime_str).tz_convert(local_timezone.zone)
        location        = pvlib.location.Location(lat, lon, tz=local_timezone.zone, altitude=0)
        solar_position  = location.get_solarposition(pd.DatetimeIndex([local_datetime]))
        elevation       = solar_position['elevation'].iloc[0]
        zenith          = solar_position['apparent_zenith'].iloc[0]
        azimuth         = solar_position['azimuth'].iloc[0]
        linke_turbidity = 2.5 if aqi <= 2 else 4.0 if aqi <= 4 else 6.0
        clear_sky       = location.get_clearsky(local_datetime, model='ineichen',
                                                linke_turbidity=linke_turbidity)
        cloud_factor    = max(0.1, 1 - cloud_cover / 100)
        dni             = clear_sky['dni'].iloc[0] * cloud_factor
        dhi             = clear_sky['dhi'].iloc[0] * (1 + (1 - cloud_factor) * 0.5)
        ghi             = dni * pvlib.irradiance.aoi_projection(0, 0, zenith, azimuth) + dhi
        return {'ghi': max(0, ghi), 'dni': max(0, dni), 'dhi': max(0, dhi),
                'solar_elevation': elevation, 'zenith': zenith}
    except Exception as e:
        if "year is out of range" in str(e):
            return {'ghi': 0, 'dni': 0, 'dhi': 0, 'solar_elevation': 0, 'zenith': 90}
        print(f"Error calculating solar radiance for {local_datetime_str}: {e}")
        return {'ghi': 0, 'dni': 0, 'dhi': 0, 'solar_elevation': 0, 'zenith': 90}


def calculate_globe_temperature(dry_bulb_temp, ghi, wind_speed, solar_elevation,
                                pressure, cloud_cover, aqi):
    """
    Calculates black globe temperature (150mm) using energy balance (ISO 7726).
    Unchanged from rev09.
    """
    GLOBE_DIAMETER   = 0.15
    GLOBE_EMISSIVITY = 0.95
    SOLAR_ABSORPTIVITY = 0.95
    STEFAN_BOLTZMANN = 5.67e-8
    t_air_k = dry_bulb_temp + 273.15
    effective_wind = max(0.1, wind_speed)
    h_c = 6.3 * (effective_wind ** 0.6) / (GLOBE_DIAMETER ** 0.4)

    if ghi <= 0 or solar_elevation <= 0:
        sky_depression = 10 - (cloud_cover / 100) * 6
        if aqi > 3:
            sky_depression -= 2
        sky_temp_k    = t_air_k - sky_depression
        net_radiation = GLOBE_EMISSIVITY * STEFAN_BOLTZMANN * (sky_temp_k ** 4 - t_air_k ** 4)
        delta_t       = net_radiation / h_c
        t_globe_c     = dry_bulb_temp + delta_t
        return max(dry_bulb_temp - 3, min(dry_bulb_temp + 1, t_globe_c))

    solar_input   = SOLAR_ABSORPTIVITY * ghi / 4.0
    ambient_rad   = GLOBE_EMISSIVITY * STEFAN_BOLTZMANN * (t_air_k ** 4)
    total_in      = solar_input + ambient_rad
    h_r_initial   = 4 * GLOBE_EMISSIVITY * STEFAN_BOLTZMANN * (t_air_k ** 3)
    delta_t_init  = solar_input / (h_c + h_r_initial)
    t_globe_k     = t_air_k + delta_t_init

    for _ in range(20):
        rad_out   = GLOBE_EMISSIVITY * STEFAN_BOLTZMANN * (t_globe_k ** 4)
        conv_out  = h_c * (t_globe_k - t_air_k)
        imbalance = total_in - (rad_out + conv_out)
        if abs(imbalance) < 0.1:
            break
        deriv     = 4 * GLOBE_EMISSIVITY * STEFAN_BOLTZMANN * (t_globe_k ** 3) + h_c
        t_globe_k += 0.7 * imbalance / deriv
        if t_globe_k < t_air_k - 5:
            t_globe_k = t_air_k
        elif t_globe_k > t_air_k + 50:
            t_globe_k = t_air_k + 20

    t_globe_c = t_globe_k - 273.15
    if t_globe_c < dry_bulb_temp:
        t_globe_c = dry_bulb_temp
    max_delta = 15 if effective_wind > 2.0 else 20
    return min(t_globe_c, dry_bulb_temp + max_delta)


def calculate_mrt(globe_temp, dry_bulb_temp, wind_speed, ghi, solar_elevation):
    """Calculates Mean Radiant Temperature from globe temperature (ISO 7726). Unchanged."""
    GLOBE_DIAMETER   = 0.15
    GLOBE_EMISSIVITY = 0.95
    STEFAN_BOLTZMANN = 5.67e-8
    t_globe_k  = globe_temp + 273.15
    t_air_k    = dry_bulb_temp + 273.15
    eff_wind   = max(0.1, wind_speed)
    h_c        = 6.3 * (eff_wind ** 0.6) / (GLOBE_DIAMETER ** 0.4)
    conv_term  = (h_c / (GLOBE_EMISSIVITY * STEFAN_BOLTZMANN)) * (t_globe_k - t_air_k)
    mrt_k4     = (t_globe_k ** 4) + conv_term
    if mrt_k4 < 0:
        return globe_temp
    mrt_c = mrt_k4 ** 0.25 - 273.15
    if ghi > 0 and solar_elevation > 0:
        mrt_c = np.clip(mrt_c, globe_temp - 2, globe_temp + 15)
    else:
        mrt_c = np.clip(mrt_c, dry_bulb_temp - 15, dry_bulb_temp + 3)
    if mrt_c > 70 or mrt_c < -50:
        return dry_bulb_temp if ghi <= 0 else globe_temp
    return mrt_c


def interpolate_weather(weather_data, start_dt, end_dt, interval_minutes=10):
    """Interpolates raw hourly weather data to a consistent time-step. Unchanged."""
    times = pd.date_range(start=start_dt, end=end_dt, freq=f"{interval_minutes}min")
    times_ts        = np.array([t.timestamp() for t in times])
    weather_ts      = np.array([e["dt"]                for e in weather_data])
    temps           = np.array([e["main"]["temp"]      for e in weather_data])
    rhs             = np.array([e["main"]["humidity"]  for e in weather_data])
    winds           = np.array([e["wind"]["speed"]     for e in weather_data])
    clouds          = np.array([e["clouds"]["all"]     for e in weather_data])
    pressures       = np.array([e["main"]["pressure"]  for e in weather_data])
    interp_temps    = np.interp(times_ts, weather_ts, temps)
    interp_rhs      = np.interp(times_ts, weather_ts, rhs)
    interp_winds    = np.interp(times_ts, weather_ts, winds)
    interp_clouds   = np.interp(times_ts, weather_ts, clouds)
    interp_pressures= np.interp(times_ts, weather_ts, pressures)
    return [
        {
            "time":     pd.Timestamp.fromtimestamp(ts, tz=start_dt.tz),
            "temp":     temp,
            "rh":       rh,
            "wind":     wind,
            "clouds":   cloud,
            "pressure": pressure,
            "twb":      wet_bulb_tmp(tdb=temp, rh=rh),
        }
        for ts, temp, rh, wind, cloud, pressure in zip(
            times_ts, interp_temps, interp_rhs, interp_winds,
            interp_clouds, interp_pressures
        )
    ]


# =============================================================================
#  RPE MODELS (unchanged from rev09)
# =============================================================================

def calculate_adult_rpe_improved(current_met, t_rect, training_factor,
                                  acclimatization_factor, mf_score):
    """
    Borg 6-20 RPE for adults. Piecewise-linear MET->RPE with heat penalty.
    Unchanged from rev09.
    """
    if current_met <= 2.0:
        base_rpe = 6.0 + (current_met - 1.0) * 2.0
    elif current_met <= 6.0:
        base_rpe = 8.0 + (current_met - 2.0) * 1.0
    elif current_met <= 10.0:
        base_rpe = 12.0 + (current_met - 6.0) * 0.75
    elif current_met <= 14.0:
        base_rpe = 15.0 + (current_met - 10.0) * 0.75
    else:
        base_rpe = 18.0 + min((current_met - 14.0) * 0.4, 2.0)
    training_reduction = training_factor * 0.15
    adjusted_base      = base_rpe * (1.0 - training_reduction)
    if t_rect > 38.0:
        heat_penalty = (t_rect - 38.0) * 1.8 * (1.0 - acclimatization_factor * 0.4)
    else:
        heat_penalty = 0.0
    rpe_physical       = np.clip(adjusted_base + heat_penalty, 6.0, 20.0)
    motivation_modifier= (mf_score - 0.5) * 2.0
    motivation_effect  = motivation_modifier * (20.0 - rpe_physical) * 0.15
    rpe_total          = np.clip(rpe_physical + motivation_effect, 6.0, 20.0)
    return rpe_physical, rpe_total


def calculate_pediatric_rpe_improved(current_met, t_core, fitness_factor,
                                      mf_score, age_group):
    """OMNI 0-10 RPE for children. Unchanged from rev09."""
    if age_group == "0-10":
        max_met_capacity  = 10.5
        thermal_threshold = 38.3
    else:
        max_met_capacity  = 13.0
        thermal_threshold = 38.2
    rel = np.clip(current_met / max_met_capacity, 0.0, 1.0)
    if rel <= 0.3:
        base_rpe = rel * 10.0
    elif rel <= 0.6:
        base_rpe = 3.0 + (rel - 0.3) * 11.0
    elif rel <= 0.85:
        base_rpe = 6.3 + (rel - 0.6) * 10.8
    else:
        base_rpe = 9.0 + (rel - 0.85) * 6.7
    fitness_reduction = fitness_factor * 0.20
    adjusted_base     = base_rpe * (1.0 - fitness_reduction)
    heat_penalty      = (t_core - thermal_threshold) * 1.2 if t_core > thermal_threshold else 0.0
    rpe_physical      = np.clip(adjusted_base + heat_penalty, 0.0, 10.0)
    motivation_modifier = (mf_score - 0.5) * 2.0
    motivation_effect   = motivation_modifier * (10.0 - rpe_physical) * 0.25
    rpe_total           = np.clip(rpe_physical + motivation_effect, 0.0, 10.0)
    return rpe_physical, rpe_total


# =============================================================================
#  CORE ADULT SIMULATION WORKER (rev10 changes marked [rev10]; rev11 marked [rev11];
#                                  rev13 changes marked [rev13])
#
#  CHANGE SUMMARY vs rev12
#  -----------------------
#  [rev13-P4] Dynamic pacing: current_met is updated each time step based on
#  the previous T_rect (closed pacing loop). The open-loop constant-MET
#  assumption is replaced by a proportional controller:
#
#    thermal_overshoot  = max(0, T_rect[t-1] - T_RECT_PACING_THRESHOLD)
#    met_reduction      = K_P_PACING * thermal_overshoot
#    current_met[t]     = max(1.5, met_initial - training_offset - met_reduction)
#
#  met_initial = base_met_from_vo2max (unchanged from rev12 pre-loop computation)
#  training_offset = training_factor * base_met_from_vo2max * 0.05 (rev11, unchanged)
#  K_P_PACING = 0.10 MET/°C (Ely 2007; module constant)
#  T_RECT_PACING_THRESHOLD = 38.5°C (module constant)
#
#  The training_offset represents running economy and is kept outside the
#  thermal feedback term -- economy and voluntary pacing are orthogonal.
#  current_met is stored per time step in results for diagnostics.
#
#  [rev13-P5] AUC_thermisch: cumulative integral of T_rect > 39.0°C (°C·min)
#  is accumulated in the loop and stored in results. This implements
#  Breslow's time-integrated thermal dose concept. T_AUC_THRESHOLD = 39.0°C
#  (module constant).
#
#  CHANGE SUMMARY vs rev09-rev12 (retained from earlier revisions)
#  ----------------------------------------------------------------
#  The met_variation field has been removed from AdultParticipantProfile.
#  MET is now computed from VO2max and pct_vo2max:
#
#    base_met_from_vo2max = participant.vo2max * participant.pct_vo2max
#                           / VO2MAX_TO_MET_FACTOR
#
#  A training_factor deterministic shift is then applied  [rev11: REVISED]:
#    training_offset = training_factor * base_met_from_vo2max * 0.05
#    final_met       = base_met_from_vo2max - training_offset
#
#  TRAINING_FACTOR DOUBLE-CORRECTION (rev11 resolution)
#  -----------------------------------------------------
#  Rev10 applied training_factor in two independent places:
#    1. generate_base_population(): mu of pct_vo2max -= training_factor * 0.08
#    2. here: MET further reduced by training_factor * base_met * 0.15
#
#  Combined effect at tf=1.0 was:
#    effective %VO2max = (0.80 - 0.08) * (1 - 0.15) = 0.72 * 0.85 = 0.61
#
#  This is too low for a competitive marathon population. Rev11 reduces the
#  step-2 coefficient from 0.15 to 0.05, giving:
#    tf=0.0: effective pct ≈ 0.80 * 1.00 = 0.80
#    tf=0.5: effective pct ≈ 0.76 * 0.975 = 0.74
#    tf=1.0: effective pct ≈ 0.72 * 0.95  = 0.68
#
#  These values are consistent with the physiological literature for
#  recreational marathon runners (Joyner 1991; Jones & Carter 2000).
#  Clinical calibration from Bongers (Radboudumc) is still needed to
#  determine the precise coefficient value.
#
#  The base_met parameter (user-selected activity MET) is NO LONGER used
#  to drive the individual simulation. It is retained as a reference value
#  for the population mean validation and for the MET threshold analysis
#  (unchanged). Each participant's MET is now fully determined by their
#  own VO2max and pct_vo2max, making cross-scenario comparisons valid.
# =============================================================================

def simulate_post_finish(t_rect_finish: float,
                          co_reserve_finish: float,
                          auc_klinisch_race: float,
                          wbgt_finish: float) -> dict:
    """
    Propagate T_rect and CO_reserve for PF_DUUR_MIN minutes after the finish.

    After crossing the finish line:
      1. The pacing controller (K_p) is no longer active -- open loop.
      2. Metabolic after-glow: residual heat production decays exponentially.
         Peak amplitude PF_DELTA_T_NAGLOED, time constant PF_TAU_NAGLOED.
      3. Venous pooling: muscle pump stops, CO drops, cutaneous heat
         dissipation falls. CO_reserve decreases acutely.
      4. Passive cooling from sweating continues but at reduced rate because
         cutaneous blood flow is compromised by pooling.

    EHS criterion post-finish:
        T_rect > 40.5 deg C  AND  CO_reserve <= 0

    Parameters
    ----------
    t_rect_finish      : float  -- T_rect at moment of finishing (deg C)
    co_reserve_finish  : float  -- CO_reserve at moment of finishing
    auc_klinisch_race  : float  -- AUC_klinisch accumulated during race
    wbgt_finish        : float  -- WBGT at finish (deg C), used for passive cooling

    Returns
    -------
    dict with keys:
        t_rect_piek_postfinish   : float  -- peak T_rect in finish zone
        co_reserve_postfinish    : float  -- CO_reserve at end of window
        auc_klinisch_postfinish  : float  -- additional AUC_klinisch post-finish
        auc_klinisch_totaal      : float  -- race + post-finish combined
        ehs_postfinish           : bool   -- EHS criterion met in finish zone
        t_ehs_postfinish_min     : float  -- minutes after finish at EHS (nan if not)

    References
    ----------
    Roberts WO (1998) Exertional heat stroke during a cool weather marathon.
    Rowell LB (1974) Human cardiovascular adjustments to exercise and heat stress.
    """
    # Passive cooling rate depends on WBGT -- higher WBGT means less cooling
    # At WBGT=25 the passive cooling is negligible; at WBGT=15 it is ~0.05 deg C/min
    koeling = max(0.0, PF_KOELING_PER_MIN * (1.0 - (wbgt_finish - 15.0) / 20.0))

    t_rect    = t_rect_finish
    co_res    = co_reserve_finish
    auc_kl_pf = 0.0
    t_rect_piek = t_rect_finish
    ehs_pf    = False
    t_ehs_pf  = float('nan')

    n_stappen = int(PF_DUUR_MIN / PF_DT_MIN)

    for stap in range(n_stappen):
        t = stap * PF_DT_MIN   # minutes after finish

        # Metabolic after-glow: exponential decay
        warmte_in  = PF_DELTA_T_NAGLOED * np.exp(-t / PF_TAU_NAGLOED) * PF_DT_MIN

        # Net temperature change
        t_rect += warmte_in - koeling * PF_DT_MIN
        t_rect  = max(36.5, t_rect)   # floor: cannot drop below resting

        # Venous pooling effect on CO_reserve
        #pooling_rate = (PF_DELTA_CO_POOLING * co_reserve_finish *
                        #np.exp(-t / PF_TAU_POOLING))
                        
        
        pooling_rate = (PF_DELTA_CO_POOLING * max(0.0, co_res) * 
                        np.exp(-t / PF_TAU_POOLING))
        
        co_res -= pooling_rate * PF_DT_MIN
        co_res  = max(-5.0, co_res)   # allow negative (clinical deficit)

        # Track peak T_rect
        if t_rect > t_rect_piek:
            t_rect_piek = t_rect

        # AUC_klinisch accumulation post-finish
        auc_kl_pf += max(0.0, t_rect - T_AUC_KLINISCH) * PF_DT_MIN

        # EHS criterion check
        if t_rect > 40.5 and co_res <= 0 and not ehs_pf:
            
        #if t_rect > 40.5 and co_res <= 0 and co_reserve_finish > 0 and not ehs_pf:    
            ehs_pf   = True
            t_ehs_pf = t + PF_DT_MIN

        # Early termination if T_rect safely below threshold
        if t_rect < 39.0 and t > 3.0:
            break

    return {
        't_rect_piek_postfinish'  : round(t_rect_piek, 4),
        'co_reserve_postfinish'   : round(co_res, 4),
        'auc_klinisch_postfinish' : round(auc_kl_pf, 3),
        'auc_klinisch_totaal'     : round(auc_klinisch_race + auc_kl_pf, 3),
        'ehs_postfinish'          : ehs_pf,
        't_ehs_postfinish_min'    : t_ehs_pf,
    }


def calculate_indices_jos3_adult(interp_data, lat, lon, met_value, clo_value,
                                  participant_params, training_factor,
                                  acclimatization_factor):
    """
    Run the core JOS-3 simulation for a single virtual adult participant.

    [rev10] MET computation change:
    MET is no longer taken from participant_params.met_variation (removed).
    Instead it is derived from VO2max and pct_vo2max:
        base_met = vo2max * pct_vo2max / VO2MAX_TO_MET_FACTOR
        final_met = base_met * (1 - training_factor * 0.05)  [rev11: was 0.15]

    [rev11] training_offset coefficient reduced from 0.15 to 0.05.
    See CORE ADULT SIMULATION WORKER block above for full rationale.
    The effective %VO2max at the simulation MET level is now:
        tf=0.0 -> ~0.80; tf=0.5 -> ~0.74; tf=1.0 -> ~0.68.

    The met_value argument is retained for backward compatibility but is
    only used by the liveability check (Vanos et al. 2023), not as the
    individual simulation MET. This is intentional: liveability limits are
    defined for the EVENT pace (user-selected MET), not the individual.

    Parameters
    ----------
    interp_data : list of dicts
    lat, lon : float
    met_value : float
        Base MET of the chosen activity (reference only; individual MET
        is derived from vo2max * pct_vo2max / VO2MAX_TO_MET_FACTOR).
    clo_value : float
    participant_params : AdultParticipantProfile
    training_factor : float
    acclimatization_factor : float

    Returns
    -------
    list of dicts -- one per time step
    """
    # Extract fields (all by name -- no positional indexing)
    height           = participant_params.height
    weight           = participant_params.weight
    runner_age       = participant_params.age
    gender           = participant_params.gender
    vo2max           = participant_params.vo2max
    pct_vo2max       = participant_params.pct_vo2max   # [rev10] replaces met_variation
    temp_var         = participant_params.temp_variation
    rh_var           = participant_params.rh_variation
    mf_score         = participant_params.mf_score
    sweat_factor     = participant_params.sweat_factor
    thirst_threshold = participant_params.thirst_threshold

    # ------------------------------------------------------------------
    # [rev10] INDIVIDUAL MET -- derived from VO2max and pct_vo2max
    #
    # This guarantees %VO2max == pct_vo2max by construction.
    # The training_offset shifts MET downward for better-trained runners
    # (improved economy: same %VO2max, lower absolute energy cost per kg).
    #
    # [rev11] Coefficient reduced from 0.15 to 0.05 to avoid compounding
    # with the pct_vo2max mu-shift already applied in generate_base_population().
    # Net effective %VO2max: tf=0.0 -> ~0.80; tf=0.5 -> ~0.74; tf=1.0 -> ~0.68.
    # ------------------------------------------------------------------
    base_met_from_vo2max = vo2max * pct_vo2max / VO2MAX_TO_MET_FACTOR
    training_offset      = training_factor * base_met_from_vo2max * 0.05
    current_met          = max(1.5, base_met_from_vo2max - training_offset)

    # [rev13-P4] Store initial MET for dynamic pacing reference.
    met_initial = current_met

    # [rev14-A] Individual pacing gain from participant profile.
    # Falls back to module constant K_P_PACING if field absent (backwards compat).
    kp_ind = getattr(participant_params, 'kp_pacing', K_P_PACING)

    # [rev13-P5] Initialise AUC_thermisch accumulator (°C·min, diagnostic).
    auc_thermisch = 0.0
    # [rev14-B] Initialise AUC_klinisch accumulator (Roberts 2007, clinical).
    auc_klinisch  = 0.0

    # Environment
    times      = [entry["time"] for entry in interp_data]
    temps      = [entry["temp"] + temp_var for entry in interp_data]
    wind_10m   = [entry["wind"] for entry in interp_data]
    wind_1_5m  = [wind_speed_at_height(w, 10, 1.5) for w in wind_10m]
    rhs        = [np.clip(entry["rh"] + rh_var, 10, 95) for entry in interp_data]
    clouds     = [entry["clouds"] for entry in interp_data]
    pressures  = [entry["pressure"] for entry in interp_data]

    aqi             = get_air_quality(lat, lon)
    radiances       = [calculate_solar_radiance(lat, lon, str(t), c, aqi)
                        for t, c in zip(times, clouds)]
    ghi             = [rad['ghi'] for rad in radiances]
    solar_elevations= [rad['solar_elevation'] for rad in radiances]
    globe_temps     = [calculate_globe_temperature(t, g, w, se, p, c, aqi)
                        for t, g, w, se, p, c in zip(
                            temps, ghi, wind_1_5m, solar_elevations, pressures, clouds)]
    mrts            = [calculate_mrt(gt, t, w, g, se)
                        for gt, t, w, g, se in zip(
                            globe_temps, temps, wind_1_5m, ghi, solar_elevations)]

    # JOS-3 setup
    fat_percentage = 25 if gender == "female" else 15
    jos3_model     = JOS3(height=height, weight=weight, age=runner_age,
                          fat=fat_percentage, sex=gender)
    jos3_model.t_core[:] = 37.0

    time_steps_minutes = [(times[i+1] - times[i]).total_seconds() / 60.0
                           for i in range(len(times) - 1)]
    time_steps_minutes.insert(0, 0)

    # CVR initialisation
    if CVR_BESCHIKBAAR:
        cvr_runner = RunnerProfile(
            mass=weight, height=height * 100,
            age=runner_age, sex=gender, vo2max=vo2max
        )
        jos3_cvr_reeks          = []
        cvr_gewichtsverlies_kg  = 0.0

    results                 = []
    cumulative_water_loss   = 0
    runner_stopped          = False
    can_drink_again         = True

    for i in range(len(times)):
        if runner_stopped and i > 0:
            results.append({**results[-1],
                            'time': times[i].strftime('%Y-%m-%d %H:%M'),
                            'stopped': True})
            continue

        # Liveability check uses the EVENT MET (met_value), not individual
        m_max_threshold = get_liveability_threshold(temps[i], rhs[i], runner_age)
        is_unliveable   = met_value > m_max_threshold

        # [rev13-P4] DYNAMIC PACING: update current_met based on previous T_rect.
        # At i==0 (initialisation step) no T_rect is available yet; use met_initial.
        # From i==1 onward, the previous step's t_rect drives voluntary pace reduction.
        # Ref: Ely et al. (2007) Med Sci Sports Exerc 39:1949-1955.
        if i > 0:
            prev_t_rect       = results[-1]['t_rect']
            thermal_overshoot = max(0.0, prev_t_rect - T_RECT_PACING_THRESHOLD)
            met_reduction     = kp_ind * thermal_overshoot   # [rev14-A] individual K_p
            current_met       = max(1.5, met_initial - met_reduction)

        # JOS-3 step
        jos3_model.par = current_met
        jos3_model.tdb = temps[i]
        jos3_model.tr  = mrts[i]
        jos3_model.rh  = rhs[i]
        jos3_model.v   = max(0.1, wind_1_5m[i])
        jos3_model.clo = clo_value
        if i > 0:
            jos3_model.simulate(times=int(round(time_steps_minutes[i])), dtime=60)

        t_rect = jos3_model.t_core.mean() + 0.5  # rectal offset (Rae 2008)

        # [rev13-P5] AUC_thermisch: accumulate time-integrated T_rect exceedance.
        # Integration floor T_AUC_THRESHOLD = 39.0°C (module constant).
        # dt_min = 0 at i==0 (initialisation step, no elapsed time).
        if i > 0:
            dt_min         = time_steps_minutes[i]
            auc_thermisch += max(0.0, t_rect - T_AUC_THRESHOLD)  * dt_min
            auc_klinisch  += max(0.0, t_rect - T_AUC_KLINISCH)   * dt_min  # [rev14-B]

        results_jos3 = jos3_model.dict_results()

        if 'weight_loss_by_evap_and_res' in results_jos3:
            water_loss_rate_g_per_s = float(np.mean(results_jos3['weight_loss_by_evap_and_res']))
        else:
            metabolic_heat_flux     = jos3_model.par * 58.2
            evap_fraction           = 0.6
            bsa                     = float(np.mean(jos3_model.bsa))
            water_loss_rate_g_per_s = (metabolic_heat_flux * bsa * evap_fraction) / 2418.0

        if CVR_BESCHIKBAAR and i > 0:
            dt_s = time_steps_minutes[i] * 60.0
            cvr_gewichtsverlies_kg += water_loss_rate_g_per_s * sweat_factor * dt_s / 1000.0

            co_lh = float(np.array(results_jos3['cardiac_output']).flat[-1]) \
                if 'cardiac_output' in results_jos3 else 900.0

            if 'bf_skin' in results_jos3:
                bfsk_arr = np.array(results_jos3['bf_skin'])
                bf_skin_tot = float(bfsk_arr[-1].sum() if bfsk_arr.ndim == 2
                                    else bfsk_arr.sum())
            else:
                bf_skin_tot = co_lh * 0.25

            bf_ava_h = float(np.array(results_jos3['bf_ava_hand']).flat[-1]) \
                if 'bf_ava_hand' in results_jos3 else 0.5
            bf_ava_f = float(np.array(results_jos3['bf_ava_foot']).flat[-1]) \
                if 'bf_ava_foot' in results_jos3 else 0.5
            t_cb_val = float(np.array(results_jos3['t_cb']).flat[-1]) \
                if 't_cb' in results_jos3 else jos3_model.t_core.mean()

            jos3_cvr_reeks.append(JOS3Outputs(
                t_min          = i,
                cardiac_output = co_lh,
                t_core_mean    = jos3_model.t_core.mean(),
                t_cb           = t_cb_val,
                weight_loss_g_s= cvr_gewichtsverlies_kg,
                bf_skin_total  = bf_skin_tot,
                bf_ava_hand    = bf_ava_h,
                bf_ava_foot    = bf_ava_f,
            ))

        sweat_rate     = water_loss_rate_g_per_s * 60.0 * sweat_factor
        water_loss     = sweat_rate * time_steps_minutes[i] if i > 0 else 0
        cumulative_water_loss += water_loss
        water_loss_pct = (cumulative_water_loss / 1000) / weight * 100
        is_thirsty     = water_loss_pct >= thirst_threshold

        if is_thirsty and can_drink_again:
            cumulative_water_loss = max(0, cumulative_water_loss - np.random.uniform(120, 180))
            can_drink_again       = False
        if not is_thirsty:
            can_drink_again = True

        utci_val = utci(tdb=temps[i], tr=mrts[i], v=wind_10m[i], rh=rhs[i]).utci
        with_solar = bool(ghi[i] > 0 and solar_elevations[i] > 0)
        wbgt_args = {"twb": wet_bulb_tmp(tdb=temps[i], rh=rhs[i]),
                     "tg": globe_temps[i],
                     "with_solar_load": with_solar,
                     "round_output": True}
        if with_solar:
            wbgt_args["tdb"] = temps[i]
        wbgt_val = float(wbgt(**wbgt_args).wbgt)

        rpe_physical, rpe_total = calculate_adult_rpe_improved(
            current_met=current_met,
            t_rect=t_rect,
            training_factor=training_factor,
            acclimatization_factor=acclimatization_factor,
            mf_score=mf_score,
        )

        if rpe_total >= 19.5:
            runner_stopped = True

        ehbo_visit = (t_rect >= 40.5) or (water_loss_pct >= 2.0) or (rpe_total >= 17)

        results.append({
            "time":           times[i].strftime('%Y-%m-%d %H:%M'),
            "utci":           utci_val,
            "wbgt":           wbgt_val,
            "t_rect":         t_rect,
            "water":          cumulative_water_loss,
            "mrt":            mrts[i],
            "rpe_physical":   rpe_physical,
            "rpe_total":      rpe_total,
            "stopped":        runner_stopped,
            "ehbo_visit":     ehbo_visit,
            "is_unliveable":  is_unliveable,
            "current_met":    current_met,        # [rev13-P4] dynamic MET
            "auc_thermisch":  auc_thermisch,      # [rev13-P5] °C·min above 39.0°C
            "auc_klinisch":   auc_klinisch,       # [rev14-B] °C·min above 40.5°C
            "kp_ind":         kp_ind,             # [rev14-A] individual pacing gain
            # CVR fields (filled post-loop)
            "hr_geschat":     float('nan'),
            "cvs_index":      float('nan'),
            "co_reserve":     float('nan'),
            "decompensatie":  False,
        })

    if CVR_BESCHIKBAAR and len(jos3_cvr_reeks) > 0:
        cvr_ts = koppel_cvr_aan_jos3(cvr_runner, jos3_cvr_reeks)
        for j, cvr_state in enumerate(cvr_ts.states):
            idx = j + 1
            if idx < len(results):
                results[idx]['hr_geschat']    = cvr_state.HR
                results[idx]['cvs_index']     = cvr_state.CVS_index
                results[idx]['co_reserve']    = cvr_state.CO_reserve
                results[idx]['decompensatie'] = cvr_state.decompensatie

    # [rev14-C] POST-FINISH MODULE
    # Propagate T_rect and CO_reserve for PF_DUUR_MIN minutes after finish.
    # Uses final simulation state as starting point.
    t_rect_finish    = results[-1]['t_rect']
    co_reserve_finish = results[-1].get('co_reserve', 2.0)
    if np.isnan(co_reserve_finish):
        co_reserve_finish = 2.0   # fallback if CVR not available
    wbgt_finish       = results[-1]['wbgt']
    auc_klinisch_race = results[-1]['auc_klinisch']

    pf = simulate_post_finish(
        t_rect_finish     = t_rect_finish,
        co_reserve_finish = co_reserve_finish,
        auc_klinisch_race = auc_klinisch_race,
        wbgt_finish       = wbgt_finish,
    )
    # Attach post-finish results to last entry for easy extraction
    results[-1].update({
        't_rect_piek_postfinish'  : pf['t_rect_piek_postfinish'],
        'co_reserve_postfinish'   : pf['co_reserve_postfinish'],
        'auc_klinisch_postfinish' : pf['auc_klinisch_postfinish'],
        'auc_klinisch_totaal'     : pf['auc_klinisch_totaal'],
        'ehs_postfinish'          : pf['ehs_postfinish'],
        't_ehs_postfinish_min'    : pf['t_ehs_postfinish_min'],
    })

    return results


# =============================================================================
#  PEDIATRIC SIMULATION WORKER (unchanged from rev09)
# =============================================================================

def run_pediatric_simulation(interp_data, lat, lon, met_value, clo_value,
                              participant_params, age_group="11-20", fitness_factor=0.5):
    """Single pediatric participant simulation. Unchanged from rev09."""
    height, weight, temp_var, rh_var, met_var, mf_score, sweat_factor, thirst_threshold = participant_params

    if age_group == "0-10":
        age            = 7
        fat_percentage = 18
        gender         = 'male'
    else:
        age            = 15
        gender         = np.random.choice(["male", "female"])
        fat_percentage = 22 if gender == 'female' else 15

    times      = [entry["time"] for entry in interp_data]
    temps      = [entry["temp"] + temp_var for entry in interp_data]
    wind_10m   = [entry["wind"] for entry in interp_data]
    wind_1_5m  = [wind_speed_at_height(w, 10, 1.0) for w in wind_10m]
    rhs        = [np.clip(entry["rh"] + rh_var, 10, 95) for entry in interp_data]
    clouds     = [entry["clouds"] for entry in interp_data]
    pressures  = [entry["pressure"] for entry in interp_data]

    aqi             = get_air_quality(lat, lon)
    radiances       = [calculate_solar_radiance(lat, lon, str(t), c, aqi)
                        for t, c in zip(times, clouds)]
    ghi             = [rad['ghi'] for rad in radiances]
    solar_elevations= [rad['solar_elevation'] for rad in radiances]
    globe_temps     = [calculate_globe_temperature(t, g, w, se, p, c, aqi)
                        for t, g, w, se, p, c in zip(
                            temps, ghi, wind_1_5m, solar_elevations, pressures, clouds)]
    mrts            = [calculate_mrt(gt, t, w, g, se)
                        for gt, t, w, g, se in zip(
                            globe_temps, temps, wind_1_5m, ghi, solar_elevations)]

    jos3_model = JOS3(height=height, weight=weight, age=age,
                       fat=fat_percentage, sex=gender)
    jos3_model.t_core[:] = 37.0
    time_steps_minutes = [(times[i+1] - times[i]).total_seconds() / 60.0
                           for i in range(len(times) - 1)]
    time_steps_minutes.insert(0, 0)

    results               = []
    cumulative_water_loss = 0
    participant_stopped   = False
    can_drink_again       = True
    current_met           = met_value + met_var

    for i in range(len(times)):
        if participant_stopped and i > 0:
            results.append({**results[-1],
                            'time': times[i].strftime('%Y-%m-%d %H:%M'),
                            'stopped': True})
            continue

        jos3_model.par = current_met
        jos3_model.tdb = temps[i]
        jos3_model.tr  = mrts[i]
        jos3_model.rh  = rhs[i]
        jos3_model.v   = max(0.1, wind_1_5m[i])
        jos3_model.clo = clo_value
        if i > 0:
            jos3_model.simulate(times=int(round(time_steps_minutes[i])), dtime=60)
        t_core_avg = jos3_model.t_core.mean()

        sweat_rate_ml = sweat_factor * (jos3_model.par * 58.2 * float(np.mean(jos3_model.bsa))) / 2418 * 1000 / 60
        water_loss    = sweat_rate_ml * time_steps_minutes[i] if i > 0 else 0
        cumulative_water_loss += water_loss
        water_loss_pct = (cumulative_water_loss / 1000) / weight * 100
        is_thirsty     = water_loss_pct >= thirst_threshold

        if is_thirsty and can_drink_again:
            cumulative_water_loss = max(0, cumulative_water_loss - np.random.uniform(100, 150))
            can_drink_again       = False
        if not is_thirsty:
            can_drink_again = True

        utci_val   = utci(tdb=temps[i], tr=mrts[i], v=wind_10m[i], rh=rhs[i]).utci
        with_solar = bool(ghi[i] > 0 and solar_elevations[i] > 0)
        wbgt_args  = {"twb": wet_bulb_tmp(tdb=temps[i], rh=rhs[i]),
                      "tg": globe_temps[i],
                      "with_solar_load": with_solar,
                      "round_output": True}
        if with_solar:
            wbgt_args["tdb"] = temps[i]
        wbgt_val = float(wbgt(**wbgt_args).wbgt)

        rpe_physical, rpe_total = calculate_pediatric_rpe_improved(
            current_met=current_met, t_core=t_core_avg,
            fitness_factor=fitness_factor, mf_score=mf_score, age_group=age_group)

        if rpe_total >= 9.5:
            participant_stopped = True

        first_aid = (t_core_avg >= 38.8) or (water_loss_pct >= 3.0) or (rpe_total >= 8.5)

        results.append({
            "time": times[i].strftime('%Y-%m-%d %H:%M'), "utci": utci_val, "wbgt": wbgt_val,
            "t_core": t_core_avg, "water_loss_ml": cumulative_water_loss, "mrt": mrts[i],
            "rpe_physical": rpe_physical, "rpe_total": rpe_total,
            "stopped": participant_stopped, "first_aid_visit": first_aid,
        })
    return results


# =============================================================================
#  PARALLEL WORKER WRAPPERS (unchanged)
# =============================================================================

def worker_monte_carlo_adult(args):
    """Worker for a single adult Monte Carlo simulation in parallel."""
    import sys, os
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    if _script_dir not in sys.path:
        sys.path.insert(0, _script_dir)
    interp_data, lat, lon, met_value, clo_value, params, training, acclimatization = args
    return calculate_indices_jos3_adult(
        interp_data, lat, lon, met_value, clo_value, params, training, acclimatization
    )


def worker_monte_carlo_pediatric(args):
    """Worker for a single pediatric Monte Carlo simulation in parallel."""
    interp_data, lat, lon, met_value, clo_value, params, age_group, fitness_factor = args
    return run_pediatric_simulation(
        interp_data, lat, lon, met_value, clo_value, params, age_group, fitness_factor
    )


# =============================================================================
#  MONTE CARLO MANAGERS
# =============================================================================

def _calculate_threshold_stats(temp_sims, n_simulations):
    """Calculates mean % exceeding each T_rect threshold with 95% CI (Wald)."""
    thresholds = [38.0, 38.5, 39.0, 39.5, 40.0, 40.5, 41.0]
    stats = {}
    for t in thresholds:
        p          = np.mean(temp_sims > t, axis=0)
        se         = np.sqrt(p * (1 - p) / n_simulations) if n_simulations > 0 else 0
        ci_margin  = 1.96 * se
        stats[f"{t:.1f}"] = {
            'mean':     p * 100,
            'ci_lower': np.maximum(0, p - ci_margin) * 100,
            'ci_upper': np.minimum(1, p + ci_margin) * 100,
        }
    return stats


# =============================================================================
#  [rev15] CALIBRATION UNCERTAINTY FUNCTIONS
# =============================================================================

def _intercept_sensitivity(z_zonder_intercept: np.ndarray,
                            intercept_nominal: float,
                            delta: float = 0.5) -> dict:
    """
    Computes key population percentages at intercept_nominal ± delta.

    The delta of 0.5 log-odds units is a conservative estimate of the
    calibration transfer uncertainty from Boston Marathon (Breslow 2021)
    to DtD conditions.  It covers:
      - Between-event EHS rate variation (~2-fold in the literature).
      - JOS-3 systematic T_rect offset vs. field measurements (~0.3°C).
      - training_offset coefficient uncertainty (±0.05, pending Bongers).

    Parameters
    ----------
    z_zonder_intercept : np.ndarray, shape (n_sim,)
        Logistic Z-score per participant WITHOUT intercept term.
    intercept_nominal : float
        The calibrated intercept for the active endpoint.
    delta : float
        Half-width of the intercept uncertainty band (default 0.5).

    Returns
    -------
    dict with keys:
        'p_collapse_mean_lo'  : float  -- mean collapse probability (%), lower bound
        'p_collapse_mean_hi'  : float  -- mean collapse probability (%), upper bound
        'pct_hoog_risico_lo'  : float  -- % individuals p_collapse > 50%, lower bound
        'pct_hoog_risico_hi'  : float  -- % individuals p_collapse > 50%, upper bound
    """
    results = {}
    for suffix, intercept in (('lo', intercept_nominal - delta),
                               ('hi', intercept_nominal + delta)):
        z       = intercept + z_zonder_intercept
        p_ind   = 1.0 / (1.0 + np.exp(-z))
        results[f'p_collapse_mean_{suffix}']  = float(np.nanmean(p_ind)) * 100
        results[f'pct_hoog_risico_{suffix}']  = float(np.nanmean(p_ind > 0.50)) * 100
    return results


def _pct_sensitivity_from_flag(flag_array: np.ndarray,
                                z_zonder_intercept: np.ndarray,
                                intercept_nominal: float,
                                delta: float = 0.5) -> tuple:
    """
    Estimates the calibration sensitivity band for a binary population
    fraction (e.g. pct_decompensatie, pct_roberts_kritiek) that depends
    indirectly on the same Z-logistic threshold.

    Because pct_decompensatie and pct_roberts_kritiek are derived from
    physiological thresholds (CO_reserve, AUC_klinisch) that are NOT
    directly controlled by intercept_kal, the intercept sensitivity does
    NOT apply to them.  Their dominant uncertainty is instead the ±0.3°C
    JOS-3 T_rect systematic offset.

    For those metrics we use a different approach: we shift the effective
    T_rect threshold by ±0.3°C and recompute the fraction.  However,
    this requires the raw arrays, which are not all available here.

    As a pragmatic approximation this function propagates the RELATIVE
    sensitivity of the logistic collapse probability to bound the
    physiological fraction:

        relative_sensitivity = (p_hi - p_lo) / p_nominal

    and applies it symmetrically to the flag-based fraction:

        lo = flag_fraction * (1 - relative_sensitivity / 2)
        hi = flag_fraction * (1 + relative_sensitivity / 2)
        (both clipped to [0, 100])

    This is explicitly an approximation.  The band is labelled '~' in
    output to signal model-based rather than statistical uncertainty.

    Parameters
    ----------
    flag_array : np.ndarray, shape (n_sim,)  -- boolean or 0/1
    z_zonder_intercept : np.ndarray, shape (n_sim,)
    intercept_nominal : float
    delta : float

    Returns
    -------
    (lo, hi) : tuple of float  -- lower and upper band values in %
    """
    flag_pct = float(np.mean(flag_array)) * 100
    if flag_pct == 0.0:
        return 0.0, 0.0

    # Relative sensitivity via logistic band
    z_nom = intercept_nominal + z_zonder_intercept
    p_nom = float(np.nanmean(1.0 / (1.0 + np.exp(-z_nom)))) * 100
    if p_nom == 0.0:
        return flag_pct, flag_pct

    s_lo = _intercept_sensitivity(z_zonder_intercept, intercept_nominal, delta)
    s_hi = s_lo  # same call returns both
    rel  = abs(s_lo['p_collapse_mean_hi'] - s_lo['p_collapse_mean_lo']) / (2.0 * p_nom)

    lo = max(0.0,   flag_pct * (1.0 - rel))
    hi = min(100.0, flag_pct * (1.0 + rel))
    return lo, hi


def _format_pct_with_band(value: float, lo: float, hi: float,
                           decimals: int = 1) -> str:
    """
    Formats a percentage with its calibration uncertainty band.

    Output: "X.X% (~Y.Y–Z.Z%)"

    The tilde signals that the band reflects model calibration uncertainty
    (intercept ± 0.5 log-odds), NOT a classical 95% confidence interval.
    For GHOR reporting this distinction is important: the band should be
    read as "plausible range given calibration source uncertainty".

    Parameters
    ----------
    value    : point estimate (%)
    lo, hi   : lower and upper band (%)
    decimals : decimal places (default 1)

    Returns
    -------
    str, e.g. "7.8% (~5.2–11.4%)"
    """
    fmt = f".{decimals}f"
    return f"{value:{fmt}}% (~{lo:{fmt}}–{hi:{fmt}}%)"


def get_age_group(age):
    """Categorises age into a predefined group string."""
    if 18 <= age <= 29:  return "18-29"
    if 30 <= age <= 39:  return "30-39"
    if 40 <= age <= 49:  return "40-49"
    if 50 <= age <= 59:  return "50-59"
    return "60-65"


def run_monte_carlo_adult(interp_data, lat, lon, met_value, clo_value,
                           n_simulations, age_configuration,
                           training_factor, acclimatization_factor,
                           use_parallel=True,
                           base_population=None,
                           random_seed=None):
    """
    Manages the entire adult Monte Carlo simulation process.

    [rev10] New parameter: base_population (optional)
    --------------------------------------------------
    If base_population is provided (list of AdultParticipantProfile from
    generate_base_population()), it is used directly without re-sampling.
    This enables reproducible scenario comparisons: the SAME population
    experiences different meteorological conditions or MET values.

    If base_population is None, a fresh population is generated with
    generate_base_population() using the given random_seed.

    The age_configuration parameter is IGNORED when base_population is
    provided (age is already embedded in each profile from rev10 onward).
    A deprecation note is printed if age_configuration != 'standard' and
    a base_population is supplied.

    Parameters
    ----------
    interp_data : list of dicts
    lat, lon : float
    met_value : float
        Base MET of the activity (used for liveability check and reference).
    clo_value : float
    n_simulations : int
        Target number of participants. Ignored if base_population is provided.
    age_configuration : str or tuple
        "standard" or (min_age, max_age). Ignored if base_population provided.
    training_factor : float
    acclimatization_factor : float
    use_parallel : bool
    base_population : list of AdultParticipantProfile or None [rev10]
    random_seed : int or None [rev10]

    Returns
    -------
    tuple : (all_results, stats, results_df)
    """
    # ------------------------------------------------------------------
    # [rev10] Population source selection
    # ------------------------------------------------------------------
    if base_population is not None:
        # Use the provided fixed population
        simulations_params_base = base_population
        if age_configuration != "standard":
            print(f"{Fore.YELLOW}[rev10] age_configuration is ignored when "
                  f"base_population is provided.{Style.RESET_ALL}")
        print(f"{Fore.CYAN}[rev10] Using provided fixed population "
              f"({len(simulations_params_base)} profiles).{Style.RESET_ALL}")
    else:
        # Generate a fresh population (backward-compatible path)
        simulations_params_base = generate_base_population(
            n_simulations      = n_simulations,
            training_factor    = training_factor,
            acclimatization_factor = acclimatization_factor,
            random_seed        = random_seed,
        )

    n_simulations = len(simulations_params_base)
    if n_simulations == 0:
        print(f"{Fore.RED}Failed to generate any valid participant profiles. Aborting.{Style.RESET_ALL}")
        return None, None, None

    # ------------------------------------------------------------------
    # [rev10] Note on age_configuration when NOT using base_population
    # ------------------------------------------------------------------
    # In rev10, age is assigned during generate_base_population(), so the
    # age_configuration option below is a no-op for the fixed-population path.
    # For the fresh-population path, age is already embedded in each profile.
    # The block below is retained for backward compatibility only and has no
    # effect on the profile.age field (which was set during generation).

    participant_demographics = []
    worker_args = []

    for params in simulations_params_base:
        # Age is already embedded in the profile from generate_base_population().
        # The age_configuration override from rev09 is removed here.
        # [rev10] If using a fresh population with age_configuration != 'standard',
        # a warning is shown but the profile age is NOT overridden.
        gender = params.gender
        worker_args.append((interp_data, lat, lon, met_value, clo_value, params,
                            training_factor, acclimatization_factor))
        participant_demographics.append({
            'age':       params.age,
            'age_group': get_age_group(params.age),
            'gender':    gender,
        })

    # Parallel execution
    if use_parallel:
        with Pool(multiprocessing.cpu_count()) as pool:
            all_results = list(tqdm(
                pool.imap(worker_monte_carlo_adult, worker_args),
                total=len(worker_args),
                desc="Monte Carlo Adult Sims", unit="sim",
            ))
    else:
        all_results = [worker_monte_carlo_adult(args)
                       for args in tqdm(worker_args, desc="Monte Carlo Adult Sims")]

    # Aggregate results
    t_rect_sims    = np.array([[r["t_rect"]       for r in res] for res in all_results])
    water_sims     = np.array([[r["water"]         for r in res] for res in all_results])
    rpe_sims       = np.array([[r["rpe_total"]     for r in res] for res in all_results])
    stopped_sims   = np.array([[r["stopped"]       for r in res] for res in all_results])
    ehbo_sims      = np.array([[r["ehbo_visit"]    for r in res] for res in all_results])
    unlive_sims    = np.array([[r["is_unliveable"] for r in res] for res in all_results])
    met_sims       = np.array([[r.get("current_met", float('nan')) for r in res] for res in all_results])
    auc_per_sim    = np.array([res[-1].get("auc_thermisch", 0.0) for res in all_results])
    # [rev14-B] AUC_klinisch (Roberts threshold 40.5 deg C)
    auc_kl_per_sim = np.array([res[-1].get("auc_klinisch", 0.0) for res in all_results])
    # [rev14-C] Post-finish outcomes
    ehs_pf_per_sim = np.array([res[-1].get("ehs_postfinish", False) for res in all_results])
    t_rect_pf_sim  = np.array([res[-1].get("t_rect_piek_postfinish", res[-1]["t_rect"])
                                for res in all_results])
    auc_kl_tot_sim = np.array([res[-1].get("auc_klinisch_totaal", auc_kl_per_sim[i])
                                for i, res in enumerate(all_results)])
    # [rev14-A] K_p population distribution
    kp_per_sim     = np.array([res[-1].get("kp_ind", K_P_PACING) for res in all_results])

    # [rev16] Vulnerable upper-tail cohort for operational plotting.
    # This is NOT a confidence interval. It is a fixed subgroup: participants
    # whose personal peak T_rect over the full event lies in the hottest 2.5%
    # of the simulated population. Keeping the cohort fixed makes the plot
    # interpretable as "the most vulnerable runners" rather than the highest
    # value at each separate time point.
    vulnerable_tail_pct_target = 2.5
    peak_t_rect_per_sim = np.nanmax(t_rect_sims, axis=1)
    vulnerable_tail_cutoff = float(np.nanpercentile(
        peak_t_rect_per_sim, 100.0 - vulnerable_tail_pct_target
    ))
    vulnerable_tail_mask = peak_t_rect_per_sim >= vulnerable_tail_cutoff
    if not np.any(vulnerable_tail_mask):
        vulnerable_tail_mask[int(np.nanargmax(peak_t_rect_per_sim))] = True
    vulnerable_tail_n = int(np.sum(vulnerable_tail_mask))
    vulnerable_tail_pct_actual = float(vulnerable_tail_n / n_simulations * 100)
    vulnerable_tail_t_rect_sims = t_rect_sims[vulnerable_tail_mask, :]
    vulnerable_tail_profiles = [
        p for p, selected in zip(simulations_params_base, vulnerable_tail_mask)
        if selected
    ]

    stats = {
        'mean_t_rect':        np.mean(t_rect_sims, axis=0),
        'lower_t_rect':       np.percentile(t_rect_sims,  5, axis=0),
        'upper_t_rect':       np.percentile(t_rect_sims, 95, axis=0),
        't_rect_p975':        np.percentile(t_rect_sims, 97.5, axis=0),
        't_rect_p99':         np.percentile(t_rect_sims, 99, axis=0),
        'vulnerable_tail_pct_target': vulnerable_tail_pct_target,
        'vulnerable_tail_pct_actual': vulnerable_tail_pct_actual,
        'vulnerable_tail_n':          vulnerable_tail_n,
        'vulnerable_tail_peak_cutoff': vulnerable_tail_cutoff,
        'vulnerable_tail_mean_t_rect': np.nanmean(vulnerable_tail_t_rect_sims, axis=0),
        'vulnerable_tail_p50_t_rect':  np.nanpercentile(vulnerable_tail_t_rect_sims, 50, axis=0),
        'vulnerable_tail_p95_t_rect':  np.nanpercentile(vulnerable_tail_t_rect_sims, 95, axis=0),
        'vulnerable_tail_age_mean':    float(np.mean([p.age for p in vulnerable_tail_profiles])),
        'vulnerable_tail_vo2max_mean': float(np.mean([p.vo2max for p in vulnerable_tail_profiles])),
        'vulnerable_tail_kp_mean':     float(np.mean([
            getattr(p, 'kp_pacing', K_P_PACING) for p in vulnerable_tail_profiles
        ])),
        'vulnerable_tail_nsaid_pct':   float(np.mean([
            getattr(p, 'nsaid_gebruik', False) for p in vulnerable_tail_profiles
        ]) * 100),
        'mean_water':         np.mean(water_sims, axis=0),
        'lower_water':        np.percentile(water_sims,  5, axis=0),
        'upper_water':        np.percentile(water_sims, 95, axis=0),
        'mean_rpe':           np.mean(rpe_sims, axis=0),
        'lower_rpe':          np.percentile(rpe_sims,  5, axis=0),
        'upper_rpe':          np.percentile(rpe_sims, 95, axis=0),
        'percent_stopped':    np.mean(stopped_sims, axis=0) * 100,
        'percent_ehbo':       np.mean(ehbo_sims, axis=0) * 100,
        'percent_unliveable': np.mean(unlive_sims, axis=0) * 100,
        'temp_threshold_stats': _calculate_threshold_stats(t_rect_sims, n_simulations),
        'mean_current_met':   np.nanmean(met_sims, axis=0),
        'lower_current_met':  np.nanpercentile(met_sims,  5, axis=0),
        'upper_current_met':  np.nanpercentile(met_sims, 95, axis=0),
        'auc_thermisch_p50':  float(np.percentile(auc_per_sim, 50)),
        'auc_thermisch_p95':  float(np.percentile(auc_per_sim, 95)),
        'auc_thermisch_mean': float(np.mean(auc_per_sim)),
        'pct_auc_above_5':    float(np.mean(auc_per_sim > 5.0) * 100),
        # [rev14-A] K_p population statistics
        'kp_mean':            float(np.mean(kp_per_sim)),
        'kp_p05':             float(np.percentile(kp_per_sim,  5)),
        'pct_kp_onder_002':   float(np.mean(kp_per_sim < 0.02) * 100),  # overgemotiveerd
        'pct_nsaid':          float(np.mean([getattr(p, 'nsaid_gebruik', False)
                               for p in simulations_params_base]) * 100),
        # [rev14-B] AUC_klinisch (Roberts 2007)
        'auc_klinisch_p50':   float(np.percentile(auc_kl_per_sim, 50)),
        'auc_klinisch_p95':   float(np.percentile(auc_kl_per_sim, 95)),
        'auc_klinisch_totaal_p95': float(np.percentile(auc_kl_tot_sim, 95)),
        'pct_roberts_kritiek': float(np.mean(auc_kl_tot_sim > AUC_ROBERTS_GRENS) * 100),
        # [rev14-C] Post-finish
        'pct_ehs_postfinish': float(np.mean(ehs_pf_per_sim) * 100),
        't_rect_pf_p95':      float(np.percentile(t_rect_pf_sim, 95)),
    }

    # CVR population statistics
    if CVR_BESCHIKBAAR:
        hr_sims  = np.array([[r.get('hr_geschat',   float('nan')) for r in res] for res in all_results])
        cvs_sims = np.array([[r.get('cvs_index',    float('nan')) for r in res] for res in all_results])
        res_sims = np.array([[r.get('co_reserve',   float('nan')) for r in res] for res in all_results])
        dec_sims = np.array([[r.get('decompensatie', False)       for r in res] for res in all_results])

        hr_piek  = np.nanmax(hr_sims,  axis=1)
        cvs_piek = np.nanmax(cvs_sims, axis=1)
        res_min  = np.nanmin(res_sims, axis=1)
        dec_ooit = np.any(dec_sims, axis=1)

        dehy_sims = np.array([
            [r.get('water', 0) / 1000 / max(1, simulations_params_base[i].weight) * 100
             for r in res]
            for i, res in enumerate(all_results)
        ])
        dehy_eind = dehy_sims[:, -1]
        vulnerable_tail_res_min = res_min[vulnerable_tail_mask]
        vulnerable_tail_dehy_eind = dehy_eind[vulnerable_tail_mask]
        hr_maxen  = [208 - 0.7 * p.age for p in simulations_params_base]

        # Collapse risk (two-phase logistic)
        # [rev12] intercept_kal and P_OBS now come from COLLAPSE_ENDPOINTS table.
        # The active endpoint is set by ACTIVE_ENDPOINT (module-level constant)
        # or overridden by the endpoint_key argument passed to this function.
        W_T1 = 1.0
        W_T2 = 4.0
        W_C  = 0.8
        W_D  = 0.5
        P_OBS = P_OBS_COLLAPSE

        t_rect_max_per_sim = np.nanmax(t_rect_sims, axis=1)
        z_zonder_intercept = (
              W_T1 * np.clip(t_rect_max_per_sim - 39.5, 0.0, None)
            + W_T2 * np.clip(t_rect_max_per_sim - 40.5, 0.0, None)
            + W_C  * np.clip(2.0 - res_min,            0.0, None)
            + W_D  * np.clip(dehy_eind - 3.0,          0.0, None)
        )

        # [rev12] Read intercept from active endpoint in COLLAPSE_ENDPOINTS table
        active_ep         = COLLAPSE_ENDPOINTS[ACTIVE_ENDPOINT]
        intercept_kal     = active_ep['intercept_kal']  # [rev12: was hardcoded -10.071]

        z_collapse          = intercept_kal + z_zonder_intercept
        p_collapse_per_sim  = 1.0 / (1.0 + np.exp(-z_collapse))

        p_collapse_mediaan  = float(np.nanpercentile(p_collapse_per_sim, 50)) * 100
        p_collapse_p95      = float(np.nanpercentile(p_collapse_per_sim, 95)) * 100
        p_collapse_gemiddeld= float(np.nanmean(p_collapse_per_sim)) * 100
        pct_hoog_risico     = float(np.nanmean(p_collapse_per_sim > 0.50)) * 100
        verwacht_per_1000   = p_collapse_gemiddeld * 10

        stats.update({
            'hr_piek_p50':        float(np.nanpercentile(hr_piek,  50)),
            'hr_piek_p95':        float(np.nanpercentile(hr_piek,  95)),
            'cvs_piek_p50':       float(np.nanpercentile(cvs_piek, 50)),
            'cvs_piek_p95':       float(np.nanpercentile(cvs_piek, 95)),
            'pct_cvs_boven_90':   float(np.nanmean(cvs_piek > 0.90) * 100),
            'pct_decompensatie':  float(np.mean(dec_ooit) * 100),
            'co_reserve_min_p50': float(np.nanpercentile(res_min,   50)),
            'co_reserve_min_p05': float(np.nanpercentile(res_min,    5)),
            'dehy_pct_p50':       float(np.nanpercentile(dehy_eind, 50)),
            'dehy_pct_p95':       float(np.nanpercentile(dehy_eind, 95)),
            'vulnerable_tail_co_reserve_min_p50': float(np.nanpercentile(vulnerable_tail_res_min, 50)),
            'vulnerable_tail_co_reserve_min_p05': float(np.nanpercentile(vulnerable_tail_res_min, 5)),
            'vulnerable_tail_dehy_pct_p50':       float(np.nanpercentile(vulnerable_tail_dehy_eind, 50)),
            'vulnerable_tail_dehy_pct_p95':       float(np.nanpercentile(vulnerable_tail_dehy_eind, 95)),
            'hr_max_populatie':   float(np.mean(hr_maxen)),
            'p_collapse_mediaan':           p_collapse_mediaan,
            'p_collapse_p95':               p_collapse_p95,
            'p_collapse_gemiddeld':         p_collapse_gemiddeld,
            'pct_hoog_collapsrisico':       pct_hoog_risico,
            'p_collapse_per_sim':           p_collapse_per_sim,
            'collapse_intercept_kal':       float(intercept_kal),
            'collapse_p_obs_pct':           P_OBS * 100,
            'verwacht_collapsen_per_1000':  verwacht_per_1000,
            # [rev12] Endpoint traceability
            'active_endpoint':              ACTIVE_ENDPOINT,
            'active_endpoint_label':        active_ep['label'],
            'active_endpoint_status':       active_ep['status'],
            'active_endpoint_source':       active_ep['source'],
            'active_endpoint_p_obs':        active_ep['p_obs'],
        })

        # [rev15] Calibration uncertainty bands (intercept ± 0.5 log-odds)
        _sens = _intercept_sensitivity(z_zonder_intercept, intercept_kal, delta=0.5)
        dec_lo, dec_hi = _pct_sensitivity_from_flag(
            dec_ooit, z_zonder_intercept, intercept_kal, delta=0.5)
        rob_lo, rob_hi = _pct_sensitivity_from_flag(
            (auc_kl_tot_sim > AUC_ROBERTS_GRENS), z_zonder_intercept, intercept_kal, delta=0.5)
        pf_lo, pf_hi   = _pct_sensitivity_from_flag(
            ehs_pf_per_sim, z_zonder_intercept, intercept_kal, delta=0.5)
        stats.update({
            # Collapse probability band
            'p_collapse_gemiddeld_lo':   _sens['p_collapse_mean_lo'],
            'p_collapse_gemiddeld_hi':   _sens['p_collapse_mean_hi'],
            'pct_hoog_collapsrisico_lo': _sens['pct_hoog_risico_lo'],
            'pct_hoog_collapsrisico_hi': _sens['pct_hoog_risico_hi'],
            # Physiological fraction bands (approximate; see docstring)
            'pct_decompensatie_lo':      dec_lo,
            'pct_decompensatie_hi':      dec_hi,
            'pct_roberts_kritiek_lo':    rob_lo,
            'pct_roberts_kritiek_hi':    rob_hi,
            'pct_ehs_postfinish_lo':     pf_lo,
            'pct_ehs_postfinish_hi':     pf_hi,
        })

    # Build Excel export DataFrame
    detailed_export_data = []
    for i, sim_result in enumerate(all_results):
        p           = simulations_params_base[i]
        demographics= participant_demographics[i]
        max_t_rect  = max(r["t_rect"] for r in sim_result)
        max_rpe     = max(r["rpe_total"] for r in sim_result)
        max_dehy    = max(r["water"] / 1000 / p.weight * 100 for r in sim_result) if p.weight > 0 else 0

        # [rev10] Include pct_vo2max and implied MET in export
        implied_met = p.vo2max * p.pct_vo2max / VO2MAX_TO_MET_FACTOR
        final_auc   = sim_result[-1].get('auc_thermisch', 0.0)
        final_auc_kl= sim_result[-1].get('auc_klinisch', 0.0)
        auc_kl_tot  = sim_result[-1].get('auc_klinisch_totaal', final_auc_kl)
        min_met     = min((r.get('current_met', implied_met) for r in sim_result), default=implied_met)
        ehs_pf      = sim_result[-1].get('ehs_postfinish', False)
        t_rect_pf   = sim_result[-1].get('t_rect_piek_postfinish', sim_result[-1]['t_rect'])

        row = {
            'participant_id':        i + 1,
            'age':                   p.age,
            'gender':                p.gender,
            'age_group':             demographics['age_group'],
            'height_m':              round(p.height, 2),
            'weight_kg':             round(p.weight, 1),
            'vo2max_ml_kg_min':      round(p.vo2max, 1),
            'pct_vo2max':            round(p.pct_vo2max, 3),
            'implied_met_initial':   round(implied_met, 2),
            'met_min_dynamic':       round(min_met, 2),
            'kp_pacing':             round(getattr(p, 'kp_pacing', K_P_PACING), 4),  # [rev14-A]
            'nsaid_gebruik':         int(getattr(p, 'nsaid_gebruik', False)),          # [rev14-A]
            'auc_thermisch_deg_min': round(final_auc, 2),
            'auc_klinisch_deg_min':  round(final_auc_kl, 2),                          # [rev14-B]
            'auc_klinisch_totaal':   round(auc_kl_tot, 2),                            # [rev14-C]
            'roberts_kritiek':       int(auc_kl_tot > AUC_ROBERTS_GRENS),             # [rev14-B]
            'ehs_postfinish':        int(ehs_pf),                                     # [rev14-C]
            't_rect_piek_postfinish':round(t_rect_pf, 3),                            # [rev14-C]
            'vulnerable_tail_top_2_5pct': int(bool(vulnerable_tail_mask[i])),          # [rev16]
            'vulnerable_tail_peak_cutoff': round(vulnerable_tail_cutoff, 3),           # [rev16]
            'temp_variation_C':      round(p.temp_variation, 2),
            'rh_variation_perc':     round(p.rh_variation, 1),
            'mf_score':              round(p.mf_score, 2),
            'sweat_factor':          round(p.sweat_factor, 2),
            'thirst_threshold_perc': round(p.thirst_threshold, 2),
            'max_t_rect':            round(max_t_rect, 4),
            'max_rpe':               round(max_rpe, 2),
            'max_water_loss_perc':   round(max_dehy, 2),
        }

        if CVR_BESCHIKBAAR:
            hr_vals  = [r.get('hr_geschat',  float('nan')) for r in sim_result]
            cvs_vals = [r.get('cvs_index',   float('nan')) for r in sim_result]
            res_vals = [r.get('co_reserve',  float('nan')) for r in sim_result]
            row.update({
                'cvr_hr_piek':        round(float(np.nanmax(hr_vals)),  1),
                'cvr_cvs_index_max':  round(float(np.nanmax(cvs_vals)), 3),
                'cvr_co_reserve_min': round(float(np.nanmin(res_vals)), 2),
                'cvr_decompensatie':  any(r.get('decompensatie', False) for r in sim_result),
                'cvr_p_collapse_pct': round(
                    float(stats['p_collapse_per_sim'][i]) * 100, 4
                ) if 'p_collapse_per_sim' in stats else float('nan'),
                # [rev12] Endpoint traceability columns
                'collapse_endpoint':        stats.get('active_endpoint', 'ehs'),
                'collapse_endpoint_label':  stats.get('active_endpoint_label', ''),
                'collapse_endpoint_status': stats.get('active_endpoint_status', ''),
                'collapse_p_obs_per_10k':   round(
                    stats.get('active_endpoint_p_obs', float('nan')) * 10_000, 2),
                'collapse_intercept_kal':   stats.get('collapse_intercept_kal', float('nan')),
            })
        detailed_export_data.append(row)

    results_df = pd.DataFrame(detailed_export_data)
    return all_results, stats, results_df


def run_monte_carlo_pediatric(interp_data, lat, lon, met_value, clo_value,
                               n_simulations=200, age_group="11-20",
                               fitness_factor=0.5, use_parallel=True):
    """Manages the pediatric Monte Carlo simulation process. Unchanged from rev09."""
    participant_params_list = simulate_child_participant(
        n_simulations, met_value, age_group, fitness_factor)

    n_simulations = len(participant_params_list)
    if n_simulations == 0:
        print(f"{Fore.RED}Failed to generate any valid profiles. Aborting.{Style.RESET_ALL}")
        return None, None

    worker_args = [(interp_data, lat, lon, met_value, clo_value, params, age_group, fitness_factor)
                   for params in participant_params_list]

    if use_parallel:
        with Pool(multiprocessing.cpu_count()) as pool:
            all_results = list(tqdm(
                pool.imap(worker_monte_carlo_pediatric, worker_args),
                total=len(worker_args), desc="Monte Carlo Pediatric Sims"))
    else:
        all_results = [worker_monte_carlo_pediatric(args)
                       for args in tqdm(worker_args, desc="Monte Carlo Pediatric Sims")]

    t_core_sims   = np.array([[r["t_core"]        for r in res] for res in all_results])
    water_sims    = np.array([[r["water_loss_ml"]  for r in res] for res in all_results])
    rpe_sims      = np.array([[r["rpe_total"]      for r in res] for res in all_results])
    stopped_sims  = np.array([[r["stopped"]        for r in res] for res in all_results])
    first_aid_sims= np.array([[r["first_aid_visit"]for r in res] for res in all_results])

    stats = {
        'mean_t_core':      np.mean(t_core_sims, axis=0),
        'lower_t_core':     np.percentile(t_core_sims,  5, axis=0),
        'upper_t_core':     np.percentile(t_core_sims, 95, axis=0),
        'mean_water':       np.mean(water_sims, axis=0),
        'lower_water':      np.percentile(water_sims,  5, axis=0),
        'upper_water':      np.percentile(water_sims, 95, axis=0),
        'mean_rpe':         np.mean(rpe_sims, axis=0),
        'lower_rpe':        np.percentile(rpe_sims,  5, axis=0),
        'upper_rpe':        np.percentile(rpe_sims, 95, axis=0),
        'percent_stopped':  np.mean(stopped_sims, axis=0) * 100,
        'percent_first_aid':np.mean(first_aid_sims, axis=0) * 100,
        'temp_threshold_stats': _calculate_threshold_stats(t_core_sims, n_simulations),
    }
    return all_results, stats


# =============================================================================
#  RISK CLASSIFICATION (unchanged from rev09)
# =============================================================================

def calculate_adult_risk_classification(utci_val, wbgt_val, t_rect, rpe_total):
    scores = {
        'utci':  0 if utci_val <= 26 else 1 if utci_val <= 32 else 2 if utci_val <= 38 else 3 if utci_val <= 46 else 4,
        'wbgt':  0 if wbgt_val <  25 else 1 if wbgt_val <  28 else 2 if wbgt_val <  31 else 3 if wbgt_val <  33 else 4,
        't_rect':0 if t_rect   <  38 else 1 if t_rect   <  39 else 2 if t_rect   < 39.5 else 3 if t_rect < 40 else 4,
        'rpe':   0 if rpe_total< 12  else 1 if rpe_total< 15  else 2 if rpe_total< 17   else 3 if rpe_total < 19 else 4,
    }
    s = max(scores.values())
    lvl = {0: ("No risk", Fore.WHITE), 1: ("Low risk", Fore.GREEN),
           2: ("Moderate risk", Fore.YELLOW), 3: ("High risk", Fore.MAGENTA),
           4: ("Extreme risk", Fore.RED)}
    text, color = lvl[s]
    return {"risk_color": color,
            "detailed_risk": f"{text} (U:{scores['utci']}/W:{scores['wbgt']}/T:{scores['t_rect']}/RPE:{scores['rpe']})"}


def calculate_pediatric_risk_classification(utci_val, wbgt_val, t_core, rpe_total):
    scores = {
        'utci':  0 if utci_val <= 26 else 1 if utci_val <= 32 else 2 if utci_val <= 38 else 3 if utci_val <= 46 else 4,
        'wbgt':  0 if wbgt_val <  23 else 1 if wbgt_val <  26 else 2 if wbgt_val <  29 else 3 if wbgt_val <  31 else 4,
        't_core':0 if t_core   < 38.2 else 1 if t_core  < 38.8 else 2 if t_core  < 39.2 else 3 if t_core < 39.8 else 4,
        'rpe':   0 if rpe_total<  4   else 1 if rpe_total< 6    else 2 if rpe_total< 8    else 3 if rpe_total < 9  else 4,
    }
    s = max(scores.values())
    lvl = {0: ("No risk", Fore.WHITE), 1: ("Low risk", Fore.GREEN),
           2: ("Moderate risk", Fore.YELLOW), 3: ("High risk", Fore.MAGENTA),
           4: ("Extreme risk", Fore.RED)}
    text, color = lvl[s]
    return {"risk_color": color,
            "detailed_risk": f"{text} (U:{scores['utci']}/W:{scores['wbgt']}/T:{scores['t_core']}/RPE:{scores['rpe']})"}


# =============================================================================
#  PLOTTING
#  [rev16] Adult T_rect panel separates distribution percentiles from a fixed
#  vulnerable tail cohort. The cohort is defined once by each runner's peak
#  T_rect over the full event, so the plotted tail line follows the same
#  high-risk runners over time instead of changing membership at every step.
# =============================================================================

def plot_adult_results(results, stats, city_name, start_time, duration_hours):
    """
    4-panel plot for adult simulation results.

    The T_rect panel includes P97.5/P99 reference curves and, when available,
    the mean trajectory of the vulnerable tail cohort: the runners whose own
    peak T_rect falls in the highest 2.5% of the Monte Carlo population.
    """
    times_str = [r["time"] for r in results[0]]
    times_dt  = [pd.to_datetime(t) for t in times_str]

    fig, axs = plt.subplots(2, 2, figsize=(20, 14), gridspec_kw={'hspace': 0.4, 'wspace': 0.3})
    fig.suptitle(f"Adult Thermal Simulation for {city_name} on {pd.to_datetime(start_time).date()}",
                 fontsize=18, weight='bold')

    ax1 = axs[0, 0]
    ax1.plot(times_dt, stats['mean_t_rect'], 'b-', label='Mean T_rect')
    ax1.fill_between(times_dt, stats['lower_t_rect'], stats['upper_t_rect'],
                     color='blue', alpha=0.2, label='5th-95th Percentile')
    if 't_rect_p975' in stats:
        ax1.plot(times_dt, stats['t_rect_p975'], color='darkorange',
                 linestyle='-.', linewidth=2.0, label='P97.5 T_rect')
    if 't_rect_p99' in stats:
        ax1.plot(times_dt, stats['t_rect_p99'], color='red',
                 linestyle=':', linewidth=2.4, label='P99 T_rect')
    if 'vulnerable_tail_mean_t_rect' in stats:
        tail_label = (
            f"Vulnerable tail mean "
            f"(top {stats.get('vulnerable_tail_pct_actual', 2.5):.1f}% by peak T_rect)"
        )
        ax1.plot(times_dt, stats['vulnerable_tail_mean_t_rect'], color='black',
                 linestyle='-', linewidth=2.6, label=tail_label)
    ax1.axhline(y=39.0, color='gold',       linestyle='--', label='Heat Stress (39°C)')
    ax1.axhline(y=40.0, color='darkorange', linestyle='--', label='Serious Risk (40°C)')
    ax1.axhline(y=42.0, color='red',        linestyle='--', label='Life-threatening (42°C)')
    ax1.axhline(y=43.0, color='black',      linestyle='-',  linewidth=2,
                label='Lethal Limit (43°C, Vanos 2023)')
    ax1.set_title("Core Body Temperature (T_rect)", fontsize=14)
    ax1.set_ylabel("T_rect (°C)", fontsize=12)
    ax1.legend()

    ax2 = axs[0, 1]
    ax2.plot(times_dt, stats['mean_rpe'], 'r-', label='Mean RPE')
    ax2.fill_between(times_dt, stats['lower_rpe'], stats['upper_rpe'],
                     color='red', alpha=0.2, label='5th-95th Percentile')
    ax2.axhline(y=17,   color='purple', linestyle=':',  label='Exhaustion (RPE 17)')
    ax2.axhline(y=19.5, color='black',  linestyle='--', label='Stop Threshold (RPE 19.5)')
    ax2.set_title("Rate of Perceived Exertion (Borg 6-20)", fontsize=14)
    ax2.set_ylabel("RPE (6-20)", fontsize=12)
    ax2.set_ylim(6, 20.5)
    ax2.legend()

    ax3  = axs[1, 0]
    utci_vals = [r["utci"] for r in results[0]]
    wbgt_vals = [r["wbgt"] for r in results[0]]
    ax3.plot(times_dt, utci_vals, color='purple', label='UTCI (°C)')
    ax3.set_ylabel("UTCI (°C)", color='purple', fontsize=12)
    ax3.legend(loc='upper left')
    ax4 = ax3.twinx()
    ax4.plot(times_dt, wbgt_vals, color='green', label='WBGT (°C)')
    ax4.set_ylabel("WBGT (°C)", color='green', fontsize=12)
    ax4.legend(loc='upper right')
    ax3.set_title("Environmental Thermal Indices", fontsize=14)

    ax5 = axs[1, 1]
    ax5.plot(times_dt, stats['percent_stopped'],    color='black',     linestyle='--', label='% Stopped')
    ax5.plot(times_dt, stats['percent_ehbo'],       color='magenta',   linestyle='--', label='% First Aid')
    ax5.plot(times_dt, stats['percent_unliveable'], color='darkviolet',linewidth=2.5,  label='% Unliveable')
    ax5.fill_between(times_dt, 0, stats['percent_unliveable'], color='darkviolet', alpha=0.2)
    ax5.set_title("Percentage of Participants at Risk", fontsize=14)
    ax5.set_ylabel("Participants (%)", fontsize=12)
    ax5.set_ylim(0, 101)
    ax5.legend()

    for ax in axs.flat:
        ax.set_xlabel("Time", fontsize=12)
        ax.grid(True, linestyle='--', alpha=0.6)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        ax.tick_params(axis='x', rotation=45)
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.show(block=False)


def plot_pediatric_results(results, stats, city_name, start_time, duration_hours):
    """4-panel plot for pediatric simulation results. Unchanged."""
    times_str = [r["time"] for r in results[0]]
    times_dt  = [pd.to_datetime(t) for t in times_str]
    fig, axs  = plt.subplots(2, 2, figsize=(20, 14), gridspec_kw={'hspace': 0.4, 'wspace': 0.3})
    fig.suptitle(f"Pediatric Thermal Simulation for {city_name} on {pd.to_datetime(start_time).date()}",
                 fontsize=18, weight='bold')

    ax1 = axs[0, 0]
    ax1.plot(times_dt, stats['mean_t_core'], 'b-', label='Mean T_core')
    ax1.fill_between(times_dt, stats['lower_t_core'], stats['upper_t_core'],
                     color='blue', alpha=0.2, label='5th-95th Percentile')
    ax1.axhline(y=38.5, color='gold',       linestyle='--', label='Moderate Stress (38.5°C)')
    ax1.axhline(y=39.2, color='darkorange', linestyle='--', label='High Risk (39.2°C)')
    ax1.set_title("Core Body Temperature (T_core)", fontsize=14)
    ax1.set_ylabel("T_core (°C)", fontsize=12)
    ax1.legend()

    ax2 = axs[0, 1]
    ax2.plot(times_dt, stats['mean_rpe'], 'r-', label='Mean RPE (OMNI 0-10)')
    ax2.fill_between(times_dt, stats['lower_rpe'], stats['upper_rpe'],
                     color='red', alpha=0.2, label='5th-95th Percentile')
    ax2.axhline(y=8.5, color='purple', linestyle=':', label='First Aid (RPE 8.5)')
    ax2.axhline(y=9.5, color='black',  linestyle='--', label='Stop Threshold (RPE 9.5)')
    ax2.set_title("Rate of Perceived Exertion (OMNI 0-10)", fontsize=14)
    ax2.set_ylabel("RPE (0-10)", fontsize=12)
    ax2.set_ylim(0, 10.5)
    ax2.legend()

    ax3 = axs[1, 0]
    utci_vals = [r["utci"] for r in results[0]]
    wbgt_vals = [r["wbgt"] for r in results[0]]
    ax3.plot(times_dt, utci_vals, color='purple', label='UTCI (°C)')
    ax3.set_ylabel("UTCI (°C)", color='purple', fontsize=12)
    ax3.legend(loc='upper left')
    ax4 = ax3.twinx()
    ax4.plot(times_dt, wbgt_vals, color='green', label='WBGT (°C)')
    ax4.set_ylabel("WBGT (°C)", color='green', fontsize=12)
    ax4.legend(loc='upper right')
    ax3.set_title("Environmental Thermal Indices", fontsize=14)

    ax5 = axs[1, 1]
    ax5.plot(times_dt, stats['percent_stopped'],   color='black',   label='% Stopped')
    ax5.plot(times_dt, stats['percent_first_aid'], color='magenta', linestyle='--', label='% First Aid')
    ax5.set_title("Percentage of Participants at Risk", fontsize=14)
    ax5.set_ylabel("Participants (%)", fontsize=12)
    ax5.set_ylim(0, 101)
    ax5.legend()

    for ax in axs.flat:
        ax.set_xlabel("Time", fontsize=12)
        ax.grid(True, linestyle='--', alpha=0.6)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        ax.tick_params(axis='x', rotation=45)
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.show(block=False)


def plot_adult_boxplots(results_df):
    """Boxplots by age group and gender. Unchanged."""
    if results_df is None or results_df.empty:
        return
    sns.set(style="whitegrid")
    variables = [
        ('max_t_rect',         'Maximum Core Temperature (°C)', [37, 42]),
        ('max_rpe',            'Maximum Perceived Exertion (RPE)', [6, 20]),
        ('max_water_loss_perc','Maximum Water Loss (% body weight)', [0, 5]),
    ]
    fig, axes = plt.subplots(3, 1, figsize=(14, 18))
    fig.suptitle("Distribution of Max Values by Age Group and Gender", fontsize=16)
    for i, (var, title, ylim) in enumerate(variables):
        ax    = axes[i]
        order = sorted(results_df['age_group'].unique())
        sns.boxplot(data=results_df, x='age_group', y=var, hue='gender', ax=ax,
                    palette={"male": "skyblue", "female": "lightpink"},
                    showmeans=True,
                    meanprops={"marker": "o", "markerfacecolor": "white", "markeredgecolor": "black"},
                    order=order)
        ax.set_title(title, fontsize=14)
        ax.set_xlabel("Age Group", fontsize=12)
        ax.set_ylabel(title.split('(')[0].strip(), fontsize=12)
        ax.set_ylim(ylim)
        ax.legend(title="Gender", loc='upper right')
        ax.grid(True, linestyle='--', alpha=0.6)
        sns.stripplot(data=results_df, x='age_group', y=var, hue='gender', ax=ax,
                      dodge=True, jitter=True, alpha=0.4,
                      palette={"male": "blue", "female": "red"}, legend=False, order=order)
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.show(block=False)


# =============================================================================
#  MET THRESHOLD ANALYSIS (unchanged from rev09)
# =============================================================================

def get_representative_participant(age, gender):
    if gender == 'male':
        return {'height': 1.80, 'weight': 78, 'age': age, 'fat': 15, 'sex': gender}
    return {'height': 1.67, 'weight': 65, 'age': age, 'fat': 25, 'sex': gender}


def run_single_simulation_for_temp(met_value, env_conditions, participant, clo_value, duration_minutes):
    jos3 = JOS3(height=participant['height'], weight=participant['weight'],
                age=participant['age'], fat=participant['fat'], sex=participant['sex'])
    jos3.t_core[:] = 37.0
    jos3.par = met_value
    jos3.clo = clo_value
    jos3.tdb = env_conditions['tdb']
    jos3.tr  = env_conditions['mrt']
    jos3.v   = env_conditions['v']
    jos3.rh  = env_conditions['rh']
    jos3.simulate(times=int(duration_minutes))
    return jos3.t_core.mean()


def find_met_for_target_temp(target_temp, env_conditions, participant, clo_value, duration_minutes):
    met_low, met_high = 1.0, 20.0
    if run_single_simulation_for_temp(met_high, env_conditions, participant, clo_value, duration_minutes) < target_temp:
        return met_high
    if run_single_simulation_for_temp(met_low,  env_conditions, participant, clo_value, duration_minutes) > target_temp:
        return met_low
    for _ in range(5):
        met_guess = (met_low + met_high) / 2
        t_guess   = run_single_simulation_for_temp(met_guess, env_conditions, participant, clo_value, duration_minutes)
        if t_guess < target_temp:
            met_low  = met_guess
        else:
            met_high = met_guess
    return (met_low + met_high) / 2


def analyze_met_thresholds(interp_data, lat, lon, clo_value):
    print("\nStarting analysis of safe MET thresholds...")
    age_groups      = {"18-29": 25, "30-39": 35, "40-49": 45, "50-59": 55, "60-65": 62}
    genders         = ['male', 'female']
    temp_thresholds = [39.0, 39.5, 40.0, 41.0]
    results         = []
    aqi             = get_air_quality(lat, lon)
    detailed_env    = []
    for entry in interp_data:
        rad      = calculate_solar_radiance(lat, lon, str(entry['time']), entry['clouds'], aqi)
        w15      = wind_speed_at_height(entry['wind'], 10, 1.5)
        gt       = calculate_globe_temperature(entry['temp'], rad['ghi'], w15,
                                               rad['solar_elevation'], entry['pressure'],
                                               entry['clouds'], aqi)
        mrt      = calculate_mrt(gt, entry['temp'], w15, rad['ghi'], rad['solar_elevation'])
        detailed_env.append({'time': entry['time'], 'tdb': entry['temp'],
                              'mrt': mrt, 'v': max(0.1, w15), 'rh': entry['rh']})

    total = len(detailed_env) * len(age_groups) * len(genders) * len(temp_thresholds)
    with tqdm(total=total, desc="Calculating MET Thresholds") as pbar:
        for i, env_step in enumerate(detailed_env):
            duration_minutes = max(1, (env_step['time'] - detailed_env[0]['time']).total_seconds() / 60)
            for age_label, age_val in age_groups.items():
                for gender in genders:
                    participant = get_representative_participant(age_val, gender)
                    for target_temp in temp_thresholds:
                        results.append({
                            'time':            env_step['time'],
                            'age_group':       age_label,
                            'gender':          gender,
                            'temp_threshold':  target_temp,
                            'permissible_met': find_met_for_target_temp(
                                target_temp, env_step, participant, clo_value, duration_minutes),
                        })
                        pbar.update(1)
    return pd.DataFrame(results)


def plot_met_thresholds(df_met, user_selected_met):
    """Plot permissible MET values by age group. Unchanged."""
    sns.set(style="whitegrid")
    age_groups   = sorted(df_met['age_group'].unique())
    n_groups     = len(age_groups)
    fig, axes    = plt.subplots(n_groups, 1, figsize=(12, 5 * n_groups), sharex=True, sharey=True)
    if n_groups == 1:
        axes = [axes]
    temp_colors  = {39.0: 'blue', 39.5: 'green', 40.0: 'orange', 41.0: 'red'}
    gender_styles= {'male': '-', 'female': '--'}
    for i, age_group in enumerate(age_groups):
        ax       = axes[i]
        df_group = df_met[df_met['age_group'] == age_group]
        for gender in ['male', 'female']:
            df_g = df_group[df_group['gender'] == gender]
            for temp, color in temp_colors.items():
                df_t = df_g[df_g['temp_threshold'] == temp]
                if not df_t.empty:
                    ax.plot(df_t['time'], df_t['permissible_met'],
                            label=f'{gender.capitalize()} < {temp}°C',
                            color=color, linestyle=gender_styles[gender],
                            marker='o', markersize=3)
        ax.axhline(y=user_selected_met, color='black', linestyle=':', linewidth=2.5,
                   label=f'Chosen Activity (MET: {user_selected_met})')
        ax.set_title(f'Permissible Activity Intensity: {age_group}', fontsize=14)
        ax.set_ylabel('MET Value', fontsize=12)
        ax.legend(title="Legend", bbox_to_anchor=(1.02, 1), loc='upper left')
        ax.grid(True, which='both', linestyle='--', linewidth=0.5)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
    axes[-1].set_xlabel("Time", fontsize=12)
    plt.ylim(bottom=1)
    fig.suptitle('Analysis of Maximum Permissible MET Values', fontsize=18, weight='bold')
    plt.tight_layout(rect=[0, 0, 0.85, 0.96])
    plt.show()


# =============================================================================
#  THERMOPOULOS EXCEL LOADER (unchanged from rev09)
# =============================================================================

def load_weather_from_thermopoulos_excel(filepath, sheet_name):
    """Load weather data from a Thermopoulos Excel file. Unchanged."""
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, index_col=0)
    except Exception as e:
        print(f"{Fore.RED}Error reading Excel file: {e}{Style.RESET_ALL}")
        return None, None

    required = ['T_air_urban', 'RH', 'pressure', 'wind_10m', 'cloud_cover']
    missing  = [col for col in required if col not in df.columns]
    if missing:
        print(f"{Fore.RED}Missing columns: {missing}{Style.RESET_ALL}")
        return None, None

    tz_str = None
    try:
        meta_df = pd.read_excel(filepath, sheet_name='Metadata', header=0)
        meta_df.columns = meta_df.columns.str.strip().str.lower()
        if 'timezone' in meta_df.columns:
            tz_str = str(meta_df.iloc[0]['timezone']).strip()
    except Exception as e:
        print(f"{Fore.YELLOW}Could not read Metadata sheet: {e}{Style.RESET_ALL}")

    if tz_str is None:
        tz_str = input("Enter timezone (e.g., Europe/Amsterdam): ").strip()
        if not tz_str:
            return None, None

    if not isinstance(df.index, pd.DatetimeIndex):
        try:
            df.index = pd.to_datetime(df.index)
        except:
            print(f"{Fore.RED}Index could not be converted to datetime.{Style.RESET_ALL}")
            return None, None

    if df.index.tz is None:
        df.index = df.index.tz_localize(tz_str)

    weather_data = []
    for idx, row in df.iterrows():
        weather_data.append({
            "dt":   int(idx.timestamp()),
            "main": {"temp": row["T_air_urban"], "humidity": row["RH"], "pressure": row["pressure"]},
            "wind": {"speed": row["wind_10m"]},
            "clouds": {"all": row["cloud_cover"]},
        })
    print(f"{Fore.GREEN}Loaded {len(weather_data)} hours from {sheet_name}.{Style.RESET_ALL}")
    return weather_data, tz_str


# =============================================================================
#  WBGT 5-DAY EXPORT (unchanged)
# =============================================================================

def calculate_and_export_wbgt(city_name, lat, lon, apply_uhi=False, population=None):
    """Calculates WBGT for the next 5 days and saves to Excel. Unchanged."""
    print(f"\n--- Calculating WBGT for {city_name} for the next 5 days ---")
    current_date   = datetime.now().strftime('%Y-%m-%d')
    sheet_name     = f"{city_name.replace(' ', '_')}_{current_date}"
    excel_filename = "HESTIA_Simulation_Results.xlsx"
    weather_data   = get_weather_forecast(lat, lon)
    if not weather_data:
        return
    start_dt    = pd.Timestamp.now(tz=UTC)
    end_dt      = start_dt + timedelta(days=5)
    interp_data = interpolate_weather(weather_data, start_dt, end_dt, interval_minutes=60)
    if apply_uhi:
        for entry in interp_data:
            entry["temp"] += calculate_uhi_effect(population, entry["wind"],
                                                   6 <= entry["time"].hour < 18)
    wbgt_results = []
    aqi = get_air_quality(lat, lon)
    for entry in interp_data:
        radiance   = calculate_solar_radiance(lat, lon, str(entry["time"]), entry["clouds"], aqi)
        wind_1_5m  = wind_speed_at_height(entry["wind"], 10, 1.5)
        globe_temp = calculate_globe_temperature(entry["temp"], radiance['ghi'], wind_1_5m,
                                                  radiance['solar_elevation'], entry["pressure"],
                                                  entry["clouds"], aqi)
        mrt        = calculate_mrt(globe_temp, entry["temp"], wind_1_5m,
                                   radiance['ghi'], radiance['solar_elevation'])
        with_solar = bool(radiance['ghi'] > 0 and radiance['solar_elevation'] > 0)
        wbgt_args  = {"twb": entry["twb"], "tg": globe_temp,
                      "with_solar_load": with_solar, "round_output": True}
        if with_solar:
            wbgt_args["tdb"] = entry["temp"]
        wbgt_val   = float(wbgt(**wbgt_args).wbgt)
        uhi_effect = calculate_uhi_effect(population, entry["wind"],
                                           6 <= entry["time"].hour < 18) if apply_uhi else 0.0
        wbgt_results.append({
            "time": entry["time"].strftime('%Y-%m-%d %H:%M'), "temperature": entry["temp"],
            "relative_humidity": entry["rh"], "wind_speed": entry["wind"],
            "cloud_cover": entry["clouds"], "pressure": entry["pressure"],
            "wet_bulb_temp": entry["twb"], "globe_temp": globe_temp, "mrt": mrt,
            "wbgt": wbgt_val, "uhi_effect_applied": uhi_effect,
        })
    try:
        if os.path.exists(excel_filename):
            with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                pd.DataFrame(wbgt_results).to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='w') as writer:
                pd.DataFrame(wbgt_results).to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"{Fore.GREEN}[SUCCESS] WBGT data exported.{Style.RESET_ALL}")
    except Exception as e:
        print(f"{Fore.RED}[ERROR] Could not export to Excel: {e}{Style.RESET_ALL}")


# =============================================================================
#  USER INPUT FUNCTIONS (unchanged)
# =============================================================================

def select_met_activity(activity_dict):
    print("\nChoose an activity:")
    for key, activity in activity_dict.items():
        print(f"{key}. {activity['description']} (MET: {activity['met']})")
    while True:
        try:
            choice = int(input("Enter the activity number: "))
            if choice in activity_dict:
                selected = activity_dict[choice]
                return choice, selected["met"], selected["speed_kmh"] * 1000 / 3600
            print("Invalid choice.")
        except ValueError:
            print("Please enter a valid number.")


def select_clo_value(clo_dict):
    print("\nChoose a clothing option:")
    for key, option in clo_dict.items():
        print(f"{key}. {option['description']} (clo: {option['clo']})")
    while True:
        try:
            choice = int(input("Enter the clothing option number: "))
            if choice in clo_dict:
                return clo_dict[choice]["clo"]
            print("Invalid choice.")
        except ValueError:
            print("Please enter a valid number.")


def select_age_configuration():
    """
    [rev10] NOTE: age_configuration is now only used as a label.
    Age is embedded in each profile from generate_base_population().
    This function is retained for backward compatibility.
    """
    print("\n[rev10] Age distribution is now set in generate_base_population().")
    print("        N(40, 10) clipped [18, 65] -- representative of NL marathon events.")
    return "standard"


def get_adaptation_profile(activity_choice):
    profiles = {
        1: {"name": "Beginner / Not acclimatized",         "training": 0.0,  "acclimatization": 0.0},
        2: {"name": "Low-average / Partially acclimatized","training": 0.25, "acclimatization": 0.3},
        3: {"name": "Average / Well-acclimatized",         "training": 0.5,  "acclimatization": 0.6},
        4: {"name": "Advanced / Fully acclimatized",       "training": 1.0,  "acclimatization": 1.0},
    }
    if activity_choice in [6, 7, 8, 9, 10]:        recommended = 4
    elif activity_choice in [1, 2, 3, 4, 5, 11]:   recommended = 3
    elif activity_choice in [15, 16, 17]:           recommended = 2
    else:                                            recommended = 1

    print(f"\n{Fore.CYAN}Recommended: {profiles[recommended]['name']}{Style.RESET_ALL}")
    while True:
        use = input("Use this profile? (y/n): ").lower().strip()
        if use in ['y', 'yes', '']:
            return profiles[recommended]
        if use in ['n', 'no']:
            for k, v in profiles.items():
                print(f"{k}. {v['name']}")
            while True:
                try:
                    c = int(input("Enter number: "))
                    if c in profiles:
                        return profiles[c]
                    print("Invalid.")
                except ValueError:
                    print("Please enter a valid number.")
        print("Please enter 'y' or 'n'.")


def clear_all_caches():
    get_lat_lon.cache_clear()
    get_timezone.cache_clear()
    get_weather_forecast.cache_clear()
    get_air_quality.cache_clear()
    calculate_solar_radiance.cache_clear()
    print("All function caches cleared.")


# =============================================================================
#  MAIN (rev10 changes: population generation, validation, and persistence)
# =============================================================================

def main():
    """
    Main simulation pipeline.

    [rev10] Population workflow:
    1. User is prompted for population parameters (N, seed).
    2. generate_base_population() creates the fixed population.
    3. validate_population() is called and results printed.
    4. User can optionally save/load the population (.pkl).
    5. run_monte_carlo_adult() uses the fixed population.

    [rev11] Population files now carry a version tag. A mismatch warning is
    printed if a file saved by a different revision is loaded.
    """
    plt.close('all')
    clear_all_caches()
    init(autoreset=True)
    print(f"{Style.BRIGHT}HESTIA: Integrated Thermal Simulation Tool (rev16){Style.RESET_ALL}")
    print(f"{Fore.CYAN}rev16 changes: T_rect P97.5/P99 plotlijnen en vaste top-2.5% "
          f"kwetsbare staartgroep op basis van piek-T_rect.{Style.RESET_ALL}")
    print(f"{Fore.CYAN}rev15 changes: calibratie-onzekerheidsband voor alle GHOR-percentages "
          f"(pct_decompensatie, pct_roberts_kritiek, pct_ehs_postfinish, collapse){Style.RESET_ALL}")
    print(f"{Fore.YELLOW}[rev14] INTERCEPT RECALIBRATION REQUIRED: K_p distributie en "
          f"post-finish module verschuiven T_rect-verdeling. Run intercept_estimation.py "
          f"met tdb 22->26 graden C referentie voor GHOR-rapportages.{Style.RESET_ALL}")

    try:
        # --- Mode Selection ---
        while True:
            print("\nSelect the simulation mode:")
            print("1. Adults (Thermal Simulation)")
            print("2. Children (Thermal Simulation)")
            print("3. WBGT Calculation (5-day forecast)")
            print("4. Use existing Thermopoulos Excel data")
            sim_mode = input("Enter your choice (1, 2, 3, or 4): ").strip()
            if sim_mode in ['1', '2', '3', '4']:
                sim_mode = ({'1': 'adult', '2': 'child', '3': 'wbgt', '4': 'excel'})[sim_mode]
                break
            print(f"{Fore.RED}Invalid choice.{Style.RESET_ALL}")

        print(f"\n--- Mode selected: {sim_mode.upper()} ---")

        # --- [rev12] Collapse endpoint selection ---
        # The active endpoint determines which intercept_kal is used for the
        # collapse risk model. Default is 'ehs' (most conservative, best documented).
        if sim_mode in ('adult', '1') or True:   # show for all adult sims
            print(f"\n[rev12] Collapse risk endpoint selection:")
            ep_keys = list(COLLAPSE_ENDPOINTS.keys())
            for i, k in enumerate(ep_keys, 1):
                ep = COLLAPSE_ENDPOINTS[k]
                status_color = Fore.GREEN if ep['status'] == 'VALIDATED' else Fore.YELLOW
                print(f"  {i}. {ep['label']}")
                print(f"     P_obs: {ep['p_obs']*10000:.1f}/10.000  |  "
                      f"intercept: {ep['intercept_kal']:.3f}  |  "
                      f"{status_color}{ep['status']}{Style.RESET_ALL}")
            print(f"  (default: 1 = EHS, most conservative)")
            ep_choice = input("Select endpoint (1/2/3, or Enter for default): ").strip()
            if ep_choice in ('2', '3'):
                import builtins
                # Override module-level ACTIVE_ENDPOINT for this session
                import sys
                _chosen_key = ep_keys[int(ep_choice) - 1]
                # Patch module globals for this run
                globals()['ACTIVE_ENDPOINT'] = _chosen_key
                globals()['P_OBS_COLLAPSE']  = COLLAPSE_ENDPOINTS[_chosen_key]['p_obs']
                print(f"{Fore.CYAN}[rev12] Active endpoint set to: "
                      f"'{COLLAPSE_ENDPOINTS[_chosen_key]['label']}'{Style.RESET_ALL}")
                if COLLAPSE_ENDPOINTS[_chosen_key]['status'] in ('PROVISIONAL', 'REQUIRES_RECALIBRATION'):
                    print(f"{Fore.YELLOW}[rev12] WARNING: This endpoint is PROVISIONAL. "
                          f"See module docstring for calibration status.{Style.RESET_ALL}")
            else:
                print(f"{Fore.CYAN}[rev12] Using default endpoint: "
                      f"'{COLLAPSE_ENDPOINTS['ehs']['label']}'{Style.RESET_ALL}")

        city_name = input("Enter the city name (e.g., Amsterdam): ")
        lat, lon  = get_lat_lon(city_name)
        if lat is None:
            print(f"{Fore.RED}Could not retrieve coordinates.{Style.RESET_ALL}")
            return

        if sim_mode == 'wbgt':
            apply_uhi  = input("Apply Urban Heat Island (UHI) effect? (y/n): ").lower() == 'y'
            population = int(input("City population: ")) if apply_uhi else None
            calculate_and_export_wbgt(city_name, lat, lon, apply_uhi, population)
            return

        sim_submode = None

        if sim_mode == 'excel':
            excel_file = input("Enter the Thermopoulos Excel filename: ").strip()
            if not os.path.isfile(excel_file):
                print(f"{Fore.RED}File not found.{Style.RESET_ALL}")
                return
            try:
                from openpyxl import load_workbook
                wb     = load_workbook(excel_file, read_only=True)
                sheets = wb.sheetnames
                print("\nAvailable sheets:")
                for i, s in enumerate(sheets, 1):
                    print(f"{i}. {s}")
                sheet_name = sheets[int(input("Select sheet number: ")) - 1]
            except Exception as e:
                print(f"{Fore.YELLOW}Could not list sheets: {e}{Style.RESET_ALL}")
                sheet_name = input("Enter sheet name: ").strip()

            weather_data, tz_str = load_weather_from_thermopoulos_excel(excel_file, sheet_name)
            if weather_data is None:
                return
            local_timezone = pytz.timezone(tz_str)

            while True:
                sub = input("\nSimulate for (1) Adults or (2) Children? ").strip()
                if sub in ['1', '2']:
                    sim_submode = 'adult' if sub == '1' else 'child'
                    break
                print(f"{Fore.RED}Invalid choice.{Style.RESET_ALL}")

            apply_uhi  = False
            population = None

        else:
            weather_data = get_weather_forecast(lat, lon)
            if not weather_data:
                return

            print("\n--- Raw Weather Forecast Summary ---")
            print("{:<20} {:<10} {:<10} {:<10}".format("Time (UTC)", "Temp (°C)", "RH (%)", "Wind (m/s)"))
            print("-" * 55)
            for entry in weather_data[:5]:
                dt_utc = datetime.fromtimestamp(entry["dt"], tz=UTC)
                print("{:<20} {:<10.1f} {:<10.0f} {:<10.1f}".format(
                    dt_utc.strftime('%Y-%m-%d %H:%M'),
                    entry["main"]["temp"], entry["main"]["humidity"], entry["wind"]["speed"]))
            print("...")

            available_dates = sorted(set([
                pd.Timestamp(datetime.fromtimestamp(e["dt"], tz=UTC)).date()
                for e in weather_data
            ]))
            print("\nAvailable forecast dates:")
            for i, d in enumerate(available_dates):
                print(f"{i+1}. {d.strftime('%Y-%m-%d')}")
            date_choice   = int(input("Choose date number: ")) - 1
            date_str      = available_dates[date_choice].strftime('%Y-%m-%d')
            local_timezone= get_timezone(lat, lon)
            apply_uhi     = input("Apply UHI effect? (y/n): ").lower() == 'y'
            population_uhi= int(input("City population: ")) if apply_uhi else None
            sim_submode   = sim_mode

        # --- Date / time selection ---
        if sim_mode == 'excel':
            available_dates = sorted(set([
                datetime.fromtimestamp(e["dt"], tz=local_timezone).date()
                for e in weather_data
            ]))
            print("\nAvailable dates:")
            for i, d in enumerate(available_dates):
                print(f"{i+1}. {d.strftime('%Y-%m-%d')}")
            date_choice = int(input("Choose date number: ")) - 1
            date_str    = available_dates[date_choice].strftime('%Y-%m-%d')

        print(f"\nDate selected: {date_str}.")
        start_time_str = input("Enter start time (HH:MM): ")
        duration       = float(input("Event duration in hours (e.g., 1.5): "))
        n_simulations  = int(input("Number of simulated participants (e.g., 5000): "))

        # --- Mode-specific inputs ---
        if sim_submode == 'adult':
            activity_choice, met_value, _ = select_met_activity(MET_ACTIVITIES_ADULT)
            adaptation_profile  = get_adaptation_profile(activity_choice)
            training_factor     = adaptation_profile['training']
            acclimatization_factor = adaptation_profile['acclimatization']
            clo_value           = select_clo_value(CLO_OPTIONS_ADULT)
            age_config          = select_age_configuration()  # [rev10] label only
        else:
            age_group_choice    = input("Select age group (1: 0-10, 2: 11-20): ")
            age_config          = "0-10" if age_group_choice == '1' else "11-20"
            _, met_value, _     = select_met_activity(MET_ACTIVITIES_CHILD)
            fitness_choice      = input("Fitness level (1: Low, 2: Medium, 3: High): ")
            fitness_factor      = {'1': 0.2, '2': 0.5, '3': 0.8}.get(fitness_choice, 0.5)
            clo_value           = select_clo_value(CLO_OPTIONS_CHILD)

        # --- [rev10] Population generation and validation ---
        base_population = None
        if sim_submode == 'adult':
            # Ask for random seed for reproducibility
            seed_input = input("\n[rev10] Enter random seed for reproducible population "
                               "(integer, or press Enter for random): ").strip()
            random_seed = int(seed_input) if seed_input else None

            # Check for existing population file
            pop_file = f"HESTIA_population_N{n_simulations}_seed{random_seed}.pkl"
            if os.path.isfile(pop_file):
                load_existing = input(f"Found existing population file '{pop_file}'. Load it? (y/n): ").lower()
                if load_existing == 'y':
                    base_population = load_population(pop_file)
                    print(f"{Fore.CYAN}Loaded {len(base_population)} profiles from disk.{Style.RESET_ALL}")

            if base_population is None:
                # Generate fresh population
                base_population = generate_base_population(
                    n_simulations         = n_simulations,
                    training_factor       = training_factor,
                    acclimatization_factor= acclimatization_factor,
                    random_seed           = random_seed,
                )
                # Validate
                print("\nRunning population validation...")
                val_result = validate_population(base_population, show_plots=True)
                if not val_result['all_pass']:
                    print(f"{Fore.YELLOW}[rev10] Population validation found issues. "
                          f"Check warnings above before proceeding.{Style.RESET_ALL}")

                # Offer to save
                save_pop = input(f"\nSave population to '{pop_file}' for future use? (y/n): ").lower()
                if save_pop == 'y':
                    save_population(base_population, pop_file)

            n_simulations = len(base_population)

        # --- Weather preparation ---
        start_dt    = pd.Timestamp(f"{date_str} {start_time_str}", tz=local_timezone.zone)
        end_dt      = start_dt + timedelta(hours=duration)
        interp_data = interpolate_weather(weather_data, start_dt, end_dt)

        if apply_uhi and sim_mode != 'excel':
            print(f"\nApplying UHI effect for {city_name}...")
            for entry in interp_data:
                is_daytime   = 6 <= entry["time"].hour < 18
                entry["temp"]+= calculate_uhi_effect(population_uhi, entry["wind"], is_daytime)

        print(f"\nStarting simulation for {n_simulations} participants in {city_name}...")

        if sim_submode == 'adult':
            all_results, stats, results_df = run_monte_carlo_adult(
                interp_data, lat, lon, met_value, clo_value, n_simulations,
                age_config, training_factor, acclimatization_factor,
                use_parallel=True,
                base_population=base_population,   # [rev10]
                random_seed=random_seed,           # [rev10]
            )
        else:
            all_results, stats = run_monte_carlo_pediatric(
                interp_data, lat, lon, met_value, clo_value, n_simulations,
                age_config, fitness_factor, use_parallel=True
            )

        if not all_results:
            print("No results were calculated.")
            return

        # --- Excel export ---
        if sim_submode == 'adult' and results_df is not None:
            excel_filename = "HESTIA_Simulation_Results.xlsx"
            sheet_name     = f"{city_name.replace(' ', '_')}_{start_dt.strftime('%Y-%m-%d')}"
            print(f"\n[INFO] Exporting to '{excel_filename}', sheet '{sheet_name}'...")
            try:
                if os.path.exists(excel_filename):
                    with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a',
                                        if_sheet_exists='replace') as writer:
                        results_df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='w') as writer:
                        results_df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"{Fore.GREEN}[SUCCESS] Exported {len(results_df)} participants.{Style.RESET_ALL}")
            except Exception as e:
                print(f"{Fore.RED}[ERROR] Excel export failed: {e}{Style.RESET_ALL}")

        # --- Environmental data summary ---
        print("\n--- Interpolated Environmental Data Used in Simulation ---")
        print("{:<20} {:<10} {:<10} {:<15} {:<10} {:<10}".format(
            "Time", "Temp (°C)", "RH (%)", "TWB (°C)", "Wind (m/s)", "MRT (°C)"))
        print("-" * 80)
        first_sim = all_results[0]
        for i, entry in enumerate(interp_data):
            print("{:<20} {:<10.1f} {:<10.0f} {:<15.1f} {:<10.1f} {:<10.1f}".format(
                entry["time"].strftime('%Y-%m-%d %H:%M'),
                entry["temp"], entry["rh"], entry["twb"],
                entry["wind"], first_sim[i]['mrt'],
            ))

        # --- Results display ---
        if sim_submode == 'adult':
            print(f"\n=== Adult Risk Classification Summary (95th Percentile + T_rect Tail) ===")
            print("{:<20} {:<10} {:<10} {:<15} {:<15} {:<10} {:<40}".format(
                "Time", "UTCI", "WBGT", "T_rect(95%)", "T_rect(97.5%)", "RPE(95%)", "Overall Risk"))
            print("-" * 122)
            for i, r_point in enumerate(all_results[0]):
                risk = calculate_adult_risk_classification(
                    r_point['utci'], r_point['wbgt'],
                    stats['upper_t_rect'][i], stats['upper_rpe'][i])
                t975 = stats.get('t_rect_p975', stats['upper_t_rect'])[i]
                print(f"{r_point['time']:<20} {r_point['utci']:<10.2f} {r_point['wbgt']:<10.2f} "
                      f"{stats['upper_t_rect'][i]:<15.2f} {t975:<15.2f} "
                      f"{stats['upper_rpe'][i]:<10.2f} "
                      f"{risk['risk_color']}{risk['detailed_risk']}")

            print(f"\n=== Core Temperature Risk Threshold Analysis (T_rect) ===")
            thresholds = sorted([float(k) for k in stats['temp_threshold_stats'].keys()])
            header = "{:<20}".format("Time") + "".join(
                [" {:<25}".format(f"% > {t}°C (95% CI)") for t in thresholds])
            print(header)
            print("-" * len(header))
            for i, r_point in enumerate(all_results[0]):
                row = "{:<20}".format(r_point['time'])
                for t in thresholds:
                    t_str = f"{t:.1f}"
                    m  = stats['temp_threshold_stats'][t_str]['mean'][i]
                    lo = stats['temp_threshold_stats'][t_str]['ci_lower'][i]
                    hi = stats['temp_threshold_stats'][t_str]['ci_upper'][i]
                    row += " {:<25}".format(f"{m:8.4f} ({lo:8.4f}-{hi:8.4f})")
                print(row)

            if 'vulnerable_tail_n' in stats:
                print(f"\n=== Vulnerable Tail Cohort (top 2.5% by peak T_rect) ===")
                print(f"Participants: {stats['vulnerable_tail_n']} "
                      f"({stats['vulnerable_tail_pct_actual']:.2f}% of simulation)")
                print(f"Peak T_rect cutoff: {stats['vulnerable_tail_peak_cutoff']:.2f} °C")
                print(f"Mean age: {stats['vulnerable_tail_age_mean']:.1f} y | "
                      f"Mean VO2max: {stats['vulnerable_tail_vo2max_mean']:.1f} mL/kg/min | "
                      f"Mean K_p: {stats['vulnerable_tail_kp_mean']:.3f} MET/°C | "
                      f"NSAID use: {stats['vulnerable_tail_nsaid_pct']:.1f}%")
                if 'vulnerable_tail_co_reserve_min_p50' in stats:
                    print(f"Tail CO reserve min P50/P05: "
                          f"{stats['vulnerable_tail_co_reserve_min_p50']:.2f} / "
                          f"{stats['vulnerable_tail_co_reserve_min_p05']:.2f} L/min | "
                          f"Tail dehydration P50/P95: "
                          f"{stats['vulnerable_tail_dehy_pct_p50']:.2f}% / "
                          f"{stats['vulnerable_tail_dehy_pct_p95']:.2f}%")

            plot_adult_results(all_results, stats, city_name, start_dt, duration)
            if results_df is not None:
                plot_adult_boxplots(results_df)

            if CVR_BESCHIKBAAR and 'hr_piek_p50' in stats:
                # [rev12] Print active endpoint before CVR summary for traceability
                if 'active_endpoint_label' in stats:
                    ep_status = stats['active_endpoint_status']
                    ep_color  = Fore.GREEN if ep_status == 'VALIDATED' else Fore.YELLOW
                    print(f"\n{'='*72}")
                    print(f"[rev12] COLLAPSE RISK MODEL -- ACTIVE ENDPOINT")
                    print(f"  Eindpunt : {stats['active_endpoint_label']}")
                    print(f"  Status   : {ep_color}{ep_status}{Style.RESET_ALL}")
                    print(f"  P_obs    : {stats['active_endpoint_p_obs']*10000:.1f} / 10.000")
                    print(f"  Intercept: {stats['collapse_intercept_kal']:.3f}")
                    print(f"  Bron     : {stats['active_endpoint_source']}")
                    if ep_status in ('PROVISIONAL', 'REQUIRES_RECALIBRATION'):
                        print(f"  {Fore.YELLOW}NOTE: Herrekalibratie vereist na rev14 "
                              f"K_p-distributie (A) en post-finish module (C). "
                              f"Run intercept_estimation.py met tdb 22→26°C.{Style.RESET_ALL}")
                    print(f"{'='*72}")

                # [rev14] GHOR OPERATIONELE SAMENVATTING
                print(f"\n{'='*72}")
                print(f"[rev14] GHOR OPERATIONELE SAMENVATTING")
                print(f"{'='*72}")

                # A — K_p populatie
                if 'kp_mean' in stats:
                    print(f"\n  [A] PACING-RESPONSIVITEIT POPULATIE")
                    print(f"      K_p gemiddeld        : {stats['kp_mean']:.3f} MET/°C")
                    print(f"      K_p P5               : {stats['kp_p05']:.3f} MET/°C")
                    pct_km = stats['pct_kp_onder_002']
                    km_col = Fore.RED if pct_km > 5 else Fore.YELLOW if pct_km > 2 else Fore.GREEN
                    # [rev15] No intercept sensitivity for K_p (it is sampled independently);
                    # report as point estimate with MC note.
                    print(f"      % K_p < 0.02 (overgemotiveerd): "
                          f"{km_col}{pct_km:.1f}%{Style.RESET_ALL}  "
                          f"[MC-steekproef; N={n_simulations}]")
                    print(f"      % NSAID-gebruik      : {stats.get('pct_nsaid', 0):.1f}%")
                    if pct_km > 3:
                        print(f"      {Fore.YELLOW}AANDACHT: {pct_km:.1f}% van de populatie pacat "
                              f"niet bij thermische overbelasting. Verhoogde "
                              f"finish-zone surveillance aanbevolen.{Style.RESET_ALL}")

                # B — AUC klinisch (Roberts)
                if 'auc_klinisch_p95' in stats:
                    print(f"\n  [B] THERMISCHE DOSIS (Roberts 2007, grens 60 graad-min)")
                    print(f"      AUC_klinisch P50     : {stats['auc_klinisch_p50']:.2f} °C·min")
                    print(f"      AUC_klinisch P95     : {stats['auc_klinisch_p95']:.2f} °C·min")
                    print(f"      AUC totaal P95       : {stats['auc_klinisch_totaal_p95']:.2f} °C·min")
                    pct_rob = stats['pct_roberts_kritiek']
                    rob_col = Fore.RED if pct_rob > 1 else Fore.YELLOW if pct_rob > 0.1 else Fore.GREEN
                    # [rev15] Calibration uncertainty band
                    rob_lo  = stats.get('pct_roberts_kritiek_lo', pct_rob)
                    rob_hi  = stats.get('pct_roberts_kritiek_hi', pct_rob)
                    rob_str = _format_pct_with_band(pct_rob, rob_lo, rob_hi)
                    print(f"      % boven Roberts-grens: "
                          f"{rob_col}{rob_str}{Style.RESET_ALL}")
                    print(f"      [rev15] Band = kalibratie-onzekerheid intercept ±0.5; "
                          f"geen klassieke 95% BI")
                    if pct_rob > 0.5:
                        print(f"      {Fore.RED}KRITIEK: >0.5% populatie overschrijdt 60 graad-min. "
                              f"Koeling binnen 5-10 min na collaps essentieel.{Style.RESET_ALL}")

                # C — Post-finish
                if 'pct_ehs_postfinish' in stats:
                    print(f"\n  [C] POST-FINISH RISICOZONE (10 min na finish)")
                    pct_pf = stats['pct_ehs_postfinish']
                    pf_col = Fore.RED if pct_pf > 0.5 else Fore.YELLOW if pct_pf > 0.1 else Fore.GREEN
                    # [rev15] Calibration uncertainty band
                    pf_lo  = stats.get('pct_ehs_postfinish_lo', pct_pf)
                    pf_hi  = stats.get('pct_ehs_postfinish_hi', pct_pf)
                    pf_str = _format_pct_with_band(pct_pf, pf_lo, pf_hi)
                    print(f"      % EHS post-finish    : "
                          f"{pf_col}{pf_str}{Style.RESET_ALL}")
                    print(f"      T_rect piek PF P95   : {stats['t_rect_pf_p95']:.2f}°C")
                    print(f"      [rev15] Band = kalibratie-onzekerheid intercept ±0.5")
                    if pct_pf > 0.1:
                        print(f"      {Fore.YELLOW}AANDACHT: Verhoogde medische bezetting "
                              f"finish-zone aanbevolen (eerste 10 min na finish).{Style.RESET_ALL}")
                print(f"{'='*72}")
                print_cvr_population_summary(
                    stats,
                    start_group_label=city_name,
                    n_participants=n_simulations,
                    edition=str(start_dt.date()),
                )

                # [rev15] CALIBRATION UNCERTAINTY SUMMARY
                # Printed after CVR console output so GHOR readers see the
                # uncertainty context for every percentage above.
                if 'pct_decompensatie_lo' in stats:
                    print(f"\n{'='*72}")
                    print(f"[rev15] KALIBRATIE-ONZEKERHEIDSBAND (intercept ±0.5 log-odds)")
                    print(f"{'='*72}")
                    print(f"  Toelichting: de '~lo–hi%' band weerspiegelt de onzekerheid")
                    print(f"  in de calibratiebron (Boston→DtD transfer), NIET een")
                    print(f"  klassieke 95% BI.  De Monte Carlo-steekproefonzekerheid")
                    print(f"  (Wald, N={n_simulations}) is hierin NIET de dominante factor.")
                    print(f"")
                    dec_str = _format_pct_with_band(
                        stats['pct_decompensatie'],
                        stats['pct_decompensatie_lo'],
                        stats['pct_decompensatie_hi'])
                    col_str = _format_pct_with_band(
                        stats['p_collapse_gemiddeld'],
                        stats['p_collapse_gemiddeld_lo'],
                        stats['p_collapse_gemiddeld_hi'])
                    hoog_str = _format_pct_with_band(
                        stats['pct_hoog_collapsrisico'],
                        stats['pct_hoog_collapsrisico_lo'],
                        stats['pct_hoog_collapsrisico_hi'])
                    print(f"  CVR decompensatie        : {dec_str}")
                    print(f"  Gem. collapsekans        : {col_str}")
                    print(f"  % hoog collapsrisico     : {hoog_str}")
                    if 'pct_roberts_kritiek_lo' in stats:
                        rob_str = _format_pct_with_band(
                            stats['pct_roberts_kritiek'],
                            stats['pct_roberts_kritiek_lo'],
                            stats['pct_roberts_kritiek_hi'])
                        print(f"  % boven Roberts-grens    : {rob_str}")
                    if 'pct_ehs_postfinish_lo' in stats:
                        pf_str = _format_pct_with_band(
                            stats['pct_ehs_postfinish'],
                            stats['pct_ehs_postfinish_lo'],
                            stats['pct_ehs_postfinish_hi'])
                        print(f"  EHS post-finish          : {pf_str}")
                    print(f"")
                    print(f"  Band wordt smaller na klinische kalibratie (Bongers,")
                    print(f"  Radboudumc) van Z-functie coëfficiënten W_T1/W_T2/W_C/W_D.")
                    print(f"{'='*72}")

            df_met_analysis = analyze_met_thresholds(interp_data, lat, lon, clo_value)
            plot_met_thresholds(df_met_analysis, met_value)

        else:
            print(f"\n=== Pediatric Risk Classification Summary (95th Percentile) ===")
            print("{:<20} {:<10} {:<10} {:<15} {:<10} {:<15} {:<40}".format(
                "Time", "UTCI", "WBGT", "T_core(95%)", "RPE(95%)", "% First Aid", "Risk"))
            print("-" * 130)
            for i, r_point in enumerate(all_results[0]):
                risk = calculate_pediatric_risk_classification(
                    r_point['utci'], r_point['wbgt'],
                    stats['upper_t_core'][i], stats['upper_rpe'][i])
                print(f"{r_point['time']:<20} {r_point['utci']:<10.2f} {r_point['wbgt']:<10.2f} "
                      f"{stats['upper_t_core'][i]:<15.2f} {stats['upper_rpe'][i]:<10.2f} "
                      f"{stats['percent_first_aid'][i]:<15.1f} "
                      f"{risk['risk_color']}{risk['detailed_risk']}")

            plot_pediatric_results(all_results, stats, city_name, start_dt, duration)

        input("\nPress Enter to close the plot windows and exit.")

    except ValueError as e:
        import traceback
        traceback.print_exc()
        print(f"{Fore.RED}Invalid input error: {e}{Style.RESET_ALL}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"{Fore.RED}An unexpected error occurred: {e}{Style.RESET_ALL}")


if __name__ == "__main__":
    main()
